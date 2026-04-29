from __future__ import annotations

import os
import io
import json
import re
import shutil
import sqlite3
import uuid
import difflib
import traceback
import ast
import operator
from pathlib import Path
from contextlib import closing
from datetime import datetime
from functools import wraps
from typing import Any, Iterable

from flask import (
    Flask,
    Response,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def get_db_path() -> str:
    """Resolve o caminho do banco SQLite.

    No Railway, quando existir um Volume montado, o banco fica dentro dele
    para não perder os dados em redeploy. Se o Volume estiver vazio, copiamos
    o dados.db que vai junto no projeto como banco inicial.
    """
    custom_db_path = os.environ.get("DB_PATH", "").strip()
    if custom_db_path:
        custom_dir = os.path.dirname(custom_db_path)
        if custom_dir:
            os.makedirs(custom_dir, exist_ok=True)
        return custom_db_path

    bundled_db_path = os.path.join(BASE_DIR, "dados.db")
    volume_mount = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "").strip()
    if volume_mount:
        os.makedirs(volume_mount, exist_ok=True)
        volume_db_path = os.path.join(volume_mount, "dados.db")
        if not os.path.exists(volume_db_path) and os.path.exists(bundled_db_path):
            shutil.copy2(bundled_db_path, volume_db_path)
        return volume_db_path

    return bundled_db_path


DB_PATH = get_db_path()
STOCK_SEED_PATH = os.path.join(BASE_DIR, "stock_seed.json")
app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"), static_folder=os.path.join(BASE_DIR, "static"))
app.secret_key = os.environ.get("SECRET_KEY", "alvorada-separacao-lojas")


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def agora_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def agora_br() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def fmt_num(value: Any) -> str:
    try:
        f = float(value or 0)
    except (TypeError, ValueError):
        return "0"
    if f.is_integer():
        return str(int(f))
    return f"{f:.3f}".rstrip("0").rstrip(".")


def fmt_money(value: Any) -> str:
    try:
        f = float(value or 0)
    except (TypeError, ValueError):
        f = 0.0
    text = f"{f:,.2f}"
    return "R$ " + text.replace(",", "X").replace(".", ",").replace("X", ".")


def parse_fator_embalagem(value: Any, field_name: str = "Fator da embalagem", default: float = 1.0) -> float:
    raw = str(value or "").strip().casefold()
    if not raw:
        return float(default)
    match = re.search(r"[-+]?\d+(?:[\.,]\d+)?", raw.replace("emb", " "))
    if not match:
        raise ValueError(f"{field_name} inválido.")
    fator = parse_float(match.group(0), field_name)
    if fator <= 0:
        raise ValueError(f"{field_name} deve ser maior que zero.")
    return fator


def quantidade_em_embalagens(quantidade_total: Any, fator_embalagem: Any) -> float:
    try:
        quantidade = float(quantidade_total or 0)
    except (TypeError, ValueError):
        quantidade = 0.0
    try:
        fator = float(fator_embalagem or 1)
    except (TypeError, ValueError):
        fator = 1.0
    if fator <= 0:
        fator = 1.0
    return quantidade / fator


def fmt_fator_embalagem(value: Any) -> str:
    return f"Emb{fmt_num(value or 1)}"


ACCESS_OPTIONS: list[tuple[str, str]] = [
    ("painel", "Painel"),
    ("separacoes", "Separações"),
    ("estoque", "Estoque"),
    ("recebimentos", "Recebimentos"),
    ("balanco", "Balanço"),
    ("relatorios", "Relatórios"),
    ("usuarios", "Usuários"),
    ("lojas", "Lojas"),
    ("pedidos", "Criar pedidos"),
    ("lotes", "Lotes"),
    ("configuracoes", "Configurações"),
    ("mcp_teste", "MCP/IA"),
    ("comunicacao", "Chat/Tarefas"),
    ("codigo_fonte", "Código fonte"),
    ("auditoria", "Auditoria"),
]
ACCESS_KEYS = {key for key, _ in ACCESS_OPTIONS}
ACCESS_LABELS = dict(ACCESS_OPTIONS)
PERMISSION_LEVEL_LABELS = {"admin": "Admin", "comum": "Comum"}
ROLE_LABELS = {
    "admin": "Admin",
    "gerente": "Gerente",
    "estoque": "Estoque",
    "separador": "Separador",
    "conferente": "Conferente",
    "balanco": "Balanço",
    "desenvolvedor": "Desenvolvedor",
    "visualizador": "Visualizador",
}
DEFAULT_ACCESS_BY_ROLE: dict[str, set[str]] = {
    "admin": set(ACCESS_KEYS),
    "gerente": {"painel", "separacoes", "estoque", "balanco", "relatorios", "lojas", "lotes", "mcp_teste", "comunicacao"},
    "estoque": {"painel", "estoque", "recebimentos", "balanco", "relatorios", "lotes", "mcp_teste", "comunicacao"},
    "separador": {"painel", "separacoes", "pedidos", "lotes", "comunicacao"},
    "conferente": {"painel", "recebimentos", "separacoes", "pedidos", "lotes", "comunicacao"},
    "balanco": {"painel", "estoque", "balanco", "relatorios", "mcp_teste", "comunicacao"},
    "desenvolvedor": {"painel", "mcp_teste", "codigo_fonte"},
    "visualizador": {"painel", "relatorios", "mcp_teste", "comunicacao"},
}
MODULE_ENDPOINTS = {
    "painel": "dashboard",
    "separacoes": "listar_separacoes",
    "estoque": "estoque",
    "recebimentos": "recebimentos",
    "balanco": "balancos",
    "relatorios": "relatorios",
    "usuarios": "usuarios",
    "lojas": "lojas",
    "pedidos": "nova_separacao",
    "lotes": "listar_lotes",
    "configuracoes": "configuracoes",
    "mcp_teste": "mcp_teste",
    "comunicacao": "comunicacao",
    "codigo_fonte": "admin_codigo_fonte",
    "auditoria": "auditoria",
}
STOCK_MOVEMENT_TYPE_OPTIONS: list[tuple[str, str]] = [
    ("ENTRADA_INICIAL", "Cadastro inicial"),
    ("IMPORTACAO_ERP", "Importação ERP"),
    ("IMPORTACAO_ERP_NOVO", "Novo via ERP"),
    ("AJUSTE_MANUAL", "Ajuste manual"),
    ("RECONTAGEM", "Ajuste de quantidade"),
    ("BALANCO_ESTOQUE", "Balanço de estoque"),
    ("REMOVIDO_ESTOQUE", "Remoção do estoque"),
    ("SAIDA_SEPARACAO", "Saída por separação"),
    ("ESTORNO_HISTORICO", "Estorno do histórico"),
    ("RECEBIMENTO_MERCADORIA", "Recebimento de mercadoria"),
]


def normalize_role(value: Any) -> str:
    role = str(value or "").strip().lower()
    allowed = {"admin", "gerente", "estoque", "separador", "conferente", "balanco", "desenvolvedor", "visualizador"}
    return role if role in allowed else "separador"



def normalize_permission_level(value: Any, role: Any = None) -> str:
    if normalize_role(role) == "admin":
        return "admin"
    level = str(value or "").strip().lower()
    if level in {"admin", "comum"}:
        return level
    return "comum"



def default_access_rules(role: Any, permission_level: Any = "comum") -> set[str]:
    role_norm = normalize_role(role)
    level_norm = normalize_permission_level(permission_level, role_norm)
    if level_norm == "admin":
        return set(ACCESS_KEYS)
    return set(DEFAULT_ACCESS_BY_ROLE.get(role_norm, {"painel"}))



def serialize_access_rules(accesses: Iterable[str]) -> str:
    normalized = sorted({str(item).strip().lower() for item in accesses if str(item).strip().lower() in ACCESS_KEYS})
    return json.dumps(normalized, ensure_ascii=False)



def parse_access_rules(raw: Any, role: Any = None, permission_level: Any = "comum") -> set[str]:
    if isinstance(raw, (list, tuple, set)):
        valores = raw
    else:
        texto = str(raw or "").strip()
        if not texto:
            return default_access_rules(role, permission_level)
        try:
            valores = json.loads(texto)
        except json.JSONDecodeError:
            return default_access_rules(role, permission_level)
    acessos = {str(item).strip().lower() for item in valores if str(item).strip().lower() in ACCESS_KEYS}
    if not acessos:
        return default_access_rules(role, permission_level)
    return acessos



def user_permission_level(user: sqlite3.Row | dict[str, Any] | None) -> str:
    if user is None:
        return "comum"
    role = user["role"] if "role" in user.keys() else None
    raw = user["permission_level"] if "permission_level" in user.keys() else None
    return normalize_permission_level(raw, role)



def user_is_admin(user: sqlite3.Row | dict[str, Any] | None) -> bool:
    if user is None:
        return False
    role = normalize_role(user["role"] if "role" in user.keys() else None)
    return role == "admin" or user_permission_level(user) == "admin"



def user_access_set(user: sqlite3.Row | dict[str, Any] | None) -> set[str]:
    if user is None:
        return set()
    role = user["role"] if "role" in user.keys() else None
    if user_is_admin(user):
        return set(ACCESS_KEYS)
    raw = user["access_rules"] if "access_rules" in user.keys() else None
    return parse_access_rules(raw, role, user_permission_level(user))



def user_has_access(user: sqlite3.Row | dict[str, Any] | None, module: str) -> bool:
    module_key = str(module or "").strip().lower()
    if module_key not in ACCESS_KEYS:
        return False
    return module_key in user_access_set(user)



def access_labels_for_user(user: sqlite3.Row | dict[str, Any] | None) -> list[str]:
    return [ACCESS_LABELS[key] for key in ACCESS_OPTIONS_KEYS_IN_ORDER if key in user_access_set(user)]


ACCESS_OPTIONS_KEYS_IN_ORDER = [key for key, _ in ACCESS_OPTIONS]



def role_label(role: Any) -> str:
    return ROLE_LABELS.get(normalize_role(role), str(role or "-").strip() or "-")



def permission_level_label(value: Any, role: Any = None) -> str:
    level = normalize_permission_level(value, role)
    return PERMISSION_LEVEL_LABELS.get(level, level.title())



def first_allowed_endpoint(user: sqlite3.Row | dict[str, Any] | None) -> str:
    for module in ACCESS_OPTIONS_KEYS_IN_ORDER:
        if user_has_access(user, module):
            return MODULE_ENDPOINTS[module]
    return "minha_conta"



def forbidden_redirect(message: str) -> Response:
    flash(message, "error")
    destino = first_allowed_endpoint(g.user)
    if request.endpoint == destino:
        destino = "minha_conta"
    return redirect(url_for(destino))



def module_required(module: str):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if g.user is None:
                return redirect(url_for("login"))
            if not user_has_access(g.user, module):
                return forbidden_redirect("Você não tem permissão para acessar essa área.")
            return view(*args, **kwargs)

        return wrapped

    return decorator



def can_adjust_stock(user: sqlite3.Row | dict[str, Any] | None) -> bool:
    if user is None:
        return False
    return user_has_access(user, "estoque")



def can_edit_stock_registration(user: sqlite3.Row | dict[str, Any] | None) -> bool:
    if user is None:
        return False
    return user_has_access(user, "estoque") and user_is_admin(user)



def count_admin_users(conn: sqlite3.Connection, exclude_user_id: int | None = None) -> int:
    rows = conn.execute("SELECT id, role, permission_level FROM users WHERE ativo = 1").fetchall()
    total = 0
    for row in rows:
        if exclude_user_id is not None and int(row["id"]) == int(exclude_user_id):
            continue
        if user_is_admin(row):
            total += 1
    return total


def natural_store_sort_key(value: Any) -> tuple[Any, ...]:
    text = str(value or '').strip()
    parts = re.split(r'(\d+)', text.casefold())
    key: list[Any] = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return tuple(key)


def sort_store_rows(rows: Iterable[sqlite3.Row]) -> list[sqlite3.Row]:
    return sorted(rows, key=lambda row: (natural_store_sort_key(row['store_nome']), row['id']))


def ensure_column(conn: sqlite3.Connection, table: str, column: str, ddl: str) -> None:
    columns = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")


def ensure_schema_updates(conn: sqlite3.Connection) -> None:
    ensure_column(conn, "separations", "lote_codigo", "lote_codigo TEXT")
    ensure_column(conn, "stock_items", "ativo", "ativo INTEGER NOT NULL DEFAULT 1")
    ensure_column(conn, "stock_items", "codigo_barras", "codigo_barras TEXT")
    ensure_column(conn, "stock_items", "fator_embalagem", "fator_embalagem REAL NOT NULL DEFAULT 1")
    ensure_column(conn, "stock_items", "linha_erp", "linha_erp TEXT")
    ensure_column(conn, "stock_items", "erp_loja", "erp_loja TEXT")
    ensure_column(conn, "stock_items", "erp_nivel", "erp_nivel TEXT")
    ensure_column(conn, "stock_items", "linha_caminho_erp", "linha_caminho_erp TEXT")
    ensure_column(conn, "stock_items", "erp_data_base", "erp_data_base TEXT")
    ensure_column(conn, "stock_items", "erp_atualizado_em", "erp_atualizado_em TEXT")
    ensure_column(conn, "separation_items", "fator_embalagem", "fator_embalagem REAL NOT NULL DEFAULT 1")
    ensure_column(conn, "separation_items", "carryover_source_item_id", "carryover_source_item_id INTEGER")
    ensure_column(conn, "separation_items", "carryover_copied", "carryover_copied INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "separation_items", "quantidade_conferida", "quantidade_conferida REAL")
    ensure_column(conn, "separation_items", "conferido_em", "conferido_em TEXT")
    ensure_column(conn, "users", "permission_level", "permission_level TEXT NOT NULL DEFAULT 'comum'")
    ensure_column(conn, "users", "access_rules", "access_rules TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "users", "store_id", "store_id INTEGER")
    ensure_column(conn, "users", "foto_perfil", "foto_perfil TEXT")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_users_store_id ON users(store_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_separations_lote_codigo ON separations(lote_codigo)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_stock_items_codigo_barras ON stock_items(codigo_barras)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_stock_items_linha_erp ON stock_items(linha_erp)")
    conn.executescript("""
    
CREATE TABLE IF NOT EXISTS chat_groups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    ativo INTEGER NOT NULL DEFAULT 1,
    FOREIGN KEY (criado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS chat_group_members (
    group_id INTEGER NOT NULL,
    user_id INTEGER NOT NULL,
    criado_em TEXT NOT NULL,
    PRIMARY KEY (group_id, user_id),
    FOREIGN KEY (group_id) REFERENCES chat_groups (id) ON DELETE CASCADE,
    FOREIGN KEY (user_id) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS chat_messages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    group_id INTEGER NOT NULL,
    user_id INTEGER,
    mensagem TEXT,
    arquivo_nome TEXT,
    arquivo_path TEXT,
    arquivo_status TEXT NOT NULL DEFAULT 'sem_arquivo',
    recebido_por INTEGER,
    recebido_em TEXT,
    expira_em TEXT,
    criado_em TEXT NOT NULL,
    FOREIGN KEY (group_id) REFERENCES chat_groups (id) ON DELETE CASCADE,
    FOREIGN KEY (user_id) REFERENCES users (id),
    FOREIGN KEY (recebido_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS team_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    titulo TEXT NOT NULL,
    descricao TEXT,
    responsavel_id INTEGER,
    status TEXT NOT NULL DEFAULT 'ABERTA',
    prazo TEXT,
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    finalizado_em TEXT,
    FOREIGN KEY (responsavel_id) REFERENCES users (id),
    FOREIGN KEY (criado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS scheduled_orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    loja_id INTEGER,
    titulo TEXT NOT NULL,
    itens_texto TEXT NOT NULL,
    agendado_para TEXT,
    status TEXT NOT NULL DEFAULT 'AGENDADO',
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    enviado_por INTEGER,
    enviado_em TEXT,
    FOREIGN KEY (loja_id) REFERENCES stores (id),
    FOREIGN KEY (criado_por) REFERENCES users (id),
    FOREIGN KEY (enviado_por) REFERENCES users (id)
);
CREATE TABLE IF NOT EXISTS store_issue_reports (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tipo TEXT NOT NULL,
    destino_user_id INTEGER,
    loja_id INTEGER,
    nota_numero TEXT,
    itens_json TEXT NOT NULL DEFAULT '[]',
    relato TEXT,
    status TEXT NOT NULL DEFAULT 'ENVIADO',
    chat_group_id INTEGER,
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    FOREIGN KEY (destino_user_id) REFERENCES users (id),
    FOREIGN KEY (loja_id) REFERENCES stores (id),
    FOREIGN KEY (chat_group_id) REFERENCES chat_groups (id),
    FOREIGN KEY (criado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS erp_stock_imports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT NOT NULL,
        loja TEXT,
        data_base TEXT,
        total_linhas INTEGER NOT NULL DEFAULT 0,
        total_produtos INTEGER NOT NULL DEFAULT 0,
        total_grupos INTEGER NOT NULL DEFAULT 0,
        duplicados INTEGER NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'PREVIEW',
        modo TEXT,
        resumo_json TEXT NOT NULL DEFAULT '{}',
        criado_por INTEGER,
        criado_em TEXT NOT NULL,
        aplicado_em TEXT,
        FOREIGN KEY (criado_por) REFERENCES users (id)
    );
    CREATE TABLE IF NOT EXISTS erp_stock_import_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        import_id INTEGER NOT NULL,
        row_number INTEGER,
        codigo TEXT NOT NULL,
        codigo_barras TEXT,
        nivel TEXT,
        descricao TEXT NOT NULL,
        linha TEXT,
        caminho_linha TEXT,
        preco_custo REAL NOT NULL DEFAULT 0,
        preco_venda REAL NOT NULL DEFAULT 0,
        saldo_qtd REAL NOT NULL DEFAULT 0,
        saldo_custo REAL NOT NULL DEFAULT 0,
        saldo_venda REAL NOT NULL DEFAULT 0,
        dias REAL NOT NULL DEFAULT 0,
        sugestao REAL NOT NULL DEFAULT 0,
        estoque_ideal REAL NOT NULL DEFAULT 0,
        stock_item_id INTEGER,
        saldo_anterior REAL,
        delta REAL,
        status TEXT NOT NULL DEFAULT 'PENDENTE',
        motivo TEXT,
        FOREIGN KEY (import_id) REFERENCES erp_stock_imports (id) ON DELETE CASCADE,
        FOREIGN KEY (stock_item_id) REFERENCES stock_items (id)
    );
    """)
    ensure_column(conn, "chat_groups", "tipo_chat", "tipo_chat TEXT NOT NULL DEFAULT 'grupo'")
    ensure_column(conn, "chat_groups", "direto_key", "direto_key TEXT")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_chat_groups_direto_key ON chat_groups(direto_key) WHERE direto_key IS NOT NULL")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_import_items_import_id ON erp_stock_import_items(import_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_import_items_codigo ON erp_stock_import_items(codigo)")
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS receipts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nota_numero TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'FINALIZADO',
        observacao TEXT,
        conferente_id INTEGER,
        criado_em TEXT NOT NULL,
        finalizado_em TEXT NOT NULL,
        FOREIGN KEY (conferente_id) REFERENCES users (id)
    );
    CREATE TABLE IF NOT EXISTS receipt_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        receipt_id INTEGER NOT NULL,
        stock_item_id INTEGER,
        codigo TEXT NOT NULL,
        codigo_barras TEXT,
        descricao TEXT NOT NULL,
        validade TEXT,
        quantidade REAL NOT NULL DEFAULT 0,
        criado_em TEXT NOT NULL,
        FOREIGN KEY (receipt_id) REFERENCES receipts (id) ON DELETE CASCADE,
        FOREIGN KEY (stock_item_id) REFERENCES stock_items (id)
    );
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_receipts_nota ON receipts(nota_numero)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_receipt_items_receipt ON receipt_items(receipt_id)")
    conn.execute("UPDATE separations SET lote_codigo = 'SEP-' || id WHERE lote_codigo IS NULL OR TRIM(lote_codigo) = ''")
    conn.execute("UPDATE stock_items SET fator_embalagem = 1 WHERE fator_embalagem IS NULL OR fator_embalagem <= 0")
    conn.execute("UPDATE separation_items SET fator_embalagem = 1 WHERE fator_embalagem IS NULL OR fator_embalagem <= 0")
    conn.execute("UPDATE separation_items SET carryover_copied = 0 WHERE carryover_copied IS NULL")
    users = conn.execute("SELECT id, role, permission_level, access_rules FROM users").fetchall()
    for user in users:
        permission_level = normalize_permission_level(user["permission_level"], user["role"])
        access_rules = parse_access_rules(user["access_rules"], user["role"], permission_level)
        role_norm = normalize_role(user["role"])
        if permission_level != "admin" and role_norm in {"separador", "conferente", "estoque", "gerente"}:
            
            if role_norm in {"conferente", "estoque"}:
                access_rules.add("recebimentos")
            access_rules.update({"pedidos", "lotes"})
        conn.execute(
            "UPDATE users SET permission_level = ?, access_rules = ? WHERE id = ?",
            (permission_level, serialize_access_rules(access_rules), user["id"]),
        )


def novo_lote_codigo() -> str:
    return "LT-" + uuid.uuid4().hex[:10].upper()


def lote_operacao_chave_expr(alias: str = "s") -> str:
    return f"""
    CASE
        WHEN {alias}.lote_codigo LIKE 'SEP-%' THEN
            'LEGADO~' || {alias}.lote_nome || '~' || {alias}.data_referencia || '~' ||
            COALESCE(CAST({alias}.responsavel_id AS TEXT), '0') || '~' ||
            COALESCE(CAST({alias}.conferente_id AS TEXT), '0') || '~' ||
            COALESCE(CAST({alias}.criado_por AS TEXT), '0') || '~' ||
            COALESCE(CAST({alias}.usar_estoque AS TEXT), '0') || '~' ||
            COALESCE(substr({alias}.criado_em, 1, 16), '')
        ELSE COALESCE(NULLIF(TRIM({alias}.lote_codigo), ''),
            'LEGADO~' || {alias}.lote_nome || '~' || {alias}.data_referencia || '~' ||
            COALESCE(CAST({alias}.responsavel_id AS TEXT), '0') || '~' ||
            COALESCE(CAST({alias}.conferente_id AS TEXT), '0') || '~' ||
            COALESCE(CAST({alias}.criado_por AS TEXT), '0') || '~' ||
            COALESCE(CAST({alias}.usar_estoque AS TEXT), '0') || '~' ||
            COALESCE(substr({alias}.criado_em, 1, 16), '')
        )
    END
    """


def lote_operacao_chave_row(row: sqlite3.Row | dict[str, Any]) -> str:
    lote_codigo = (row["lote_codigo"] or "").strip() if "lote_codigo" in row.keys() else ""
    if lote_codigo and not lote_codigo.startswith("SEP-"):
        return lote_codigo
    return "LEGADO~{lote_nome}~{data_referencia}~{responsavel_id}~{conferente_id}~{criado_por}~{usar_estoque}~{criado_minuto}".format(
        lote_nome=row["lote_nome"],
        data_referencia=row["data_referencia"],
        responsavel_id=row["responsavel_id"] or 0,
        conferente_id=row["conferente_id"] or 0,
        criado_por=row["criado_por"] or 0,
        usar_estoque=row["usar_estoque"] or 0,
        criado_minuto=(row["criado_em"] or "")[:16],
    )


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    username TEXT NOT NULL UNIQUE,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL,
    permission_level TEXT NOT NULL DEFAULT 'comum',
    access_rules TEXT NOT NULL DEFAULT '',
    store_id INTEGER,
    ativo INTEGER NOT NULL DEFAULT 1,
    criado_em TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS stores (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL UNIQUE,
    ativo INTEGER NOT NULL DEFAULT 1,
    criado_em TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS stock_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT NOT NULL UNIQUE,
    codigo_barras TEXT,
    descricao TEXT NOT NULL,
    fator_embalagem REAL NOT NULL DEFAULT 1,
    quantidade_atual REAL NOT NULL DEFAULT 0,
    custo_unitario REAL NOT NULL DEFAULT 0,
    linha_erp TEXT,
    erp_loja TEXT,
    erp_nivel TEXT,
    linha_caminho_erp TEXT,
    erp_data_base TEXT,
    erp_atualizado_em TEXT,
    ativo INTEGER NOT NULL DEFAULT 1,
    atualizado_em TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS stock_movements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    stock_item_id INTEGER NOT NULL,
    tipo TEXT NOT NULL,
    quantidade REAL NOT NULL,
    observacao TEXT,
    referencia_tipo TEXT,
    referencia_id INTEGER,
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    FOREIGN KEY (stock_item_id) REFERENCES stock_items (id),
    FOREIGN KEY (criado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS separations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lote_nome TEXT NOT NULL,
    data_referencia TEXT NOT NULL,
    store_id INTEGER NOT NULL,
    responsavel_id INTEGER,
    conferente_id INTEGER,
    status TEXT NOT NULL DEFAULT 'ABERTA',
    usar_estoque INTEGER NOT NULL DEFAULT 1,
    observacao TEXT,
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    enviado_conferencia_em TEXT,
    finalizado_em TEXT,
    FOREIGN KEY (store_id) REFERENCES stores (id),
    FOREIGN KEY (responsavel_id) REFERENCES users (id),
    FOREIGN KEY (conferente_id) REFERENCES users (id),
    FOREIGN KEY (criado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS separation_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    separation_id INTEGER NOT NULL,
    codigo TEXT NOT NULL,
    descricao TEXT NOT NULL,
    fator_embalagem REAL NOT NULL DEFAULT 1,
    quantidade_pedida REAL NOT NULL,
    quantidade_separada REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'PENDENTE',
    custo_unitario_ref REAL NOT NULL DEFAULT 0,
    carryover_source_item_id INTEGER,
    carryover_copied INTEGER NOT NULL DEFAULT 0,
    criado_em TEXT NOT NULL,
    atualizado_em TEXT NOT NULL,
    FOREIGN KEY (separation_id) REFERENCES separations (id) ON DELETE CASCADE
);


CREATE TABLE IF NOT EXISTS balance_counts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    titulo TEXT NOT NULL,
    observacao TEXT,
    status TEXT NOT NULL DEFAULT 'ABERTO',
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    confirmado_por INTEGER,
    confirmado_em TEXT,
    FOREIGN KEY (criado_por) REFERENCES users (id),
    FOREIGN KEY (confirmado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS balance_count_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    balance_count_id INTEGER NOT NULL,
    stock_item_id INTEGER NOT NULL,
    codigo TEXT NOT NULL,
    descricao TEXT NOT NULL,
    linha_erp TEXT,
    quantidade_sistema REAL NOT NULL DEFAULT 0,
    quantidade_contada REAL NOT NULL DEFAULT 0,
    delta REAL NOT NULL DEFAULT 0,
    custo_unitario REAL NOT NULL DEFAULT 0,
    criado_em TEXT NOT NULL,
    atualizado_em TEXT NOT NULL,
    FOREIGN KEY (balance_count_id) REFERENCES balance_counts (id) ON DELETE CASCADE,
    FOREIGN KEY (stock_item_id) REFERENCES stock_items (id),
    UNIQUE(balance_count_id, stock_item_id)
);

CREATE TABLE IF NOT EXISTS mcp_saved_queries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    titulo TEXT NOT NULL,
    pergunta TEXT NOT NULL,
    contexto TEXT NOT NULL DEFAULT 'mcp_teste',
    linha TEXT,
    filtros TEXT NOT NULL DEFAULT '{}',
    publico INTEGER NOT NULL DEFAULT 0,
    criado_em TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS mcp_query_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    contexto TEXT NOT NULL DEFAULT 'mcp_teste',
    pergunta TEXT NOT NULL,
    linha TEXT,
    filtros TEXT NOT NULL DEFAULT '{}',
    tool TEXT,
    total_registros INTEGER NOT NULL DEFAULT 0,
    exportado_formato TEXT,
    criado_em TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id)
);


CREATE TABLE IF NOT EXISTS mcp_prepared_actions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    contexto TEXT NOT NULL DEFAULT 'mcp_teste',
    titulo TEXT NOT NULL,
    action_type TEXT NOT NULL,
    pergunta TEXT NOT NULL,
    payload TEXT NOT NULL DEFAULT '{}',
    status TEXT NOT NULL DEFAULT 'rascunho',
    criado_em TEXT NOT NULL,
    confirmado_em TEXT,
    FOREIGN KEY (user_id) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS receipts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nota_numero TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'FINALIZADO',
    observacao TEXT,
    conferente_id INTEGER,
    criado_em TEXT NOT NULL,
    finalizado_em TEXT NOT NULL,
    FOREIGN KEY (conferente_id) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS receipt_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    receipt_id INTEGER NOT NULL,
    stock_item_id INTEGER,
    codigo TEXT NOT NULL,
    codigo_barras TEXT,
    descricao TEXT NOT NULL,
    validade TEXT,
    quantidade REAL NOT NULL DEFAULT 0,
    criado_em TEXT NOT NULL,
    FOREIGN KEY (receipt_id) REFERENCES receipts (id) ON DELETE CASCADE,
    FOREIGN KEY (stock_item_id) REFERENCES stock_items (id)
);

CREATE TABLE IF NOT EXISTS audit_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    action TEXT NOT NULL,
    entity_type TEXT,
    entity_ref TEXT,
    details TEXT NOT NULL DEFAULT '{}',
    created_at TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS code_edit_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    arquivo TEXT NOT NULL,
    backup_path TEXT,
    acao TEXT NOT NULL DEFAULT 'save',
    detalhes TEXT NOT NULL DEFAULT '{}',
    criado_em TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS erp_stock_imports (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    filename TEXT NOT NULL,
    loja TEXT,
    data_base TEXT,
    total_linhas INTEGER NOT NULL DEFAULT 0,
    total_produtos INTEGER NOT NULL DEFAULT 0,
    total_grupos INTEGER NOT NULL DEFAULT 0,
    duplicados INTEGER NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'PREVIEW',
    modo TEXT,
    resumo_json TEXT NOT NULL DEFAULT '{}',
    criado_por INTEGER,
    criado_em TEXT NOT NULL,
    aplicado_em TEXT,
    FOREIGN KEY (criado_por) REFERENCES users (id)
);

CREATE TABLE IF NOT EXISTS erp_stock_import_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id INTEGER NOT NULL,
    row_number INTEGER,
    codigo TEXT NOT NULL,
    codigo_barras TEXT,
    nivel TEXT,
    descricao TEXT NOT NULL,
    linha TEXT,
    caminho_linha TEXT,
    preco_custo REAL NOT NULL DEFAULT 0,
    preco_venda REAL NOT NULL DEFAULT 0,
    saldo_qtd REAL NOT NULL DEFAULT 0,
    saldo_custo REAL NOT NULL DEFAULT 0,
    saldo_venda REAL NOT NULL DEFAULT 0,
    dias REAL NOT NULL DEFAULT 0,
    sugestao REAL NOT NULL DEFAULT 0,
    estoque_ideal REAL NOT NULL DEFAULT 0,
    stock_item_id INTEGER,
    saldo_anterior REAL,
    delta REAL,
    status TEXT NOT NULL DEFAULT 'PENDENTE',
    motivo TEXT,
    FOREIGN KEY (import_id) REFERENCES erp_stock_imports (id) ON DELETE CASCADE,
    FOREIGN KEY (stock_item_id) REFERENCES stock_items (id)
);
"""




def carregar_seed_estoque() -> dict[str, Any] | None:
    if not os.path.exists(STOCK_SEED_PATH):
        return None
    try:
        with open(STOCK_SEED_PATH, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    return payload


def garantir_produtos_seed(conn: sqlite3.Connection) -> None:
    payload = carregar_seed_estoque()
    if not payload:
        return

    version = str(payload.get("version") or "").strip()
    items = payload.get("items") or []
    if not version or not isinstance(items, list):
        return

    atual = conn.execute(
        "SELECT value FROM settings WHERE key = ?",
        ("stock_seed_version",),
    ).fetchone()
    if atual and (atual["value"] or "") == version:
        return

    rows: list[tuple[str, str, str, float, float, float, str]] = []
    agora = agora_iso()

    def normalizar_texto(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        if text.endswith(".0"):
            text = text[:-2]
        return text

    for item in items:
        if not isinstance(item, dict):
            continue
        codigo = normalizar_texto(item.get("codigo"))
        descricao = str(item.get("descricao") or "").strip()
        codigo_barras = normalizar_texto(item.get("codigo_barras"))
        if not codigo or not descricao:
            continue
        try:
            fator_embalagem = float(item.get("fator_embalagem") or 1)
        except (TypeError, ValueError):
            fator_embalagem = 1.0
        try:
            quantidade_atual = float(item.get("quantidade_atual") or 0)
        except (TypeError, ValueError):
            quantidade_atual = 0.0
        try:
            custo_unitario = float(item.get("custo_unitario") or 0)
        except (TypeError, ValueError):
            custo_unitario = 0.0
        if fator_embalagem <= 0:
            fator_embalagem = 1.0
        rows.append((codigo, codigo_barras, descricao, fator_embalagem, quantidade_atual, custo_unitario, agora))

    if rows:
        conn.executemany(
            """
            INSERT INTO stock_items (codigo, codigo_barras, descricao, fator_embalagem, quantidade_atual, custo_unitario, ativo, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(codigo) DO UPDATE SET
                descricao = excluded.descricao,
                codigo_barras = COALESCE(NULLIF(excluded.codigo_barras, ''), stock_items.codigo_barras),
                ativo = 1,
                fator_embalagem = CASE
                    WHEN stock_items.fator_embalagem IS NULL OR stock_items.fator_embalagem <= 0 THEN excluded.fator_embalagem
                    ELSE stock_items.fator_embalagem
                END,
                quantidade_atual = CASE
                    WHEN ABS(excluded.quantidade_atual) > 0.000001 THEN excluded.quantidade_atual
                    ELSE stock_items.quantidade_atual
                END,
                custo_unitario = CASE
                    WHEN ABS(excluded.custo_unitario) > 0.000001 THEN excluded.custo_unitario
                    ELSE stock_items.custo_unitario
                END,
                atualizado_em = CASE
                    WHEN stock_items.ativo = 0 OR ABS(excluded.quantidade_atual) > 0.000001 OR ABS(excluded.custo_unitario) > 0.000001 THEN excluded.atualizado_em
                    ELSE stock_items.atualizado_em
                END
            """,
            rows,
        )

    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        ("stock_seed_version", version),
    )


def get_setting(key: str, default: str = "") -> str:
    with closing(get_conn()) as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with closing(get_conn()) as conn:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )
        conn.commit()


def registrar_auditoria(action: str, entity_type: str = "", entity_ref: str = "", details: dict[str, Any] | None = None) -> None:
    """Registra uma ação importante sem deixar o log quebrar o fluxo principal."""
    try:
        user_id = g.user["id"] if getattr(g, "user", None) is not None else None
        with closing(get_conn()) as conn:
            conn.execute(
                """
                INSERT INTO audit_logs (user_id, action, entity_type, entity_ref, details, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    str(action or "")[:120],
                    str(entity_type or "")[:120],
                    str(entity_ref or "")[:300],
                    json.dumps(details or {}, ensure_ascii=False, default=str)[:5000],
                    agora_iso(),
                ),
            )
            conn.commit()
    except Exception:
        return


def query_one(sql: str, params: Iterable[Any] = ()) -> sqlite3.Row | None:
    with closing(get_conn()) as conn:
        return conn.execute(sql, tuple(params)).fetchone()


def query_all(sql: str, params: Iterable[Any] = ()) -> list[sqlite3.Row]:
    with closing(get_conn()) as conn:
        return conn.execute(sql, tuple(params)).fetchall()


def garantir_admin_inicial(conn: sqlite3.Connection) -> None:
    """Cria um admin inicial somente quando o banco estiver totalmente vazio.

    Isso protege deploys novos em Railway/Volume vazio. Em bancos já existentes,
    nada é alterado. A senha pode ser definida pela variável DEFAULT_ADMIN_PASSWORD.
    """
    total_users = conn.execute("SELECT COUNT(*) AS total FROM users").fetchone()["total"]
    if int(total_users or 0) > 0:
        return

    username = os.environ.get("DEFAULT_ADMIN_USERNAME", "admin").strip() or "admin"
    password = os.environ.get("DEFAULT_ADMIN_PASSWORD", "123456").strip() or "123456"
    nome = os.environ.get("DEFAULT_ADMIN_NAME", "Administrador").strip() or "Administrador"
    conn.execute(
        """
        INSERT INTO users (nome, username, password_hash, role, permission_level, access_rules, ativo, criado_em)
        VALUES (?, ?, ?, 'admin', 'admin', ?, 1, ?)
        """,
        (nome, username, generate_password_hash(password), serialize_access_rules(ACCESS_KEYS), agora_iso()),
    )


def ensure_default_data() -> None:
    with closing(get_conn()) as conn:
        conn.executescript(SCHEMA_SQL)
        ensure_schema_updates(conn)
        garantir_produtos_seed(conn)
        garantir_admin_inicial(conn)
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('vincular_estoque', '1')")
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('usar_conferente', '1')")
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('maintenance_mode', '0')")
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('code_editor_extra_password', '1')")
        conn.execute("INSERT OR IGNORE INTO stores (nome, ativo, criado_em) VALUES ('CD', 1, ?)", (agora_iso(),))
        conn.commit()


ensure_default_data()


@app.before_request
def bootstrap() -> Response | None:
    g.user = current_user()
    g.maintenance_mode = get_setting("maintenance_mode", "0") == "1"
    endpoint = request.endpoint or ""
    allowed = {"static", "login", "logout", "health", "favicon"}
    if g.maintenance_mode and endpoint not in allowed:
        if g.user is None or not user_is_admin(g.user):
            return render_template("manutencao.html", title="Sistema em manutenção"), 503
    return None


def current_user() -> sqlite3.Row | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    with closing(get_conn()) as conn:
        return conn.execute(
            """
            SELECT u.*, st.nome AS store_nome
            FROM users u
            LEFT JOIN stores st ON st.id = u.store_id
            WHERE u.id = ? AND u.ativo = 1
            """,
            (user_id,),
        ).fetchone()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def roles_required(*roles: str):
    normalized_roles = {normalize_role(role) for role in roles}

    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if g.user is None:
                return redirect(url_for("login"))
            if user_is_admin(g.user):
                return view(*args, **kwargs)
            if normalize_role(g.user["role"]) not in normalized_roles:
                return forbidden_redirect("Você não tem permissão para acessar essa área.")
            return view(*args, **kwargs)

        return wrapped

    return decorator


def parse_float(raw: str, field_name: str) -> float:
    valor = (raw or "").strip().replace(".", "").replace(",", ".") if "," in (raw or "") and (raw or "").count(",") == 1 and (raw or "").count(".") >= 1 else (raw or "").strip().replace(",", ".")
    try:
        number = float(valor)
    except ValueError as exc:
        raise ValueError(f"{field_name} inválido.") from exc
    if number < 0:
        raise ValueError(f"{field_name} não pode ser negativo.")
    return number


def role_badge(role: str) -> str:
    classes = {
        "admin": "badge badge-admin",
        "gerente": "badge badge-admin",
        "estoque": "badge badge-conferente",
        "separador": "badge badge-separador",
        "conferente": "badge badge-conferente",
        "balanco": "badge badge-admin",
        "desenvolvedor": "badge badge-admin",
        "visualizador": "badge",
    }
    return classes.get(role, "badge")


def status_class(status: str) -> str:
    normalized = status.lower().replace(" ", "_")
    return f"badge status-{normalized}"


def stock_movement_label(value: Any) -> str:
    labels = {
        "ENTRADA_INICIAL": "Cadastro inicial",
        "IMPORTACAO_ERP": "Importação ERP",
        "IMPORTACAO_ERP_NOVO": "Novo via ERP",
        "AJUSTE_MANUAL": "Ajuste manual",
        "RECONTAGEM": "Ajuste de quantidade",
        "REMOVIDO_ESTOQUE": "Remoção do estoque",
        "SAIDA_SEPARACAO": "Saída por separação",
        "ESTORNO_HISTORICO": "Estorno do histórico",
        "RECEBIMENTO_MERCADORIA": "Recebimento de mercadoria",
    }
    key = str(value or "").strip().upper()
    return labels.get(key, key.replace("_", " ").title() or "Movimentação")


def sanitize_stock_history_filters(args: Any) -> dict[str, str]:
    termo = str(args.get("q", "") or "").strip()
    hist_usuario = str(args.get("hist_usuario", "") or "").strip()
    hist_tipo = str(args.get("hist_tipo", "") or "").strip().upper()
    hist_data_inicial = str(args.get("hist_data_inicial", "") or "").strip()
    hist_data_final = str(args.get("hist_data_final", "") or "").strip()

    if not hist_usuario.isdigit():
        hist_usuario = ""
    if hist_tipo not in {key for key, _ in STOCK_MOVEMENT_TYPE_OPTIONS}:
        hist_tipo = ""
    return {
        "q": termo,
        "hist_usuario": hist_usuario,
        "hist_tipo": hist_tipo,
        "hist_data_inicial": hist_data_inicial,
        "hist_data_final": hist_data_final,
    }


def build_stock_history_query(filters: dict[str, str], limit: int | None = 80) -> tuple[str, list[Any]]:
    movement_filters: list[str] = []
    movement_params: list[Any] = []
    termo = filters["q"]
    if termo:
        movement_filters.append("(si.codigo = ? OR si.codigo_barras = ? OR si.codigo LIKE ? OR si.descricao LIKE ? OR si.codigo_barras LIKE ?)")
        movement_like = f"%{termo}%"
        movement_params.extend([termo, termo, movement_like, movement_like, movement_like])
    if filters["hist_usuario"]:
        movement_filters.append("sm.criado_por = ?")
        movement_params.append(int(filters["hist_usuario"]))
    if filters["hist_tipo"]:
        movement_filters.append("sm.tipo = ?")
        movement_params.append(filters["hist_tipo"])
    if filters["hist_data_inicial"]:
        movement_filters.append("date(sm.criado_em) >= date(?)")
        movement_params.append(filters["hist_data_inicial"])
    if filters["hist_data_final"]:
        movement_filters.append("date(sm.criado_em) <= date(?)")
        movement_params.append(filters["hist_data_final"])

    movement_where = "WHERE " + " AND ".join(movement_filters) if movement_filters else ""
    sql = f"""
        SELECT sm.*, si.codigo, si.codigo_barras, si.descricao,
               u.nome AS usuario_nome,
               u.username AS usuario_login
        FROM stock_movements sm
        JOIN stock_items si ON si.id = sm.stock_item_id
        LEFT JOIN users u ON u.id = sm.criado_por
        {movement_where}
        ORDER BY sm.id DESC
    """
    if limit is not None:
        sql += f"\n        LIMIT {int(limit)}"
    return sql, movement_params


def fetch_stock_movements(filters: dict[str, str], limit: int | None = 80) -> list[sqlite3.Row]:
    sql, params = build_stock_history_query(filters, limit=limit)
    return query_all(sql, params)


def stock_history_filter_labels(filters: dict[str, str]) -> list[str]:
    labels: list[str] = []
    if filters["q"]:
        labels.append(f"Busca: {filters['q']}")
    if filters["hist_usuario"]:
        usuario = query_one("SELECT nome, username FROM users WHERE id = ?", (int(filters["hist_usuario"]),))
        if usuario:
            nome = usuario["nome"] or usuario["username"] or "Usuário"
            if usuario["username"] and usuario["username"] != nome:
                nome = f"{nome} ({usuario['username']})"
            labels.append(f"Usuário: {nome}")
    if filters["hist_tipo"]:
        labels.append(f"Tipo: {stock_movement_label(filters['hist_tipo'])}")
    if filters["hist_data_inicial"]:
        labels.append(f"Data inicial: {filters['hist_data_inicial']}")
    if filters["hist_data_final"]:
        labels.append(f"Data final: {filters['hist_data_final']}")
    return labels


def stock_history_export_filename(extensao: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"historico-estoque-{stamp}.{extensao}"


app.jinja_env.globals.update(
    fmt_num=fmt_num,
    fmt_money=fmt_money,
    fmt_fator_embalagem=fmt_fator_embalagem,
    quantidade_em_embalagens=quantidade_em_embalagens,
    role_badge=role_badge,
    role_label=role_label,
    permission_level_label=permission_level_label,
    user_is_admin=user_is_admin,
    user_has_access=user_has_access,
    user_access_set=user_access_set,
    access_labels_for_user=access_labels_for_user,
    access_options=ACCESS_OPTIONS,
    permission_level_options=[("comum", "Comum"), ("admin", "Admin")],
    status_class=status_class,
    lote_operacao_chave_row=lote_operacao_chave_row,
    can_adjust_stock=can_adjust_stock,
    can_edit_stock_registration=can_edit_stock_registration,
    stock_movement_label=stock_movement_label,
)




@app.route("/login", methods=["GET", "POST"])
def login() -> str | Response:
    if g.user is not None:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = query_one(
            "SELECT * FROM users WHERE username = ? AND ativo = 1", (username,)
        )
        if user is None or not check_password_hash(user["password_hash"], password):
            flash("Usuário ou senha inválidos.", "error")
            return render_template("login.html", title="Login")

        session["user_id"] = user["id"]
        g.user = user
        registrar_auditoria("login", "user", str(user["id"]), {"username": username})
        flash("Login realizado com sucesso.", "success")
        return redirect(url_for(first_allowed_endpoint(user)))

    return render_template("login.html", title="Login")


@app.get("/logout")
def logout() -> Response:
    session.clear()
    flash("Sessão encerrada.", "success")
    return redirect(url_for("login"))


@app.errorhandler(Exception)
def tratar_erro_geral(exc: Exception):
    if isinstance(exc, HTTPException):
        return exc
    error_id = uuid.uuid4().hex[:10].upper()
    detalhe = "".join(traceback.format_exception_only(type(exc), exc)).strip()
    registrar_auditoria("erro_sistema", "exception", error_id, {
        "endpoint": request.endpoint,
        "path": request.path,
        "erro": detalhe,
    })
    if app.debug:
        raise exc
    return render_template(
        "erro_sistema.html",
        title="Erro no sistema",
        error_id=error_id,
        endpoint=request.endpoint or "-",
        path=request.path,
        detalhe=detalhe,
        is_admin=bool(g.user is not None and user_is_admin(g.user)),
    ), 500


@app.route("/minha-conta", methods=["GET", "POST"])
@login_required
def minha_conta() -> str | Response:
    if request.method == "POST":
        acao = request.form.get("acao", "senha")

        if acao == "foto":
            arquivo = request.files.get("foto")
            if arquivo is None or not arquivo.filename:
                flash("Escolha uma foto para atualizar o perfil.", "error")
                return redirect(url_for("minha_conta"))
            nome = secure_filename(arquivo.filename)
            ext = os.path.splitext(nome)[1].lower()
            if ext not in {".png", ".jpg", ".jpeg", ".webp"}:
                flash("Envie uma imagem PNG, JPG, JPEG ou WEBP.", "error")
                return redirect(url_for("minha_conta"))
            folder = os.path.join(BASE_DIR, "static", "uploads", "perfis")
            os.makedirs(folder, exist_ok=True)
            final_name = f"user_{g.user['id']}_{int(datetime.now().timestamp())}{ext}"
            arquivo.save(os.path.join(folder, final_name))
            foto = f"uploads/perfis/{final_name}"
            with closing(get_conn()) as conn:
                conn.execute("UPDATE users SET foto_perfil = ? WHERE id = ?", (foto, g.user["id"]))
                conn.commit()
            registrar_auditoria("atualizar_foto_perfil", "users", str(g.user["id"]), {"foto": foto})
            flash("Foto do perfil atualizada.", "success")
            return redirect(url_for("minha_conta"))

        senha_atual = request.form.get("senha_atual", "")
        nova_senha = request.form.get("nova_senha", "")
        confirmar_senha = request.form.get("confirmar_senha", "")

        if not check_password_hash(g.user["password_hash"], senha_atual):
            flash("A senha atual está incorreta.", "error")
            return redirect(url_for("minha_conta"))
        if len(nova_senha) < 4:
            flash("A nova senha precisa ter pelo menos 4 caracteres.", "error")
            return redirect(url_for("minha_conta"))
        if nova_senha != confirmar_senha:
            flash("A confirmação da nova senha não confere.", "error")
            return redirect(url_for("minha_conta"))
        if check_password_hash(g.user["password_hash"], nova_senha):
            flash("Escolha uma senha diferente da atual.", "error")
            return redirect(url_for("minha_conta"))

        with closing(get_conn()) as conn:
            conn.execute(
                "UPDATE users SET password_hash = ? WHERE id = ?",
                (generate_password_hash(nova_senha), g.user["id"]),
            )
            conn.commit()

        flash("Senha alterada com sucesso.", "success")
        return redirect(url_for("minha_conta"))

    return render_template("minha_conta.html", title="Minha conta")


def ultimos_lotes_resumo(limit: int = 8, include_canceladas: bool = True) -> list[sqlite3.Row]:
    chave_expr = lote_operacao_chave_expr("s")
    where_parts = ["1=1"]
    if not include_canceladas:
        where_parts.append("s.status <> 'CANCELADA'")
    if normalize_role(g.user["role"]) == "separador" and not user_is_admin(g.user):
        where_parts.append("s.responsavel_id = ?")
        params: list[Any] = [g.user["id"], limit]
    elif normalize_role(g.user["role"]) == "conferente" and not user_is_admin(g.user):
        where_parts.append("s.conferente_id = ?")
        params = [g.user["id"], limit]
    else:
        params = [limit]
    where = " AND ".join(where_parts)
    return query_all(
        f"""
        SELECT {chave_expr} AS operacao_chave,
               s.lote_nome,
               s.data_referencia,
               MAX(COALESCE(s.finalizado_em, s.criado_em)) AS data_evento,
               MAX(r.nome) AS responsavel_nome,
               MAX(c.nome) AS conferente_nome,
               COUNT(*) AS total_lojas,
               GROUP_CONCAT(st.nome, ' • ') AS lojas,
               CASE
                   WHEN SUM(CASE WHEN s.status = 'FINALIZADA' THEN 1 ELSE 0 END) > 0 THEN 'FINALIZADA'
                   WHEN SUM(CASE WHEN s.status = 'AGUARDANDO_CONFERENCIA' THEN 1 ELSE 0 END) > 0 THEN 'AGUARDANDO_CONFERENCIA'
                   WHEN SUM(CASE WHEN s.status = 'EM_SEPARACAO' THEN 1 ELSE 0 END) > 0 THEN 'EM_SEPARACAO'
                   WHEN SUM(CASE WHEN s.status = 'ABERTA' THEN 1 ELSE 0 END) > 0 THEN 'ABERTA'
                   ELSE 'CANCELADA'
               END AS status_resumo
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        WHERE {where}
        GROUP BY operacao_chave, s.lote_nome, s.data_referencia
        ORDER BY MAX(COALESCE(s.finalizado_em, s.criado_em)) DESC, MAX(s.id) DESC
        LIMIT ?
        """,
        params,
    )


def carregar_lote_completo(operacao_chave: str) -> list[sqlite3.Row]:
    chave_expr = lote_operacao_chave_expr("s")
    rows = query_all(
        f"""
        SELECT s.*, st.nome AS store_nome,
               r.nome AS responsavel_nome,
               c.nome AS conferente_nome,
               {chave_expr} AS operacao_chave
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        WHERE {chave_expr} = ?
        ORDER BY s.id ASC
        """,
        (operacao_chave,),
    )
    return sort_store_rows(rows)


def excluir_separacao_cancelada_no_conn(conn: sqlite3.Connection, separation_id: int) -> None:
    separation = conn.execute("SELECT id, status FROM separations WHERE id = ?", (separation_id,)).fetchone()
    if separation is None:
        raise ValueError("Separação não encontrada.")
    if separation["status"] != "CANCELADA":
        raise ValueError("Só é possível excluir de vez uma separação cancelada.")
    conn.execute("DELETE FROM separation_items WHERE separation_id = ?", (separation_id,))
    conn.execute("DELETE FROM separations WHERE id = ?", (separation_id,))


def apagar_historico_separacao_no_conn(conn: sqlite3.Connection, separation_id: int, actor_id: int) -> None:
    separation = conn.execute(
        """
        SELECT s.*, st.nome AS store_nome
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        WHERE s.id = ?
        """,
        (separation_id,),
    ).fetchone()
    if separation is None:
        raise ValueError("Separação não encontrada.")
    if separation["status"] != "FINALIZADA":
        raise ValueError("Só é possível apagar do histórico uma separação finalizada.")

    usar_controle_global = get_setting("vincular_estoque", "1") == "1"
    precisa_estornar = usar_controle_global and bool(separation["usar_estoque"])
    itens = conn.execute("SELECT * FROM separation_items WHERE separation_id = ?", (separation_id,)).fetchall()
    if precisa_estornar:
        for item in itens:
            stock = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item["codigo"],)).fetchone()
            if stock is None:
                conn.execute(
                    "INSERT INTO stock_items (codigo, descricao, quantidade_atual, custo_unitario, ativo, atualizado_em) VALUES (?, ?, 0, ?, 1, ?)",
                    (item["codigo"], item["descricao"], item["custo_unitario_ref"], agora_iso()),
                )
                stock = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item["codigo"],)).fetchone()
            novo_saldo = float(stock["quantidade_atual"]) + float(item["quantidade_separada"])
            conn.execute(
                "UPDATE stock_items SET quantidade_atual = ?, ativo = 1, atualizado_em = ? WHERE id = ?",
                (novo_saldo, agora_iso(), stock["id"]),
            )
            conn.execute(
                "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, 'ESTORNO_HISTORICO', ?, ?, 'SEPARACAO', ?, ?, ?)",
                (stock["id"], float(item["quantidade_separada"]), f"Estorno do histórico da separação {separation['lote_nome']} - {separation['store_nome']}", separation_id, actor_id, agora_iso()),
            )
    observacao_atual = (separation["observacao"] or "").strip()
    nota_cancelamento = f"Histórico removido pelo admin em {agora_br()}."
    nova_observacao = (observacao_atual + "\n" + nota_cancelamento).strip() if observacao_atual else nota_cancelamento
    conn.execute(
        "UPDATE separations SET status = 'CANCELADA', finalizado_em = NULL, enviado_conferencia_em = NULL, observacao = ? WHERE id = ?",
        (nova_observacao, separation_id),
    )


def dashboard_stats() -> dict[str, Any]:
    user = g.user
    where_clauses = ["s.status <> 'CANCELADA'"]
    params: list[Any] = []
    if user and normalize_role(user["role"]) == "separador" and not user_is_admin(user):
        where_clauses.append("s.responsavel_id = ?")
        params.append(user["id"])
    elif user and normalize_role(user["role"]) == "conferente" and not user_is_admin(user):
        where_clauses.append("s.conferente_id = ?")
        params.append(user["id"])
    where = "WHERE " + " AND ".join(where_clauses)

    resumo = query_one(
        f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN status = 'ABERTA' THEN 1 ELSE 0 END) AS abertas,
            SUM(CASE WHEN status = 'AGUARDANDO_CONFERENCIA' THEN 1 ELSE 0 END) AS aguardando,
            SUM(CASE WHEN status = 'FINALIZADA' THEN 1 ELSE 0 END) AS finalizadas
        FROM separations s
        {where}
        """,
        params,
    )
    estoque = query_one(
        "SELECT COUNT(*) AS itens, COALESCE(SUM(quantidade_atual), 0) AS total_quantidade FROM stock_items"
    )
    return {
        "total": resumo["total"] if resumo else 0,
        "abertas": resumo["abertas"] if resumo else 0,
        "aguardando": resumo["aguardando"] if resumo else 0,
        "finalizadas": resumo["finalizadas"] if resumo else 0,
        "itens_estoque": estoque["itens"] if estoque else 0,
        "qtd_estoque": estoque["total_quantidade"] if estoque else 0,
        "vincular_estoque": get_setting("vincular_estoque", "1") == "1",
        "usar_conferente": get_setting("usar_conferente", "1") == "1",
    }





def painel_gerencial_padaria_cd() -> dict[str, Any]:
    """Resumo gerencial simplificado para o painel, focado em Padaria - Industria CD.

    Esta versão evita quebrar o dashboard quando o banco antigo ainda não tem alguma
    coluna nova. Também calcula o valor considerando o fator de embalagem.
    """
    vazio = {
        "total_itens": 0,
        "quantidade_total": 0.0,
        "valor_estoque": 0.0,
        "itens_zerados": 0,
        "itens_abaixo": 0,
        "linhas": [],
    }
    try:
        with closing(get_conn()) as conn:
            cols = {row[1] for row in conn.execute("PRAGMA table_info(stock_items)").fetchall()}
            if not cols:
                return vazio

            def col(name: str, fallback: str = "NULL") -> str:
                return name if name in cols else fallback

            codigo_col = col("codigo")
            descricao_col = col("descricao", "''")
            ativo_expr = col("ativo", "1")
            qtd_expr = col("quantidade_atual", "0")
            custo_expr = col("custo_unitario", "0")
            fator_expr = col("fator_embalagem", "1")
            linha_expr = col("linha_erp", "''")
            caminho_expr = col("linha_caminho_erp", "''")
            nivel_expr = col("erp_nivel", "''")
            loja_expr = col("erp_loja", "''")

            rows = conn.execute(f"""
                SELECT
                    {codigo_col} AS codigo,
                    {descricao_col} AS descricao,
                    COALESCE({qtd_expr}, 0) AS quantidade_atual,
                    COALESCE({custo_expr}, 0) AS custo_unitario,
                    COALESCE({fator_expr}, 1) AS fator_embalagem,
                    COALESCE({linha_expr}, '') AS linha_erp,
                    COALESCE({caminho_expr}, '') AS linha_caminho_erp,
                    COALESCE({nivel_expr}, '') AS erp_nivel,
                    COALESCE({loja_expr}, '') AS erp_loja
                FROM stock_items
                WHERE COALESCE({ativo_expr}, 1) = 1
            """).fetchall()

        total_itens = 0
        quantidade_total = 0.0
        valor_estoque = 0.0
        itens_zerados = 0
        itens_abaixo = 0
        grupos: dict[str, dict[str, Any]] = {}

        for item in rows:
            texto_linha = " ".join([
                str(item["linha_erp"] or ""),
                str(item["linha_caminho_erp"] or ""),
                str(item["erp_nivel"] or ""),
                str(item["erp_loja"] or ""),
            ]).upper()
            # Filtro mais tolerante para evitar excluir itens por diferença de cadastro.
            if "PADARIA" not in texto_linha:
                continue
            if "INDUSTRIA" not in texto_linha and "INDÚSTRIA" not in texto_linha:
                continue
            if "CD" not in texto_linha and "CENTRO" not in texto_linha:
                # Mantém compatibilidade: alguns ERPs não gravam CD no caminho, mas a linha ainda é a correta.
                pass

            try:
                quantidade = float(item["quantidade_atual"] or 0)
            except (TypeError, ValueError):
                quantidade = 0.0
            try:
                custo = float(item["custo_unitario"] or 0)
            except (TypeError, ValueError):
                custo = 0.0
            try:
                fator = float(item["fator_embalagem"] or 1)
            except (TypeError, ValueError):
                fator = 1.0
            if fator <= 0:
                fator = 1.0

            total_itens += 1
            quantidade_total += quantidade
            valor_estoque += quantidade * fator * custo
            if quantidade <= 0:
                itens_zerados += 1
            elif quantidade <= 10:
                itens_abaixo += 1

            nome_linha = (item["linha_erp"] or "Sem linha")
            grupo = grupos.setdefault(nome_linha, {"linha": nome_linha, "total_itens": 0, "abaixo": 0, "zerados": 0})
            grupo["total_itens"] += 1
            if quantidade <= 0:
                grupo["zerados"] += 1
            elif quantidade <= 10:
                grupo["abaixo"] += 1

        linhas = sorted(grupos.values(), key=lambda x: x["total_itens"], reverse=True)[:8]
        return {
            "total_itens": int(total_itens),
            "quantidade_total": float(quantidade_total),
            "valor_estoque": float(valor_estoque),
            "itens_zerados": int(itens_zerados),
            "itens_abaixo": int(itens_abaixo),
            "linhas": linhas,
        }
    except Exception as exc:
        try:
            registrar_auditoria("erro_painel_gerencial_padaria", "dashboard", "padaria_cd", {"erro": str(exc)})
        except Exception:
            pass
        return vazio


def painel_chat_resumo() -> dict[str, Any]:
    """Mostra mensagens recentes dos chats em que o usuário participa sem derrubar o dashboard."""
    try:
        uid = int(g.user["id"])
        pendentes = query_one("""
            SELECT COUNT(*) AS c
            FROM chat_messages m
            JOIN chat_group_members gm ON gm.group_id = m.group_id AND gm.user_id = ?
            JOIN chat_groups cg ON cg.id = m.group_id AND cg.ativo = 1
            WHERE COALESCE(m.user_id, 0) <> ?
              AND m.criado_em >= datetime('now', '-2 days')
        """, (uid, uid))
        ultimas = query_all("""
            SELECT m.id, m.mensagem, m.criado_em, cg.nome AS grupo_nome, u.nome AS usuario_nome
            FROM chat_messages m
            JOIN chat_group_members gm ON gm.group_id = m.group_id AND gm.user_id = ?
            JOIN chat_groups cg ON cg.id = m.group_id AND cg.ativo = 1
            LEFT JOIN users u ON u.id = m.user_id
            WHERE COALESCE(m.user_id, 0) <> ?
            ORDER BY m.id DESC
            LIMIT 3
        """, (uid, uid))
        return {"novas": int((pendentes["c"] if pendentes else 0) or 0), "ultimas": ultimas}
    except Exception as exc:
        try:
            registrar_auditoria("erro_painel_chat", "dashboard", "chat_resumo", {"erro": str(exc)})
        except Exception:
            pass
        return {"novas": 0, "ultimas": []}

@app.get("/")
@login_required
@module_required("painel")
def dashboard() -> str:
    lojas_ativas = query_one("SELECT COUNT(*) AS c FROM stores WHERE ativo = 1")["c"]
    usuarios_ativos = query_one("SELECT COUNT(*) AS c FROM users WHERE ativo = 1")["c"]
    finalizadas_hoje = query_one(
        "SELECT COUNT(*) AS c FROM separations WHERE finalizado_em LIKE ?", (datetime.now().strftime("%Y-%m-%d") + "%",)
    )["c"]
    ultimos_lotes = ultimos_lotes_resumo(8, include_canceladas=True)
    return render_template(
        "dashboard.html",
        title="Painel",
        stats=dashboard_stats(),
        ultimos_lotes=ultimos_lotes,
        lojas_ativas=lojas_ativas,
        usuarios_ativos=usuarios_ativos,
        finalizadas_hoje=finalizadas_hoje,
        gerencial_padaria=painel_gerencial_padaria_cd(),
        chat_resumo=painel_chat_resumo(),
    )


@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@module_required("usuarios")
@roles_required("admin")
def usuarios() -> str | Response:
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        role = normalize_role(request.form.get("role", "separador"))
        permission_level = normalize_permission_level(request.form.get("permission_level", "comum"), role)
        store_id_raw = request.form.get("store_id", "").strip()
        store_id = int(store_id_raw) if store_id_raw.isdigit() else None
        access_rules = set(request.form.getlist("access_rules"))
        if permission_level == "admin":
            access_rules = set(ACCESS_KEYS)
        else:
            access_rules = {item for item in access_rules if item in ACCESS_KEYS}
        if not nome or not username or not password:
            flash("Preencha os dados do usuário corretamente.", "error")
            return redirect(url_for("usuarios"))
        if permission_level != "admin" and not access_rules:
            flash("Selecione pelo menos um acesso para o usuário comum.", "error")
            return redirect(url_for("usuarios"))
        try:
            with closing(get_conn()) as conn:
                conn.execute(
                    "INSERT INTO users (nome, username, password_hash, role, permission_level, access_rules, store_id, ativo, criado_em) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)",
                    (nome, username, generate_password_hash(password), role, permission_level, serialize_access_rules(access_rules), store_id, agora_iso()),
                )
                conn.commit()
            flash("Usuário criado com sucesso.", "success")
        except sqlite3.IntegrityError:
            flash("Esse login já existe.", "error")
        return redirect(url_for("usuarios"))

    users = query_all("""
        SELECT u.*, st.nome AS store_nome
        FROM users u
        LEFT JOIN stores st ON st.id = u.store_id
        ORDER BY u.ativo DESC, CASE u.permission_level WHEN 'admin' THEN 0 ELSE 1 END, u.role, u.nome
    """)
    lojas = query_all("SELECT id, nome FROM stores WHERE ativo = 1 ORDER BY nome")
    return render_template(
        "usuarios.html",
        title="Usuários",
        users=users,
        access_options=ACCESS_OPTIONS,
        role_options=list(ROLE_LABELS.items()),
        lojas=lojas,
    )


@app.post("/usuarios/<int:user_id>/salvar")
@login_required
@module_required("usuarios")
@roles_required("admin")
def salvar_usuario(user_id: int) -> Response:
    nome = request.form.get("nome", "").strip()
    username = request.form.get("username", "").strip()
    role = normalize_role(request.form.get("role", "separador"))
    permission_level = normalize_permission_level(request.form.get("permission_level", "comum"), role)
    nova_senha = request.form.get("nova_senha", "")
    store_id_raw = request.form.get("store_id", "").strip()
    store_id = int(store_id_raw) if store_id_raw.isdigit() else None
    access_rules = set(request.form.getlist("access_rules"))
    if permission_level == "admin":
        access_rules = set(ACCESS_KEYS)
    else:
        access_rules = {item for item in access_rules if item in ACCESS_KEYS}

    if not nome or not username:
        flash("Preencha nome e login do usuário.", "error")
        return redirect(url_for("usuarios"))
    if permission_level != "admin" and not access_rules:
        flash("Selecione pelo menos um acesso para o usuário comum.", "error")
        return redirect(url_for("usuarios"))
    if nova_senha and len(nova_senha) < 4:
        flash("A nova senha precisa ter pelo menos 4 caracteres.", "error")
        return redirect(url_for("usuarios"))

    with closing(get_conn()) as conn:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if user is None:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("usuarios"))

        if user_id == g.user["id"] and not user_is_admin({**dict(user), "role": role, "permission_level": permission_level}):
            if count_admin_users(conn, exclude_user_id=user_id) == 0:
                flash("Não é possível remover o nível admin do último admin ativo.", "error")
                return redirect(url_for("usuarios"))

        try:
            conn.execute(
                "UPDATE users SET nome = ?, username = ?, role = ?, permission_level = ?, access_rules = ?, store_id = ? WHERE id = ?",
                (nome, username, role, permission_level, serialize_access_rules(access_rules), store_id, user_id),
            )
            if nova_senha:
                conn.execute(
                    "UPDATE users SET password_hash = ? WHERE id = ?",
                    (generate_password_hash(nova_senha), user_id),
                )
            conn.commit()
        except sqlite3.IntegrityError:
            flash("Esse login já existe.", "error")
            return redirect(url_for("usuarios"))

    flash("Usuário atualizado com sucesso.", "success")
    return redirect(url_for("usuarios"))


@app.post("/usuarios/<int:user_id>/alternar")
@login_required
@module_required("usuarios")
@roles_required("admin")
def alternar_usuario(user_id: int) -> Response:
    if user_id == g.user["id"]:
        flash("Você não pode desativar seu próprio usuário por aqui.", "error")
        return redirect(url_for("usuarios"))
    with closing(get_conn()) as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if row is None:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("usuarios"))
        novo = 0 if row["ativo"] else 1
        if novo == 0 and user_is_admin(row) and count_admin_users(conn, exclude_user_id=user_id) == 0:
            flash("Não é possível desativar o último admin ativo do sistema.", "error")
            return redirect(url_for("usuarios"))
        conn.execute("UPDATE users SET ativo = ? WHERE id = ?", (novo, user_id))
        conn.commit()
    flash("Usuário atualizado.", "success")
    return redirect(url_for("usuarios"))


def tabela_existe(conn: sqlite3.Connection, tabela: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
        (tabela,),
    ).fetchone()
    return row is not None


def colunas_tabela(conn: sqlite3.Connection, tabela: str) -> set[str]:
    if not tabela_existe(conn, tabela):
        return set()
    return {str(row["name"]) for row in conn.execute(f"PRAGMA table_info({tabela})").fetchall()}


def contar_vinculos_usuario(conn: sqlite3.Connection, user_id: int) -> dict[str, int]:
    checks = {
        "separações como responsável": ("separations", "responsavel_id"),
        "separações como conferente": ("separations", "conferente_id"),
        "separações criadas": ("separations", "criado_por"),
        "movimentações de estoque": ("stock_movements", "criado_por"),
        "balanços criados": ("balance_counts", "criado_por"),
        "balanços confirmados": ("balance_counts", "confirmado_por"),
        "consultas MCP salvas": ("mcp_saved_queries", "user_id"),
        "histórico de consultas MCP": ("mcp_query_history", "user_id"),
        "ações MCP preparadas": ("mcp_prepared_actions", "user_id"),
        "auditoria": ("audit_logs", "user_id"),
        "edições de código": ("code_edit_history", "user_id"),
        "importações ERP": ("erp_stock_imports", "criado_por"),
    }
    resultado: dict[str, int] = {}
    for label, (tabela, coluna) in checks.items():
        if coluna not in colunas_tabela(conn, tabela):
            continue
        row = conn.execute(f"SELECT COUNT(*) AS c FROM {tabela} WHERE {coluna} = ?", (user_id,)).fetchone()
        total = int((row["c"] if row else 0) or 0)
        if total > 0:
            resultado[label] = total
    return resultado




def carregar_historicos_vinculados_usuario(conn: sqlite3.Connection, user_id: int) -> list[sqlite3.Row]:
    """Lista separações finalizadas vinculadas ao usuário para revisão/estorno seguro."""
    return conn.execute(
        """
        SELECT s.*, st.nome AS store_nome,
               r.nome AS responsavel_nome,
               c.nome AS conferente_nome,
               creator.nome AS criado_por_nome,
               COUNT(si.id) AS total_itens,
               COALESCE(SUM(si.quantidade_separada), 0) AS total_separado
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        LEFT JOIN users creator ON creator.id = s.criado_por
        LEFT JOIN separation_items si ON si.separation_id = s.id
        WHERE s.status = 'FINALIZADA'
          AND (s.responsavel_id = ? OR s.conferente_id = ? OR s.criado_por = ?)
        GROUP BY s.id
        ORDER BY COALESCE(s.finalizado_em, s.criado_em) DESC, s.id DESC
        """,
        (user_id, user_id, user_id),
    ).fetchall()


def resumo_estorno_historicos_usuario(conn: sqlite3.Connection, separation_ids: list[int]) -> list[sqlite3.Row]:
    if not separation_ids:
        return []
    placeholders = ",".join("?" for _ in separation_ids)
    return conn.execute(
        f"""
        SELECT si.codigo,
               MAX(si.descricao) AS descricao,
               COUNT(DISTINCT si.separation_id) AS total_historicos,
               COALESCE(SUM(si.quantidade_separada), 0) AS quantidade_a_devolver
        FROM separation_items si
        WHERE si.separation_id IN ({placeholders})
        GROUP BY si.codigo
        ORDER BY MAX(si.descricao) COLLATE NOCASE
        """,
        tuple(separation_ids),
    ).fetchall()


def usuario_tem_vinculos(conn: sqlite3.Connection, user_id: int) -> bool:
    return bool(contar_vinculos_usuario(conn, user_id))


def desativar_usuario_com_seguranca(conn: sqlite3.Connection, user_id: int) -> None:
    conn.execute("UPDATE users SET ativo = 0 WHERE id = ?", (user_id,))



@app.route("/usuarios/<int:user_id>/historico", methods=["GET", "POST"])
@login_required
@module_required("usuarios")
@roles_required("admin")
def historico_usuario(user_id: int) -> str | Response:
    with closing(get_conn()) as conn:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if user is None:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("usuarios"))

        historicos = carregar_historicos_vinculados_usuario(conn, user_id)
        historico_ids_validos = {int(row["id"]) for row in historicos}

        if request.method == "POST":
            selecionados = [int(v) for v in request.form.getlist("separation_ids") if str(v).isdigit()]
            selecionados = [sid for sid in selecionados if sid in historico_ids_validos]
            if not selecionados:
                flash("Selecione pelo menos um histórico finalizado para excluir/estornar.", "error")
                return redirect(url_for("historico_usuario", user_id=user_id))

            acao = request.form.get("acao", "preview")
            resumo = resumo_estorno_historicos_usuario(conn, selecionados)

            if acao == "preview":
                selecionados_set = set(selecionados)
                historicos_preview = [row for row in historicos if int(row["id"]) in selecionados_set]
                return render_template(
                    "usuario_historico.html",
                    title="Histórico vinculado ao usuário",
                    user=user,
                    historicos=historicos,
                    historicos_preview=historicos_preview,
                    resumo=resumo,
                    selecionados=selecionados,
                    modo_preview=True,
                )

            if acao == "confirmar":
                senha_admin = request.form.get("senha_admin", "")
                if not check_password_hash(g.user["password_hash"], senha_admin):
                    flash("Senha do admin incorreta. Nada foi alterado.", "error")
                    return redirect(url_for("historico_usuario", user_id=user_id))

                total_estornado = 0.0
                total_historicos = 0
                for sid in selecionados:
                    sep = conn.execute("SELECT status FROM separations WHERE id = ?", (sid,)).fetchone()
                    if sep is None or sep["status"] != "FINALIZADA":
                        continue
                    before = conn.execute(
                        "SELECT COALESCE(SUM(quantidade_separada), 0) AS total FROM separation_items WHERE separation_id = ?",
                        (sid,),
                    ).fetchone()
                    apagar_historico_separacao_no_conn(conn, sid, g.user["id"])
                    total_estornado += float((before["total"] if before else 0) or 0)
                    total_historicos += 1

                conn.commit()
                registrar_auditoria(
                    "excluir_historico_usuario_com_estorno",
                    "user",
                    str(user_id),
                    {
                        "usuario_alvo": user["username"],
                        "historicos": selecionados,
                        "total_historicos": total_historicos,
                        "total_estornado": total_estornado,
                    },
                )
                flash(
                    f"Histórico removido com segurança: {total_historicos} registro(s). "
                    f"Total devolvido ao estoque: {total_estornado:g}.",
                    "success",
                )
                return redirect(url_for("historico_usuario", user_id=user_id))

        return render_template(
            "usuario_historico.html",
            title="Histórico vinculado ao usuário",
            user=user,
            historicos=historicos,
            historicos_preview=[],
            resumo=[],
            selecionados=[],
            modo_preview=False,
        )

@app.post("/usuarios/<int:user_id>/excluir")
@login_required
@module_required("usuarios")
@roles_required("admin")
def excluir_usuario(user_id: int) -> Response:
    if user_id == g.user["id"]:
        flash("Você não pode excluir o próprio usuário logado.", "error")
        return redirect(url_for("usuarios"))

    with closing(get_conn()) as conn:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if user is None:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("usuarios"))

        outros_admins = count_admin_users(conn, exclude_user_id=user_id)
        if user_is_admin(user) and int(outros_admins or 0) == 0:
            flash("Não é possível excluir ou desativar o último admin do sistema.", "error")
            return redirect(url_for("usuarios"))

        vinculos = contar_vinculos_usuario(conn, user_id)
        if vinculos:
            desativar_usuario_com_seguranca(conn, user_id)
            conn.commit()
            registrar_auditoria(
                "desativar_usuario_com_vinculos",
                "user",
                str(user_id),
                {"username": user["username"], "vinculos": vinculos},
            )
            resumo = ", ".join(f"{nome}: {qtd}" for nome, qtd in list(vinculos.items())[:4])
            if len(vinculos) > 4:
                resumo += ", ..."
            flash(
                "Esse usuário possui histórico no sistema e não pode ser apagado sem quebrar registros antigos. "
                f"Ele foi desativado com segurança. Vínculos encontrados: {resumo}.",
                "warning",
            )
            return redirect(url_for("usuarios"))

        try:
            conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.commit()
            registrar_auditoria("excluir_usuario", "user", str(user_id), {"username": user["username"]})
        except sqlite3.IntegrityError as exc:
            conn.rollback()
            desativar_usuario_com_seguranca(conn, user_id)
            conn.commit()
            registrar_auditoria(
                "desativar_usuario_por_integridade",
                "user",
                str(user_id),
                {"username": user["username"], "erro": str(exc)},
            )
            flash(
                "Não foi possível apagar esse usuário porque ele está ligado ao histórico do sistema. "
                "Para preservar os registros, ele foi desativado com segurança.",
                "warning",
            )
            return redirect(url_for("usuarios"))

    flash("Usuário excluído com sucesso.", "success")
    return redirect(url_for("usuarios"))


@app.route("/lojas", methods=["GET", "POST"])
@login_required
@module_required("lojas")
@roles_required("admin")
def lojas() -> str | Response:
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        if not nome:
            flash("Informe o nome da loja.", "error")
            return redirect(url_for("lojas"))
        try:
            with closing(get_conn()) as conn:
                conn.execute(
                    "INSERT INTO stores (nome, ativo, criado_em) VALUES (?, 1, ?)",
                    (nome, agora_iso()),
                )
                conn.commit()
            flash("Loja cadastrada com sucesso.", "success")
        except sqlite3.IntegrityError:
            flash("Essa loja já existe.", "error")
        return redirect(url_for("lojas"))

    stores = query_all("SELECT * FROM stores ORDER BY ativo DESC, nome")
    return render_template("lojas.html", title="Lojas", stores=stores)


@app.post("/lojas/<int:store_id>/alternar")
@login_required
@module_required("lojas")
@roles_required("admin")
def alternar_loja(store_id: int) -> Response:
    with closing(get_conn()) as conn:
        row = conn.execute("SELECT ativo FROM stores WHERE id = ?", (store_id,)).fetchone()
        if row is None:
            flash("Loja não encontrada.", "error")
            return redirect(url_for("lojas"))
        conn.execute("UPDATE stores SET ativo = ? WHERE id = ?", (0 if row["ativo"] else 1, store_id))
        conn.commit()
    flash("Loja atualizada.", "success")
    return redirect(url_for("lojas"))


@app.post("/lojas/<int:store_id>/excluir")
@login_required
@module_required("lojas")
@roles_required("admin")
def excluir_loja(store_id: int) -> Response:
    with closing(get_conn()) as conn:
        loja = conn.execute("SELECT * FROM stores WHERE id = ?", (store_id,)).fetchone()
        if loja is None:
            flash("Loja não encontrada.", "error")
            return redirect(url_for("lojas"))

        usos = conn.execute("SELECT COUNT(*) AS c FROM separations WHERE store_id = ?", (store_id,)).fetchone()["c"]
        if int(usos or 0) > 0:
            flash("Essa loja já foi usada em separações. Desative em vez de excluir.", "error")
            return redirect(url_for("lojas"))

        conn.execute("DELETE FROM stores WHERE id = ?", (store_id,))
        conn.commit()

    flash("Loja excluída com sucesso.", "success")
    return redirect(url_for("lojas"))






MCP_LABELS: dict[str, str] = {
    "id": "ID",
    "chave": "Campo",
    "valor": "Valor",
    "codigo": "Código",
    "codigo_barras": "Código de barras",
    "linha": "Linha/Categoria",
    "amostra": "Exemplo",
    "descricao": "Descrição",
    "nome": "Nome",
    "tipo": "Tipo",
    "status": "Status",
    "ativo": "Ativo",
    "fator_formatado": "Emb.",
    "fator_embalagem": "Fator",
    "quantidade": "Quantidade",
    "quantidade_formatada": "Quantidade",
    "quantidade_atual": "Qtd. atual",
    "quantidade_pedida": "Qtd. pedida",
    "quantidade_separada": "Qtd. separada",
    "quantidade_conferida": "Qtd. conferida",
    "custo_unitario_formatado": "Custo unit.",
    "valor_total_estimado_formatado": "Valor estimado",
    "quantidade_total_formatada": "Qtd. total",
    "valor_total_estimado": "Valor estimado",
    "quantidade_total": "Qtd. total",
    "itens_zerados": "Zerados",
    "itens_estoque_baixo_ate_10": "Estoque baixo",
    "menor_estoque_formatado": "Menor estoque",
    "maior_estoque_formatado": "Maior estoque",
    "valor_estimado_formatado": "Valor estimado",
    "total_itens": "Itens",
    "total_separacoes": "Separações",
    "loja_nome": "Loja",
    "lojas": "Lojas",
    "lote_codigo": "Lote",
    "lote_nome": "Nome do lote",
    "data_referencia": "Data",
    "status_encontrados": "Status",
    "total_quantidade_pedida": "Total pedido",
    "total_quantidade_separada": "Total separado",
    "ultima_atualizacao": "Última atualização",
    "atualizado_em": "Atualizado em",
    "criado_em": "Criado em",
    "criado_por_nome": "Usuário",
    "observacao": "Observação",
    "referencia_tipo": "Referência",
    "referencia_id": "Ref. ID",
    "encontrado": "Encontrado",
    "mensagem": "Mensagem",
    "db_path": "Banco",
    "origem": "Origem",
}

MCP_COLUMN_PREFERENCE: dict[str, list[str]] = {
    "produtos": [
        "linha",
        "codigo",
        "codigo_barras",
        "descricao",
        "quantidade_formatada",
        "fator_formatado",
        "custo_unitario_formatado",
        "valor_total_estimado_formatado",
        "atualizado_em",
    ],
    "lojas": ["id", "nome", "ativo", "criado_em"],
    "lotes": [
        "lote_codigo",
        "lote_nome",
        "data_referencia",
        "status_encontrados",
        "total_separacoes",
        "lojas",
        "total_quantidade_pedida",
        "total_quantidade_separada",
        "ultima_atualizacao",
    ],
    "itens_lote": [
        "loja_nome",
        "codigo",
        "descricao",
        "quantidade_pedida",
        "quantidade_separada",
        "quantidade_conferida",
        "status",
        "custo_unitario_ref",
        "atualizado_em",
    ],
    "movimentacoes": [
        "criado_em",
        "tipo",
        "codigo",
        "descricao",
        "quantidade_formatada",
        "criado_por_nome",
        "observacao",
        "referencia_tipo",
    ],
    "categorias": ["linha", "total_itens", "amostra"],
    "resumo_linhas": [
        "linha",
        "total_itens",
        "quantidade_total_formatada",
        "valor_total_estimado_formatado",
        "itens_zerados",
        "itens_estoque_baixo_ate_10",
        "menor_estoque_formatado",
        "maior_estoque_formatado",
        "amostra",
    ],
    "chave_valor": ["chave", "valor"],
    "geral": ["origem", "linha", "codigo", "nome", "descricao", "quantidade_formatada", "status", "criado_em", "atualizado_em"],
}


def _mcp_formatar_valor(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False, default=str)
    return str(valor)


def _mcp_label(chave: str) -> str:
    return MCP_LABELS.get(chave, str(chave).replace("_", " ").strip().capitalize())


def _mcp_detectar_tipo_tabela(resultado: Any, tool: str = "") -> str:
    if isinstance(resultado, dict) and any(isinstance(resultado.get(k), list) for k in ("produtos", "lojas", "lotes", "movimentacoes")):
        return "geral"
    if isinstance(resultado, dict) and isinstance(resultado.get("itens"), list):
        return "itens_lote"
    nome = str(tool or "").casefold()
    if "loja" in nome:
        return "lojas"
    if "lote" in nome:
        return "lotes"
    if "moviment" in nome or "historico" in nome or "histórico" in nome:
        return "movimentacoes"
    if "resumo_estoque_por_linha" in nome or "resumo_linhas" in nome:
        return "resumo_linhas"
    if "categoria" in nome or "linha" in nome:
        return "categorias"
    if "produto" in nome or "estoque" in nome or "baixo" in nome:
        return "produtos"
    return "chave_valor"


def _mcp_numero_seguro(valor: Any) -> float:
    """Converte números vindos do banco ou texto formatado em float para resumos."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    texto = re.sub(r"[^0-9.\-]", "", texto)
    try:
        return float(texto)
    except (TypeError, ValueError):
        return 0.0


def _mcp_resumo_tabela(rows: list[dict[str, Any]], tipo: str = "") -> dict[str, Any]:
    """Cria um resumo simples para exibir na tela e na exportação."""
    resumo: dict[str, Any] = {"total_registros": len(rows)}
    if not rows:
        return resumo

    linhas = {str(row.get("linha") or "").strip() for row in rows if str(row.get("linha") or "").strip()}
    if linhas:
        resumo["linhas_distintas"] = len(linhas)

    quantidade_total = 0.0
    encontrou_quantidade = False
    valor_total = 0.0
    encontrou_valor = False
    zerados = 0
    baixo = 0

    for row in rows:
        if "quantidade_atual" in row:
            qtd = _mcp_numero_seguro(row.get("quantidade_atual"))
            quantidade_total += qtd
            encontrou_quantidade = True
            if qtd <= 0:
                zerados += 1
            if qtd <= 10:
                baixo += 1
        elif "quantidade_total" in row:
            quantidade_total += _mcp_numero_seguro(row.get("quantidade_total"))
            encontrou_quantidade = True

        if "valor_total_estimado" in row:
            valor_total += _mcp_numero_seguro(row.get("valor_total_estimado"))
            encontrou_valor = True

    if encontrou_quantidade:
        resumo["quantidade_total"] = quantidade_total
        resumo["quantidade_total_formatada"] = fmt_num(quantidade_total)
        resumo["itens_zerados"] = zerados
        resumo["itens_estoque_baixo_ate_10"] = baixo
    if encontrou_valor:
        resumo["valor_total_estimado"] = valor_total
        resumo["valor_total_estimado_formatado"] = fmt_money(valor_total)
    return resumo


def _mcp_normalizar_para_tabela(resultado: Any, tool: str = "", titulo: str = "Resultado") -> dict[str, Any]:
    """Converte qualquer retorno do MCP em uma tabela exibível/exportável."""
    tipo = _mcp_detectar_tipo_tabela(resultado, tool)
    rows: list[dict[str, Any]] = []

    if isinstance(resultado, dict) and any(isinstance(resultado.get(k), list) for k in ("produtos", "lojas", "lotes", "movimentacoes")):
        tipo = "geral"
        rows = []
        for origem, chave_lista in [("Produto", "produtos"), ("Loja", "lojas"), ("Lote", "lotes"), ("Movimentação", "movimentacoes")]:
            for item in resultado.get(chave_lista, []) or []:
                if isinstance(item, dict):
                    novo = {"origem": origem}
                    novo.update(item)
                    rows.append(novo)
        titulo = f"Pesquisa geral: {resultado.get('termo') or titulo}"
    elif isinstance(resultado, dict) and isinstance(resultado.get("itens"), list):
        rows = [dict(item) for item in resultado.get("itens", []) if isinstance(item, dict)]
        resumo = resultado.get("resumo") if isinstance(resultado.get("resumo"), dict) else {}
        titulo = f"Lote {resultado.get('lote_codigo') or ''}".strip() or titulo
        if resumo:
            titulo += " - " + ", ".join(
                f"{_mcp_label(k)}: {_mcp_formatar_valor(v)}"
                for k, v in resumo.items()
                if k in {"total_itens", "valor_estimado_formatado", "total_quantidade_pedida", "total_quantidade_separada"}
            )
    elif isinstance(resultado, list):
        rows = [dict(item) if isinstance(item, dict) else {"valor": item} for item in resultado]
    elif isinstance(resultado, dict):
        rows = []
        for chave, valor in resultado.items():
            if isinstance(valor, (dict, list)):
                valor = json.dumps(valor, ensure_ascii=False, default=str)
            rows.append({"chave": _mcp_label(chave), "valor": valor})
        tipo = "chave_valor"
    elif resultado is not None:
        rows = [{"chave": "Resultado", "valor": str(resultado)}]
        tipo = "chave_valor"

    if not rows:
        return {"title": titulo, "columns": [], "labels": {}, "rows": [], "count": 0}

    preferred = MCP_COLUMN_PREFERENCE.get(tipo, [])
    found_keys: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in found_keys:
                found_keys.append(key)

    columns = [key for key in preferred if key in found_keys]
    extras = [key for key in found_keys if key not in columns]
    columns.extend(extras[: max(0, 10 - len(columns))])

    labels = {key: _mcp_label(key) for key in columns}
    normalized_rows = [{key: _mcp_formatar_valor(row.get(key, "")) for key in columns} for row in rows]
    summary = _mcp_resumo_tabela(rows, tipo=tipo)

    return {
        "title": titulo,
        "columns": columns,
        "labels": labels,
        "rows": normalized_rows,
        "count": len(normalized_rows),
        "summary": summary,
    }


def _resumir_resultado_mcp(resultado: Any, limite_linhas: int = 12, tool: str = "") -> str:
    """Transforma o retorno das ferramentas MCP em texto simples para a página de teste."""
    if resultado is None:
        return "Nenhum resultado retornado."

    if isinstance(resultado, dict) and resultado.get("encontrado") is False:
        return str(resultado.get("mensagem") or "Nada encontrado.")

    tabela = _mcp_normalizar_para_tabela(resultado, tool=tool)
    total = tabela.get("count", 0)
    if not total:
        return "Nenhum item encontrado."

    if tabela["columns"] == ["chave", "valor"]:
        linhas = []
        for row in tabela["rows"][:limite_linhas]:
            linhas.append(f"{row.get('chave', '')}: {row.get('valor', '')}")
        if total > limite_linhas:
            linhas.append(f"... e mais {total - limite_linhas} campo(s).")
        return "\n".join(linhas)

    linhas = [f"Encontrei {total} resultado(s)."]
    cols = tabela["columns"][:4]
    for row in tabela["rows"][:limite_linhas]:
        partes = [row.get(col, "") for col in cols if row.get(col, "") != ""]
        linhas.append("- " + " | ".join(partes))
    if total > limite_linhas:
        linhas.append(f"... e mais {total - limite_linhas} item(ns). Use Exportar Excel/PDF para ver tudo.")
    return "\n".join(linhas)




# =========================
# MCP INTELIGENTE - CÁLCULOS SEGUROS
# =========================
_MCP_OPERADORES_CALCULO = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod,
    ast.Pow: operator.pow,
}
_MCP_OPERADORES_UNARIOS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def _mcp_avaliar_no_calculo(node: ast.AST) -> float:
    """Avalia apenas números e operações matemáticas básicas. Não executa código Python."""
    if isinstance(node, ast.Expression):
        return _mcp_avaliar_no_calculo(node.body)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return node.value
    if hasattr(ast, "Num") and isinstance(node, ast.Num):  # compatibilidade
        return node.n
    if isinstance(node, ast.BinOp) and type(node.op) in _MCP_OPERADORES_CALCULO:
        esquerda = _mcp_avaliar_no_calculo(node.left)
        direita = _mcp_avaliar_no_calculo(node.right)
        if isinstance(node.op, ast.Pow) and abs(direita) > 12:
            raise ValueError("Expoente muito alto para cálculo seguro.")
        return _MCP_OPERADORES_CALCULO[type(node.op)](esquerda, direita)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _MCP_OPERADORES_UNARIOS:
        return _MCP_OPERADORES_UNARIOS[type(node.op)](_mcp_avaliar_no_calculo(node.operand))
    raise ValueError("Use apenas números e operadores +, -, *, /, %, //, ** e parênteses.")


def _mcp_normalizar_expressao_calculo(texto: str) -> str:
    bruto = str(texto or "").strip().lower()
    bruto = bruto.replace("×", "*").replace("x", "*").replace("÷", "/")
    bruto = bruto.replace(",", ".")
    trocas = [
        (r"\bcalcule\b", " "),
        (r"\bcalcular\b", " "),
        (r"\bquantos?\s+é\b", " "),
        (r"\bquantos?\s+e\b", " "),
        (r"\bquanto\s+é\b", " "),
        (r"\bquanto\s+e\b", " "),
        (r"\bqual\s+é\s+o\s+resultado\s+de\b", " "),
        (r"\bfaz\b", " "),
        (r"\bvezes\b", " * "),
        (r"\bmais\b", " + "),
        (r"\bmenos\b", " - "),
        (r"\bdividido\s+por\b", " / "),
        (r"\bdividido\b", " / "),
        (r"\bporcento\b", " /100 "),
    ]
    for padrao, repl in trocas:
        bruto = re.sub(padrao, repl, bruto, flags=re.IGNORECASE)
    bruto = re.sub(r"[^0-9\.\+\-\*\/\%\(\)\s]", " ", bruto)
    bruto = re.sub(r"\s+", "", bruto)
    return bruto.strip()


def _mcp_tentar_calculo(texto: str) -> dict[str, Any] | None:
    """Detecta comandos de cálculo e devolve uma resposta MCP tabular."""
    original = str(texto or "").strip()
    lower = original.casefold()
    parece_calculo = any(p in lower for p in ["calcule", "calcular", "quanto é", "quanto e", "quantos é", "quantos e", "resultado de", "vezes", "dividido"]) or bool(re.fullmatch(r"[\d\s\.,\+\-\*\/\%\(\)]+", original))
    if not parece_calculo:
        return None
    expr = _mcp_normalizar_expressao_calculo(original)
    if not expr or not re.search(r"\d", expr) or not re.search(r"[\+\-\*\/\%]", expr):
        return None
    try:
        resultado = _mcp_avaliar_no_calculo(ast.parse(expr, mode="eval"))
        if isinstance(resultado, float) and resultado.is_integer():
            resultado_fmt = str(int(resultado))
        else:
            resultado_fmt = (f"{resultado:.6f}".rstrip("0").rstrip(".") if isinstance(resultado, float) else str(resultado))
    except Exception as exc:
        return {
            "tool": "calculadora_segura",
            "answer": f"Não consegui calcular com segurança: {exc}",
            "raw": {"expressao": expr, "erro": str(exc)},
            "table": _mcp_normalizar_para_tabela({"expressao": expr, "erro": str(exc)}, "calculadora_segura", "Cálculo"),
            "query": original,
            "linha": "",
            "filtros": {},
            "acoes_sugeridas": [],
        }
    return {
        "tool": "calculadora_segura",
        "answer": f"Resultado: {resultado_fmt}",
        "raw": {"expressao": expr, "resultado": resultado_fmt},
        "table": _mcp_normalizar_para_tabela({"expressao": expr, "resultado": resultado_fmt}, "calculadora_segura", "Cálculo"),
        "query": original,
        "linha": "",
        "filtros": {},
        "acoes_sugeridas": [],
    }


def _mcp_extrair_limite(texto: str, padrao: int = 30, maximo: int = 200) -> int:
    bruto = str(texto or "")
    patterns = [
        r"(?:limite|listar|mostra(?:r)?|trazer|ultimos|últimos)\s+(\d+)",
        r"(\d+)\s+(?:itens|produtos|resultados|linhas|movimenta)",
    ]
    for pattern in patterns:
        match = re.search(pattern, bruto, flags=re.IGNORECASE)
        if match:
            try:
                return max(1, min(int(match.group(1)), maximo))
            except ValueError:
                pass
    return padrao


def _mcp_extrair_primeiro_numero(texto: str, padrao: int = 10) -> int:
    match = re.search(r"\b\d+\b", str(texto or ""))
    if not match:
        return padrao
    try:
        return int(match.group(0))
    except ValueError:
        return padrao


def _mcp_limpar_termo(texto: str) -> str:
    termo = str(texto or "").strip()
    limpar = [
        r"\b(pesquisar|pesquise|buscar|busque|procurar|procure|consultar|consulte|listar|lista|mostre|mostrar|trazer|traga)\b",
        r"\b(produtos?|itens?|estoque|categorias?|linhas?|cadastrados?|com saldo|sem saldo|no sistema|na lista|tudo|geral)\b",
        r"\b(limite|listar|mostrar|trazer|ultimos|últimos)\s+\d+\b",
        r"\b\d+\s+(itens|produtos|resultados|linhas)\b",
    ]
    for pattern in limpar:
        termo = re.sub(pattern, " ", termo, flags=re.IGNORECASE)
    termo = re.sub(r"\s+", " ", termo).strip(" :-")
    return termo.strip()


def _mcp_extrair_codigo_depois_de(palavras: str, texto: str) -> str:
    pattern = rf"(?:{palavras})\s*(?:n[ºo°.]?\s*)?([A-Za-z0-9._/-]+)"
    match = re.search(pattern, texto, flags=re.IGNORECASE)
    return (match.group(1) if match else "").strip()


def _mcp_extrair_linha(texto: str) -> str:
    """Extrai linha/categoria de comandos como: linha pão de alho, categoria frios."""
    bruto = str(texto or "")
    pattern = r"(?:linha|categoria)\s*[:=-]?\s*([A-Za-zÀ-ÿ0-9\s\-/]+)"
    match = re.search(pattern, bruto, flags=re.IGNORECASE)
    if not match:
        return ""

    valor = match.group(1)
    valor = re.split(
        r"\b(?:limite|listar|mostrar|trazer|produto|produtos|itens|estoque|baixo|baixos|movimenta|lote|loja|saldo|com saldo|entre|minimo|mínimo|maximo|máximo|ate|até)\b",
        valor,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    return re.sub(r"\s+", " ", valor).strip(" :-")


def _mcp_remover_linha_do_texto(texto: str) -> str:
    """Remove apenas o trecho da linha/categoria para não atrapalhar o termo de busca."""
    bruto = str(texto or "")
    linha = _mcp_extrair_linha(bruto)
    if not linha:
        return bruto.strip()

    pattern = rf"(?:linha|categoria)\s*[:=-]?\s*{re.escape(linha)}"
    sem_linha = re.sub(pattern, " ", bruto, count=1, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", sem_linha).strip()


def _mcp_filtro_float(filtros: dict[str, Any], chave: str) -> float | None:
    valor = filtros.get(chave)
    if valor in (None, "", "null"):
        return None
    texto = str(valor).strip().replace(",", ".")
    try:
        return float(texto)
    except (TypeError, ValueError):
        return None


def _mcp_filtro_bool(filtros: dict[str, Any], chave: str, padrao: bool = False) -> bool:
    valor = filtros.get(chave)
    if valor in (None, ""):
        return padrao
    if isinstance(valor, bool):
        return valor
    return str(valor).strip().casefold() in {"1", "true", "sim", "s", "yes", "on"}


def _mcp_filtro_texto(filtros: dict[str, Any], chave: str, padrao: str = "") -> str:
    return str(filtros.get(chave) if filtros.get(chave) is not None else padrao).strip()



def _mcp_resposta_chat_natural(texto: str) -> dict[str, Any] | None:
    """Respostas naturais para o assistente flutuante quando a mensagem não é uma consulta do sistema."""
    original = str(texto or "").strip()
    lower = original.casefold()
    if not original:
        return None

    saudacoes = {"oi", "olá", "ola", "bom dia", "boa tarde", "boa noite", "e ai", "e aí", "opa"}
    if lower in saudacoes or any(lower.startswith(s + " ") for s in saudacoes):
        nome = ""
        try:
            nome = (g.user["nome"] or g.user["username"] or "").split()[0]
        except Exception:
            nome = ""
        return {
            "tool": "chat_natural",
            "answer": f"Bom dia{', ' + nome if nome else ''}! Eu sou o assistente MCP/IA. Pode conversar comigo normalmente ou pedir ações do sistema. Ex.: calcule 10+20, estoque baixo, lotes abertos, status do sistema, ou procurar um produto.",
            "raw": {"mensagem": original},
            "table": _mcp_normalizar_para_tabela([], "chat_natural", "Conversa"),
            "query": original,
            "linha": "",
            "filtros": {},
            "acoes_sugeridas": [],
        }

    if any(p in lower for p in ["o que voce faz", "o que você faz", "me ajuda", "ajuda", "como funciona", "o que consegue", "comandos"]):
        return {
            "tool": "chat_natural",
            "answer": "Eu consigo conversar, calcular e consultar o sistema. Posso verificar estoque, produtos, lotes abertos, movimentações, status do sistema e preparar ações como relatório ou lista de balanço. Quando for algo sensível, eu peço sua senha antes.",
            "raw": {"mensagem": original},
            "table": _mcp_normalizar_para_tabela([], "chat_natural", "Ajuda do assistente"),
            "query": original,
            "linha": "",
            "filtros": {},
            "acoes_sugeridas": [],
        }

    if any(p in lower for p in ["obrigado", "obrigada", "valeu", "show", "blz", "beleza"]):
        return {
            "tool": "chat_natural",
            "answer": "Fechado! Quando precisar, é só me chamar por aqui.",
            "raw": {"mensagem": original},
            "table": _mcp_normalizar_para_tabela([], "chat_natural", "Conversa"),
            "query": original,
            "linha": "",
            "filtros": {},
            "acoes_sugeridas": [],
        }

    # Pergunta comum sem intenção clara de consulta: responde como chatbot, sem forçar pesquisa vazia.
    palavras_sistema = [
        "estoque", "produto", "item", "codigo", "código", "barras", "lote", "loja", "movimenta",
        "histórico", "historico", "baixo", "zerado", "saldo", "categoria", "linha", "status", "sistema",
        "listar", "buscar", "pesquisar", "procurar", "relatorio", "relatório", "balanço", "balanco",
    ]
    if not any(p in lower for p in palavras_sistema):
        return {
            "tool": "chat_natural",
            "answer": "Entendi. Eu posso conversar com você por aqui, mas minha inteligência principal está ligada ao sistema. Para eu executar algo real, me peça de forma direta, por exemplo: 'calcule 10+20', 'ver estoque baixo', 'listar lotes abertos' ou 'consultar produto 123'.",
            "raw": {"mensagem": original},
            "table": _mcp_normalizar_para_tabela([], "chat_natural", "Conversa"),
            "query": original,
            "linha": "",
            "filtros": {},
            "acoes_sugeridas": [],
        }
    return None

def _executar_pergunta_mcp(mensagem: str, *, modo: str = "chat", linha: str = "", filtros: dict[str, Any] | None = None) -> dict[str, Any]:
    """Roteador mais tolerante para testar as mesmas consultas expostas no MCP."""
    texto = str(mensagem or "").strip()
    lower = texto.casefold()
    linha_filtro = str(linha or "").strip() or _mcp_extrair_linha(texto)
    texto_sem_linha = _mcp_remover_linha_do_texto(texto) if linha_filtro else texto
    lower_sem_linha = texto_sem_linha.casefold()
    if not texto:
        return {
            "tool": "nenhuma",
            "answer": "Digite uma pergunta, por exemplo: calcule 10*3, estoque baixo, farinha, produto 123, lote ABC ou movimentações do estoque.",
            "raw": [],
            "table": _mcp_normalizar_para_tabela([], "nenhuma"),
        }

    calculo = _mcp_tentar_calculo(texto)
    if calculo is not None:
        return calculo

    conversa = _mcp_resposta_chat_natural(texto)
    if conversa is not None:
        return conversa

    try:
        from . import mcp_server as mcp_tools
    except Exception as exc:
        erro = f"Não consegui carregar o módulo MCP. Verifique se as dependências foram instaladas. Erro: {exc}"
        return {
            "tool": "erro_importacao_mcp",
            "answer": erro,
            "raw": {"mensagem": erro},
            "table": _mcp_normalizar_para_tabela({"mensagem": erro}, "erro_importacao_mcp"),
        }

    filtros = dict(filtros or {})
    tipo_forcado = _mcp_filtro_texto(filtros, "tipo")
    limite = _mcp_extrair_limite(texto_sem_linha, padrao=30)
    limite_filtro = _mcp_filtro_float(filtros, "limite")
    if limite_filtro is not None:
        limite = max(1, min(int(limite_filtro), 200))

    estoque_min = _mcp_filtro_float(filtros, "estoque_min")
    estoque_max = _mcp_filtro_float(filtros, "estoque_max")
    intervalo_match = re.search(r"entre\s+(\d+(?:[\.,]\d+)?)\s+e\s+(\d+(?:[\.,]\d+)?)", lower)
    if intervalo_match and estoque_min is None and estoque_max is None:
        estoque_min = _mcp_numero_seguro(intervalo_match.group(1))
        estoque_max = _mcp_numero_seguro(intervalo_match.group(2))
    if estoque_max is None:
        max_match = re.search(r"(?:saldo|estoque|quantidade|qtd)\s*(?:até|ate|<=|menor(?:\s+que)?|maximo|máximo)\s*(\d+(?:[\.,]\d+)?)", lower)
        if max_match:
            estoque_max = _mcp_numero_seguro(max_match.group(1))
    if estoque_min is None:
        min_match = re.search(r"(?:saldo|estoque|quantidade|qtd)\s*(?:>=|maior(?:\s+que)?|minimo|mínimo|acima\s+de)\s*(\d+(?:[\.,]\d+)?)", lower)
        if min_match:
            estoque_min = _mcp_numero_seguro(min_match.group(1))
    somente_com_saldo = _mcp_filtro_bool(filtros, "somente_com_saldo", "com saldo" in lower)
    ordenar = _mcp_filtro_texto(filtros, "ordenar", "descricao")
    direcao = _mcp_filtro_texto(filtros, "direcao", "asc")
    termo_filtro = _mcp_filtro_texto(filtros, "termo")
    titulo = "Resultado da consulta"

    if tipo_forcado == "resumo_linhas" or (
        any(p in lower for p in ["resumo por linha", "resumo por categoria", "por linha", "por categoria"])
        and any(p in lower for p in ["estoque", "resumo", "categoria", "linha"])
    ):
        tool = "resumo_estoque_por_linha"
        resultado = mcp_tools.resumo_estoque_por_linha(linha=linha_filtro, limite=limite)
        titulo = "Resumo do estoque por linha/categoria"

    elif tipo_forcado in {"avancado", "busca_avancada"}:
        tool = "buscar_produtos_avancado"
        termo = termo_filtro or _mcp_limpar_termo(texto_sem_linha)
        resultado = mcp_tools.buscar_produtos_avancado(
            termo=termo,
            linha=linha_filtro,
            estoque_min=estoque_min,
            estoque_max=estoque_max,
            somente_com_saldo=somente_com_saldo,
            ordenar=ordenar,
            direcao=direcao,
            limite=limite,
        )
        titulo = f"Busca avançada: {termo or 'todos os produtos'}"

    elif any(palavra in lower for palavra in ["categorias", "categoria", "linhas", "linha"]) and not linha_filtro and not any(p in lower for p in ["produto", "produtos", "item", "itens", "estoque", "baixo", "pesquis", "buscar", "procurar"]):
        tool = "listar_categorias_linha"
        termo = _mcp_limpar_termo(texto_sem_linha)
        resultado = mcp_tools.listar_categorias_linha(termo=termo, limite=min(limite, 80))
        titulo = "Categorias/Linhas encontradas"

    elif any(palavra in lower for palavra in ["status", "resumo", "sistema", "geral", "painel"]) and not any(p in lower for p in ["pesquis", "buscar", "procurar"]):
        tool = "status_sistema"
        resultado = mcp_tools.status_sistema()
        titulo = "Status do sistema"

    elif any(palavra in lower for palavra in ["baixo", "baixa", "acabando", "critico", "crítico", "pouco", "repor", "reposição", "reposicao", "zerado", "sem saldo"]):
        tool = "listar_estoque_baixo"
        limite_quantidade = 0 if ("zerado" in lower or "sem saldo" in lower) else _mcp_extrair_primeiro_numero(texto, 10)
        resultado = mcp_tools.listar_estoque_baixo(limite_quantidade=limite_quantidade, limite=limite, linha=linha_filtro)
        titulo = f"Estoque baixo até {limite_quantidade}"

    elif "loja" in lower:
        tool = "listar_lojas"
        resultado = mcp_tools.listar_lojas(ativas=("inativa" not in lower and "todas" not in lower), limite=min(limite, 200))
        titulo = "Lojas cadastradas"

    elif "movimenta" in lower or "histórico" in lower or "historico" in lower:
        tool = "listar_movimentacoes_estoque"
        codigo = _mcp_extrair_codigo_depois_de(r"codigo|código|produto|item|barras", texto_sem_linha)
        tipo = ""
        if "entrada" in lower:
            tipo = "ENTRADA_INICIAL"
        elif "saida" in lower or "saída" in lower:
            tipo = "SAIDA_SEPARACAO"
        elif "ajuste" in lower:
            tipo = "AJUSTE_MANUAL"
        elif "recont" in lower or "balan" in lower:
            tipo = "RECONTAGEM"
        elif "remov" in lower or "exclu" in lower:
            tipo = "REMOVIDO_ESTOQUE"
        resultado = mcp_tools.listar_movimentacoes_estoque(codigo=codigo, tipo=tipo, limite=limite)
        titulo = "Movimentações de estoque"

    elif "lote" in lower:
        codigo = _mcp_extrair_codigo_depois_de(r"lote", texto_sem_linha)
        if codigo and codigo.casefold() not in {"aberto", "abertos", "todos", "todas"}:
            tool = "consultar_lote"
            resultado = mcp_tools.consultar_lote(codigo)
            titulo = f"Consulta do lote {codigo}"
        else:
            tool = "listar_lotes_abertos"
            resultado = mcp_tools.listar_lotes_abertos(limite=limite)
            titulo = "Lotes abertos"

    elif any(palavra in lower for palavra in ["produto", "codigo", "código", "barras", "item"]):
        codigo = _mcp_extrair_codigo_depois_de(r"produto|codigo|código|barras|item", texto_sem_linha)
        termo = _mcp_limpar_termo(texto_sem_linha)
        somente_com_saldo = somente_com_saldo or ("com saldo" in lower)
        if codigo and not any(palavra in lower for palavra in ["nome", "descrição", "descricao", "contem", "contém", "parecido"]):
            tool = "consultar_produto"
            resultado = mcp_tools.consultar_produto(codigo)
            titulo = f"Consulta do produto {codigo}"
            if isinstance(resultado, dict) and resultado.get("encontrado") is False:
                tool = "listar_produtos_estoque"
                resultado = mcp_tools.listar_produtos_estoque(termo=codigo, somente_com_saldo=somente_com_saldo, limite=limite, linha=linha_filtro)
                titulo = f"Produtos encontrados: {codigo}"
        else:
            if estoque_min is not None or estoque_max is not None or ordenar != "descricao" or direcao != "asc":
                tool = "buscar_produtos_avancado"
                resultado = mcp_tools.buscar_produtos_avancado(
                    termo=termo,
                    linha=linha_filtro,
                    estoque_min=estoque_min,
                    estoque_max=estoque_max,
                    somente_com_saldo=somente_com_saldo,
                    ordenar=ordenar,
                    direcao=direcao,
                    limite=limite,
                )
            else:
                tool = "listar_produtos_estoque"
                resultado = mcp_tools.listar_produtos_estoque(termo=termo, somente_com_saldo=somente_com_saldo, limite=limite, linha=linha_filtro)
            titulo = f"Produtos encontrados: {termo or 'todos'}"

    elif "estoque" in lower or "listar" in lower or "lista" in lower or "pesquis" in lower or "buscar" in lower or "procur" in lower:
        termo = termo_filtro or _mcp_limpar_termo(texto_sem_linha)
        if estoque_min is not None or estoque_max is not None or ordenar != "descricao" or direcao != "asc" or somente_com_saldo:
            tool = "buscar_produtos_avancado"
            resultado = mcp_tools.buscar_produtos_avancado(
                termo=termo,
                linha=linha_filtro,
                estoque_min=estoque_min,
                estoque_max=estoque_max,
                somente_com_saldo=somente_com_saldo,
                ordenar=ordenar,
                direcao=direcao,
                limite=limite,
            )
        else:
            tool = "listar_produtos_estoque"
            resultado = mcp_tools.listar_produtos_estoque(termo=termo, somente_com_saldo=False, limite=limite, linha=linha_filtro)
        titulo = f"Estoque encontrado: {termo or 'todos'}"

    else:
        # Fallback mais útil: trata a mensagem como termo de busca do estoque.
        tool = "listar_produtos_estoque"
        termo = _mcp_limpar_termo(texto_sem_linha) or texto
        resultado = mcp_tools.listar_produtos_estoque(termo=termo, somente_com_saldo=False, limite=limite, linha=linha_filtro)
        titulo = f"Pesquisa por: {termo}"
        if not resultado and hasattr(mcp_tools, "pesquisar_geral"):
            tool = "pesquisar_geral"
            resultado = mcp_tools.pesquisar_geral(termo=termo, limite=min(limite, 50), linha=linha_filtro)
            titulo = f"Pesquisa geral por: {termo}"
        if not resultado:
            return {
                "tool": "chat_natural",
                "answer": f"Não encontrei nada para '{termo}'. Posso tentar de outro jeito se você mandar o código, parte da descrição, lote ou loja.",
                "raw": [],
                "table": _mcp_normalizar_para_tabela([], "chat_natural", "Conversa"),
                "query": texto,
                "linha": linha_filtro,
                "filtros": filtros,
                "acoes_sugeridas": [],
            }

    if linha_filtro and tool in {"listar_produtos_estoque", "listar_estoque_baixo", "pesquisar_geral", "buscar_produtos_avancado", "resumo_estoque_por_linha"}:
        titulo = f"{titulo} | Linha: {linha_filtro}"

    tabela = _mcp_normalizar_para_tabela(resultado, tool=tool, titulo=titulo)
    resposta = {
        "tool": tool,
        "answer": _resumir_resultado_mcp(resultado, tool=tool),
        "raw": resultado,
        "table": tabela,
        "query": texto,
        "linha": linha_filtro,
        "filtros": filtros,
    }
    resposta["acoes_sugeridas"] = _mcp_sugerir_acoes(resposta)
    return resposta


def _mcp_total_registros_resposta(resposta: dict[str, Any]) -> int:
    table = resposta.get("table") or {}
    rows = table.get("rows") or []
    try:
        return int(len(rows))
    except TypeError:
        return 0


def _mcp_sugerir_acoes(resposta: dict[str, Any]) -> list[dict[str, str]]:
    tool = str(resposta.get("tool") or "")
    total = _mcp_total_registros_resposta(resposta)
    if total <= 0:
        return []
    acoes: list[dict[str, str]] = []
    if tool in {"listar_produtos_estoque", "listar_estoque_baixo", "buscar_produtos_avancado", "pesquisar_geral", "resumo_estoque_por_linha"}:
        acoes.append({"type": "criar_lista_balanco", "label": "Preparar lista de balanço"})
        acoes.append({"type": "gerar_relatorio_gerencial", "label": "Preparar relatório gerencial"})
    if tool in {"listar_lotes_abertos", "consultar_lote"}:
        acoes.append({"type": "revisar_lote", "label": "Preparar revisão do lote"})
    acoes.append({"type": "exportar_excel", "label": "Exportar Excel"})
    acoes.append({"type": "exportar_pdf", "label": "Exportar PDF"})
    return acoes




def _mcp_senha_atual_valida(password: str) -> bool:
    if g.user is None:
        return False
    try:
        return check_password_hash(g.user["password_hash"], password or "")
    except Exception:
        return False


def _mcp_acao_sensivel(action_type: str) -> bool:
    return str(action_type or "").strip().lower() in {
        "criar_lista_balanco",
        "gerar_relatorio_gerencial",
        "revisar_lote",
    }


def _mcp_executar_criar_lista_balanco(pergunta: str, linha: str = "", filtros: dict[str, Any] | None = None) -> dict[str, Any]:
    if not user_has_access(g.user, "balanco") and not user_is_admin(g.user):
        raise ValueError("Seu usuário não tem permissão para criar balanço.")
    resposta = _executar_pergunta_mcp(pergunta or "listar estoque limite 50", linha=linha, filtros=filtros or {})
    raw = resposta.get("raw")
    produtos = raw if isinstance(raw, list) else []
    produtos = [p for p in produtos if isinstance(p, dict) and (p.get("id") or p.get("codigo"))]
    if not produtos:
        raise ValueError("A consulta não trouxe produtos suficientes para criar uma lista de balanço.")
    agora = agora_iso()
    titulo = f"MCP - {str(pergunta or 'Lista de balanço')[:80]}"
    observacao = "Criado pelo assistente MCP. Revise antes de confirmar qualquer atualização de estoque."
    with closing(get_conn()) as conn:
        cur = conn.execute(
            "INSERT INTO balance_counts (titulo, observacao, status, criado_por, criado_em) VALUES (?, ?, 'ABERTO', ?, ?)",
            (titulo, observacao, g.user["id"], agora),
        )
        balance_id = int(cur.lastrowid)
        inseridos = 0
        for prod in produtos[:200]:
            item = None
            if prod.get("id"):
                item = conn.execute("SELECT * FROM stock_items WHERE id = ? AND ativo = 1", (prod.get("id"),)).fetchone()
            if item is None and prod.get("codigo"):
                item = conn.execute("SELECT * FROM stock_items WHERE ativo = 1 AND (codigo = ? OR codigo_barras = ?) LIMIT 1", (prod.get("codigo"), prod.get("codigo"))).fetchone()
            if item is None:
                continue
            sistema = float(item["quantidade_atual"] or 0)
            conn.execute(
                """
                INSERT OR IGNORE INTO balance_count_items
                (balance_count_id, stock_item_id, codigo, descricao, linha_erp, quantidade_sistema, quantidade_contada, delta, custo_unitario, criado_em, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?)
                """,
                (balance_id, item["id"], item["codigo"], item["descricao"], item["linha_erp"], sistema, -sistema, item["custo_unitario"] or 0, agora, agora),
            )
            inseridos += 1
        conn.commit()
    registrar_auditoria("mcp_executar_criar_balanco", "balance_counts", str(balance_id), {"pergunta": pergunta[:300], "itens": inseridos})
    return {
        "ok": True,
        "mensagem": f"Balanço criado com {inseridos} item(ns).",
        "redirect_url": url_for("detalhe_balanco", balance_id=balance_id),
        "balance_id": balance_id,
        "itens": inseridos,
    }


def _mcp_executar_acao_real(action_type: str, pergunta: str, linha: str = "", filtros: dict[str, Any] | None = None) -> dict[str, Any]:
    action_type = str(action_type or "").strip().lower()
    if action_type == "criar_lista_balanco":
        return _mcp_executar_criar_lista_balanco(pergunta, linha=linha, filtros=filtros)
    if action_type == "gerar_relatorio_gerencial":
        if not user_has_access(g.user, "relatorios") and not user_is_admin(g.user):
            raise ValueError("Seu usuário não tem permissão para abrir relatórios.")
        registrar_auditoria("mcp_abrir_relatorio", "mcp", "relatorio_gerencial", {"pergunta": pergunta[:300]})
        return {"ok": True, "mensagem": "Relatório gerencial liberado.", "redirect_url": url_for("relatorio_gerencial")}
    if action_type == "revisar_lote":
        if not user_has_access(g.user, "lotes") and not user_is_admin(g.user):
            raise ValueError("Seu usuário não tem permissão para revisar lotes.")
        registrar_auditoria("mcp_abrir_lotes", "mcp", "lotes", {"pergunta": pergunta[:300]})
        return {"ok": True, "mensagem": "Tela de lotes liberada.", "redirect_url": url_for("listar_lotes")}
    if action_type in {"exportar_excel", "exportar_pdf"}:
        formato = "excel" if action_type == "exportar_excel" else "pdf"
        registrar_auditoria("mcp_exportacao_orientada", "mcp", formato, {"pergunta": pergunta[:300]})
        return {"ok": True, "mensagem": f"Para baixar em {formato.upper()}, use o botão de exportação da página MCP depois da consulta."}
    raise ValueError("Esta ação ainda não possui execução real segura.")

def _mcp_contexto_permitido(contexto: str) -> bool:
    contexto = str(contexto or "mcp_teste").strip().lower()
    mapa = {
        "painel": "painel",
        "estoque": "estoque",
        "lotes": "lotes",
        "relatorios": "relatorios",
        "mcp_teste": "mcp_teste",
    "comunicacao": "comunicacao",
    }
    modulo = mapa.get(contexto, "mcp_teste")
    if g.user is None:
        return False
    if user_is_admin(g.user):
        return True
    return user_has_access(g.user, modulo)


def _mcp_registrar_historico(
    *,
    contexto: str,
    pergunta: str,
    linha: str = "",
    filtros: dict[str, Any] | None = None,
    resposta: dict[str, Any] | None = None,
    exportado_formato: str = "",
) -> None:
    try:
        user_id = g.user["id"] if g.user is not None else None
        total = _mcp_total_registros_resposta(resposta or {}) if resposta else 0
        tool = str((resposta or {}).get("tool") or "")
        with closing(get_conn()) as conn:
            conn.execute(
                """
                INSERT INTO mcp_query_history
                (user_id, contexto, pergunta, linha, filtros, tool, total_registros, exportado_formato, criado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    str(contexto or "mcp_teste")[:60],
                    str(pergunta or "")[:800],
                    str(linha or "")[:120],
                    json.dumps(filtros or {}, ensure_ascii=False),
                    tool[:120],
                    total,
                    str(exportado_formato or "")[:20],
                    agora_iso(),
                ),
            )
            conn.commit()
    except Exception:
        # Histórico não pode quebrar a consulta principal.
        return


def _mcp_lista_consultas_salvas(contexto: str = "") -> list[dict[str, Any]]:
    contexto = str(contexto or "").strip().lower()
    user_id = g.user["id"] if g.user is not None else None
    params: list[Any] = [user_id]
    where = "(q.user_id = ? OR q.publico = 1)"
    if contexto:
        where += " AND q.contexto = ?"
        params.append(contexto)
    rows = query_all(
        f"""
        SELECT q.*, u.nome AS usuario_nome
        FROM mcp_saved_queries q
        LEFT JOIN users u ON u.id = q.user_id
        WHERE {where}
        ORDER BY q.criado_em DESC, q.id DESC
        LIMIT 80
        """,
        params,
    )
    result: list[dict[str, Any]] = []
    for row in rows:
        try:
            filtros = json.loads(row["filtros"] or "{}")
        except json.JSONDecodeError:
            filtros = {}
        result.append({
            "id": row["id"],
            "titulo": row["titulo"],
            "pergunta": row["pergunta"],
            "contexto": row["contexto"],
            "linha": row["linha"] or "",
            "filtros": filtros,
            "publico": bool(row["publico"]),
            "usuario_nome": row["usuario_nome"] or "Sistema",
            "criado_em": row["criado_em"],
            "pode_remover": bool(user_id and (row["user_id"] == user_id or user_is_admin(g.user))),
        })
    return result


def _mcp_lista_historico(contexto: str = "", limite: int = 20) -> list[dict[str, Any]]:
    contexto = str(contexto or "").strip().lower()
    limite_seguro = max(1, min(int(limite or 20), 80))
    params: list[Any] = []
    where_parts: list[str] = []
    if not user_is_admin(g.user):
        where_parts.append("h.user_id = ?")
        params.append(g.user["id"])
    if contexto:
        where_parts.append("h.contexto = ?")
        params.append(contexto)
    where = "WHERE " + " AND ".join(where_parts) if where_parts else ""
    rows = query_all(
        f"""
        SELECT h.*, u.nome AS usuario_nome
        FROM mcp_query_history h
        LEFT JOIN users u ON u.id = h.user_id
        {where}
        ORDER BY h.criado_em DESC, h.id DESC
        LIMIT {limite_seguro}
        """,
        params,
    )
    historico: list[dict[str, Any]] = []
    for row in rows:
        try:
            filtros = json.loads(row["filtros"] or "{}")
        except json.JSONDecodeError:
            filtros = {}
        historico.append({
            "id": row["id"],
            "contexto": row["contexto"],
            "pergunta": row["pergunta"],
            "linha": row["linha"] or "",
            "filtros": filtros,
            "tool": row["tool"] or "",
            "total_registros": row["total_registros"],
            "exportado_formato": row["exportado_formato"] or "",
            "usuario_nome": row["usuario_nome"] or "Sistema",
            "criado_em": row["criado_em"],
        })
    return historico


def _mcp_criar_excel(resposta: dict[str, Any]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    table = resposta.get("table") or {}
    columns = list(table.get("columns") or [])
    labels = dict(table.get("labels") or {})
    rows = list(table.get("rows") or [])
    title = str(table.get("title") or "Resultado MCP")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Consulta: {resposta.get('query', '')}"
    ws["A3"] = f"Gerado em: {agora_br()}"
    ws["A4"] = f"Ferramenta: {resposta.get('tool', '')}"
    if resposta.get("linha"):
        ws["A5"] = f"Linha/Categoria: {resposta.get('linha', '')}"
    if resposta.get("filtros"):
        ws["A6"] = "Filtros: " + json.dumps(resposta.get("filtros") or {}, ensure_ascii=False)

    header_row = 8
    if not columns:
        ws.cell(row=header_row, column=1, value="Resultado")
        ws.cell(row=header_row + 1, column=1, value=resposta.get("answer", "Sem resultado."))
    else:
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=labels.get(col, col))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5D2E")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, start=header_row + 1):
            for col_idx, col in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))

        for col_idx, col in enumerate(columns, start=1):
            values = [labels.get(col, col)] + [str(row.get(col, "")) for row in rows[:200]]
            width = min(max(len(v) for v in values) + 2, 55)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + max(1, len(rows))}"

    summary = dict(table.get("summary") or {})
    if summary:
        ws_resumo = wb.create_sheet("Resumo")
        ws_resumo["A1"] = "Resumo da consulta"
        ws_resumo["A1"].font = Font(bold=True, size=14)
        ws_resumo["A3"] = "Campo"
        ws_resumo["B3"] = "Valor"
        ws_resumo["A3"].font = Font(bold=True, color="FFFFFF")
        ws_resumo["B3"].font = Font(bold=True, color="FFFFFF")
        ws_resumo["A3"].fill = PatternFill("solid", fgColor="2F5D2E")
        ws_resumo["B3"].fill = PatternFill("solid", fgColor="2F5D2E")
        for idx, (key, value) in enumerate(summary.items(), start=4):
            ws_resumo.cell(row=idx, column=1, value=_mcp_label(str(key)))
            ws_resumo.cell(row=idx, column=2, value=_mcp_formatar_valor(value))
        ws_resumo.column_dimensions["A"].width = 32
        ws_resumo.column_dimensions["B"].width = 28

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _mcp_criar_pdf(resposta: dict[str, Any]) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    table_data = resposta.get("table") or {}
    columns = list(table_data.get("columns") or [])
    labels = dict(table_data.get("labels") or {})
    rows = list(table_data.get("rows") or [])
    title = str(table_data.get("title") or "Resultado MCP")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Consulta: {_mcp_formatar_valor(resposta.get('query', ''))}", styles["Normal"]),
        Paragraph(f"Gerado em: {agora_br()} | Ferramenta: {_mcp_formatar_valor(resposta.get('tool', ''))}", styles["Normal"]),
    ]
    if resposta.get("linha"):
        story.append(Paragraph(f"Linha/Categoria: {_mcp_formatar_valor(resposta.get('linha'))}", styles["Normal"]))

    summary = dict(table_data.get("summary") or {})
    if summary:
        resumo_linhas = [[Paragraph("Resumo", styles["BodyText"]), Paragraph("Valor", styles["BodyText"])]]
        for key in ["total_registros", "linhas_distintas", "quantidade_total_formatada", "valor_total_estimado_formatado", "itens_zerados", "itens_estoque_baixo_ate_10"]:
            if key in summary:
                resumo_linhas.append([Paragraph(_mcp_label(key), styles["BodyText"]), Paragraph(_mcp_formatar_valor(summary.get(key)), styles["BodyText"])])
        resumo_table = Table(resumo_linhas, repeatRows=1)
        resumo_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5D2E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8D0C2")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        story.append(Spacer(1, 0.25 * cm))
        story.append(resumo_table)

    story.append(Spacer(1, 0.35 * cm))

    if not columns:
        story.append(Paragraph(_mcp_formatar_valor(resposta.get("answer", "Sem resultado.")), styles["Normal"]))
    else:
        safe_columns = columns[:8]
        header = [Paragraph(labels.get(col, col), styles["BodyText"]) for col in safe_columns]
        body = [header]
        for row in rows[:120]:
            body.append([Paragraph(_mcp_formatar_valor(row.get(col, ""))[:350], styles["BodyText"]) for col in safe_columns])

        pdf_table = Table(body, repeatRows=1)
        pdf_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5D2E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8D0C2")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF8F3")]),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(pdf_table)
        if len(rows) > 120:
            story.append(Spacer(1, 0.25 * cm))
            story.append(Paragraph(f"O PDF mostra os primeiros 120 registros. Para todos os dados, use Exportar Excel.", styles["Italic"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


@app.route("/api/mcp/contexto", methods=["GET"])
@login_required
def api_mcp_contexto() -> Response:
    contexto = str(request.args.get("contexto") or "mcp_teste").strip().lower()
    if not _mcp_contexto_permitido(contexto):
        return jsonify({"error": "Você não tem permissão para este contexto."}), 403
    try:
        from .intelligence import sugestoes_contextuais
        dados = sugestoes_contextuais(contexto)
    except Exception:
        dados = {"contexto": contexto, "label": "MCP/IA", "help": "Consulta inteligente do sistema.", "sugestoes": ["status do sistema"]}
    dados["consultas_salvas"] = _mcp_lista_consultas_salvas(contexto)
    dados["historico"] = _mcp_lista_historico(contexto, limite=12)
    return jsonify(dados)


@app.route("/api/mcp/consultas-salvas", methods=["POST"])
@login_required
def api_mcp_salvar_consulta() -> Response:
    payload = request.get_json(silent=True) or {}
    contexto = str(payload.get("contexto") or "mcp_teste").strip().lower()
    titulo = str(payload.get("titulo") or "").strip()
    pergunta = str(payload.get("pergunta") or "").strip()
    linha = str(payload.get("linha") or "").strip()
    filtros = payload.get("filtros") if isinstance(payload.get("filtros"), dict) else {}
    publico = bool(payload.get("publico")) and user_is_admin(g.user)
    if not _mcp_contexto_permitido(contexto):
        return jsonify({"error": "Você não tem permissão para salvar este contexto."}), 403
    if not titulo or not pergunta:
        return jsonify({"error": "Informe título e pergunta."}), 400
    with closing(get_conn()) as conn:
        conn.execute(
            """
            INSERT INTO mcp_saved_queries (user_id, titulo, pergunta, contexto, linha, filtros, publico, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (g.user["id"], titulo[:160], pergunta[:800], contexto[:60], linha[:120], json.dumps(filtros, ensure_ascii=False), 1 if publico else 0, agora_iso()),
        )
        conn.commit()
    return jsonify({"ok": True, "consultas_salvas": _mcp_lista_consultas_salvas(contexto)})


@app.route("/api/mcp/consultas-salvas/<int:consulta_id>/excluir", methods=["POST", "DELETE"])
@login_required
def api_mcp_excluir_consulta(consulta_id: int) -> Response:
    row = query_one("SELECT * FROM mcp_saved_queries WHERE id = ?", (consulta_id,))
    if row is None:
        return jsonify({"error": "Consulta não encontrada."}), 404
    if not user_is_admin(g.user) and row["user_id"] != g.user["id"]:
        return jsonify({"error": "Você não tem permissão para excluir esta consulta."}), 403
    contexto = row["contexto"] or "mcp_teste"
    with closing(get_conn()) as conn:
        conn.execute("DELETE FROM mcp_saved_queries WHERE id = ?", (consulta_id,))
        conn.commit()
    return jsonify({"ok": True, "consultas_salvas": _mcp_lista_consultas_salvas(contexto)})


CODE_EDITOR_ALLOWED_EXTENSIONS = {
    ".py", ".html", ".css", ".js", ".json", ".md", ".txt", ".yml", ".yaml",
    ".toml", ".ini", ".cfg", ".bat", ".sh", ".gitignore", ".procfile",
}
CODE_EDITOR_BLOCKED_DIRS = {".git", "__pycache__", ".pytest_cache", ".mypy_cache", ".venv", "venv", "node_modules", ".code_backups"}
CODE_EDITOR_BLOCKED_FILES = {"dados.db", ".env", "*.sqlite", "*.sqlite3"}
CODE_EDITOR_MAX_FILE_SIZE = 1024 * 1024


def _code_editor_root() -> str:
    root = os.environ.get("CODE_EDITOR_ROOT", BASE_DIR).strip() or BASE_DIR
    return os.path.realpath(root)


def _code_editor_relpath(abs_path: str) -> str:
    rel = os.path.relpath(abs_path, _code_editor_root())
    if rel == ".":
        return ""
    return rel.replace("\\", "/")


def _code_editor_safe_path(path_value: Any = "") -> str:
    root = _code_editor_root()
    rel = str(path_value or "").replace("\\", "/").lstrip("/")
    abs_path = os.path.realpath(os.path.join(root, rel))
    if abs_path != root and not abs_path.startswith(root + os.sep):
        raise ValueError("Caminho fora do projeto bloqueado.")
    return abs_path


def _code_editor_extension(path_value: str) -> str:
    name = os.path.basename(path_value).casefold()
    if name in {".gitignore", "procfile"}:
        return ".gitignore" if name == ".gitignore" else ".procfile"
    return os.path.splitext(name)[1]


def _code_editor_is_blocked(path_value: str) -> bool:
    rel = _code_editor_relpath(path_value)
    parts = {part.casefold() for part in rel.split("/") if part}
    if parts & CODE_EDITOR_BLOCKED_DIRS:
        return True
    name = os.path.basename(path_value).casefold()
    if name in {"dados.db", ".env"}:
        return True
    if name.endswith((".db", ".sqlite", ".sqlite3", ".pyc", ".pyo", ".exe", ".dll", ".so")):
        return True
    return False


def _code_editor_is_editable(path_value: str) -> bool:
    if not os.path.isfile(path_value) or _code_editor_is_blocked(path_value):
        return False
    if os.path.getsize(path_value) > CODE_EDITOR_MAX_FILE_SIZE:
        return False
    return _code_editor_extension(path_value) in CODE_EDITOR_ALLOWED_EXTENSIONS


def _code_editor_list_dir(path_value: str = "") -> dict[str, Any]:
    abs_dir = _code_editor_safe_path(path_value)
    if not os.path.isdir(abs_dir):
        abs_dir = os.path.dirname(abs_dir)
    entries: list[dict[str, Any]] = []
    for name in sorted(os.listdir(abs_dir), key=lambda item: (not os.path.isdir(os.path.join(abs_dir, item)), item.casefold())):
        if name.casefold() in CODE_EDITOR_BLOCKED_DIRS:
            continue
        full = os.path.join(abs_dir, name)
        if _code_editor_is_blocked(full):
            continue
        try:
            stat = os.stat(full)
        except OSError:
            continue
        entries.append({
            "name": name,
            "path": _code_editor_relpath(full),
            "is_dir": os.path.isdir(full),
            "editable": _code_editor_is_editable(full),
            "size": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M"),
        })
    current_rel = _code_editor_relpath(abs_dir)
    breadcrumb = []
    acumulado = ""
    for part in [p for p in current_rel.split("/") if p]:
        acumulado = f"{acumulado}/{part}".strip("/")
        breadcrumb.append({"name": part, "path": acumulado})
    return {"current": current_rel, "entries": entries, "breadcrumb": breadcrumb}


def _code_editor_read_file(path_value: str) -> dict[str, Any]:
    abs_file = _code_editor_safe_path(path_value)
    if not _code_editor_is_editable(abs_file):
        raise ValueError("Arquivo bloqueado, muito grande ou com extensão não permitida.")
    for encoding in ("utf-8", "latin-1"):
        try:
            content = Path(abs_file).read_text(encoding=encoding)
            return {"path": _code_editor_relpath(abs_file), "content": content, "encoding": encoding, "extension": _code_editor_extension(abs_file)}
        except UnicodeDecodeError:
            continue
    raise ValueError("Não foi possível ler o arquivo como texto.")


def _code_editor_check_password(password: str) -> None:
    if get_setting("code_editor_extra_password", "1") != "1":
        return
    if g.user is None or not check_password_hash(g.user["password_hash"], password or ""):
        raise ValueError("Confirme sua senha atual para salvar ou restaurar código.")


def _code_editor_make_backup(abs_file: str) -> str:
    backup_dir = os.path.join(_code_editor_root(), ".code_backups", datetime.now().strftime("%Y%m%d"))
    os.makedirs(backup_dir, exist_ok=True)
    rel = _code_editor_relpath(abs_file).replace("/", "__")
    backup_name = f"{datetime.now().strftime('%H%M%S')}_{uuid.uuid4().hex[:6]}_{rel}"
    backup_path = os.path.join(backup_dir, backup_name)
    shutil.copy2(abs_file, backup_path)
    return _code_editor_relpath(backup_path)


def _code_editor_diff(before: str, after: str, arquivo: str) -> str:
    linhas = list(difflib.unified_diff(
        before.splitlines(),
        after.splitlines(),
        fromfile=f"antes/{arquivo}",
        tofile=f"depois/{arquivo}",
        lineterm="",
        n=3,
    ))
    if len(linhas) > 500:
        linhas = linhas[:500] + ["... diff muito grande, cortado para visualização ..."]
    return "\n".join(linhas)


def _code_editor_registrar(arquivo: str, backup: str, acao: str, detalhes: dict[str, Any]) -> int:
    with closing(get_conn()) as conn:
        cur = conn.execute(
            """
            INSERT INTO code_edit_history (user_id, arquivo, backup_path, acao, detalhes, criado_em)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (g.user["id"] if g.user is not None else None, arquivo, backup, acao, json.dumps(detalhes, ensure_ascii=False, default=str), agora_iso()),
        )
        conn.commit()
        history_id = int(cur.lastrowid)
    registrar_auditoria(f"codigo_{acao}", "arquivo", arquivo, {"backup": backup, **detalhes})
    return history_id


def _code_editor_write_file(path_value: str, content: str, password: str = "") -> tuple[str, str]:
    _code_editor_check_password(password)
    abs_file = _code_editor_safe_path(path_value)
    if not _code_editor_is_editable(abs_file):
        raise ValueError("Arquivo bloqueado, muito grande ou com extensão não permitida.")
    old_content = Path(abs_file).read_text(encoding="utf-8", errors="replace")
    new_content = str(content or "")
    if old_content == new_content:
        raise ValueError("Nenhuma alteração detectada no arquivo.")
    backup = _code_editor_make_backup(abs_file)
    Path(abs_file).write_text(new_content, encoding="utf-8", newline="")
    diff = _code_editor_diff(old_content, new_content, _code_editor_relpath(abs_file))
    _code_editor_registrar(_code_editor_relpath(abs_file), backup, "save", {
        "antes_bytes": len(old_content.encode("utf-8", errors="ignore")),
        "depois_bytes": len(new_content.encode("utf-8", errors="ignore")),
        "diff": diff,
    })
    return backup, diff


def _code_editor_historico(limit: int = 30) -> list[dict[str, Any]]:
    rows = query_all(
        f"""
        SELECT h.*, u.nome AS usuario_nome, u.username AS usuario_login
        FROM code_edit_history h
        LEFT JOIN users u ON u.id = h.user_id
        ORDER BY h.id DESC
        LIMIT {max(1, min(int(limit or 30), 100))}
        """
    )
    result: list[dict[str, Any]] = []
    for row in rows:
        try:
            detalhes = json.loads(row["detalhes"] or "{}")
        except json.JSONDecodeError:
            detalhes = {}
        result.append({
            "id": row["id"],
            "arquivo": row["arquivo"],
            "backup_path": row["backup_path"] or "",
            "acao": row["acao"],
            "detalhes": detalhes,
            "criado_em": row["criado_em"],
            "usuario_nome": row["usuario_nome"] or row["usuario_login"] or "Sistema",
        })
    return result


def _code_editor_restore_backup(history_id: int, password: str = "") -> tuple[str, str]:
    _code_editor_check_password(password)
    row = query_one("SELECT * FROM code_edit_history WHERE id = ?", (history_id,))
    if row is None or not row["backup_path"]:
        raise ValueError("Backup não encontrado no histórico.")
    target_abs = _code_editor_safe_path(row["arquivo"])
    backup_abs = _code_editor_safe_path(row["backup_path"])
    if not os.path.isfile(backup_abs) or not _code_editor_is_editable(target_abs):
        raise ValueError("Backup ou arquivo atual indisponível para restauração.")
    current_content = Path(target_abs).read_text(encoding="utf-8", errors="replace")
    backup_content = Path(backup_abs).read_text(encoding="utf-8", errors="replace")
    safety_backup = _code_editor_make_backup(target_abs)
    Path(target_abs).write_text(backup_content, encoding="utf-8", newline="")
    diff = _code_editor_diff(current_content, backup_content, row["arquivo"])
    _code_editor_registrar(row["arquivo"], safety_backup, "restore", {"restaurado_de": row["backup_path"], "diff": diff})
    return row["arquivo"], diff


@app.route("/admin/codigo", methods=["GET", "POST"])
@login_required
@module_required("codigo_fonte")
def admin_codigo_fonte() -> str | Response:
    path_atual = request.values.get("path", "")
    arquivo_atual = request.values.get("arquivo", "")
    selected: dict[str, Any] | None = None
    aviso = ""
    diff_preview = ""

    if request.method == "POST":
        action = request.form.get("action", "save")
        password = request.form.get("confirm_password", "")
        try:
            if action == "restore":
                arquivo_restaurado, diff_preview = _code_editor_restore_backup(int(request.form.get("history_id") or 0), password)
                flash(f"Backup restaurado em {arquivo_restaurado}. Se for Python, reinicie o Railway para aplicar.", "success")
                return redirect(url_for("admin_codigo_fonte", arquivo=arquivo_restaurado, restored="1"))
            arquivo_atual = request.form.get("arquivo", "")
            content = request.form.get("content", "")
            backup, diff_preview = _code_editor_write_file(arquivo_atual, content, password)
            flash(f"Arquivo salvo com backup em {backup}.", "success")
        except Exception as exc:
            flash(str(exc), "error")
        return redirect(url_for("admin_codigo_fonte", arquivo=arquivo_atual, path=os.path.dirname(arquivo_atual).replace("\\", "/")))

    try:
        if arquivo_atual:
            selected = _code_editor_read_file(arquivo_atual)
            path_atual = os.path.dirname(selected["path"]).replace("\\", "/")
        listing = _code_editor_list_dir(path_atual)
    except Exception as exc:
        aviso = str(exc)
        listing = _code_editor_list_dir("")

    historico_codigo = _code_editor_historico(40)
    diff_id = request.args.get("diff_id", "").strip()
    if diff_id.isdigit():
        for item in historico_codigo:
            if int(item["id"]) == int(diff_id):
                diff_preview = str(item.get("detalhes", {}).get("diff") or "")
                break

    return render_template(
        "admin_codigo_fonte.html",
        title="Painel ADM - Código fonte",
        listing=listing,
        selected=selected,
        aviso=aviso,
        root_path=_code_editor_root(),
        historico_codigo=historico_codigo,
        diff_preview=diff_preview,
        exige_senha_codigo=get_setting("code_editor_extra_password", "1") == "1",
    )


def _audit_backup_dir() -> Path:
    """Pasta segura para backups de auditoria removida pelo admin."""
    folder = Path(BASE_DIR) / "backups" / "auditoria"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def _backup_auditoria_rows(rows: list[sqlite3.Row], motivo: str = "") -> str:
    """Salva os registros antes de remover, para permitir conferência posterior."""
    if not rows:
        return ""
    payload = {
        "gerado_em": agora_iso(),
        "admin_id": g.user["id"] if getattr(g, "user", None) is not None else None,
        "admin_username": g.user["username"] if getattr(g, "user", None) is not None else None,
        "motivo": motivo,
        "total_registros": len(rows),
        "registros": [dict(row) for row in rows],
    }
    filename = f"auditoria_removida_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.json"
    path = _audit_backup_dir() / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return str(path.relative_to(BASE_DIR))


def _parse_ids(raw_ids: Iterable[str]) -> list[int]:
    ids: list[int] = []
    for raw in raw_ids:
        try:
            value = int(str(raw).strip())
        except Exception:
            continue
        if value > 0 and value not in ids:
            ids.append(value)
    return ids


@app.get("/auditoria")
@login_required
@module_required("auditoria")
def auditoria() -> str:
    q = str(request.args.get("q") or "").strip()
    action = str(request.args.get("action") or "").strip()
    usuario = str(request.args.get("usuario") or "").strip()
    data_inicio = str(request.args.get("data_inicio") or "").strip()
    data_fim = str(request.args.get("data_fim") or "").strip()
    params: list[Any] = []
    where_parts: list[str] = []
    if q:
        where_parts.append("(a.action LIKE ? OR a.entity_type LIKE ? OR a.entity_ref LIKE ? OR a.details LIKE ? OR u.nome LIKE ? OR u.username LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like, like, like, like, like])
    if action:
        where_parts.append("a.action = ?")
        params.append(action)
    if usuario:
        where_parts.append("(u.nome LIKE ? OR u.username LIKE ?)")
        like_user = f"%{usuario}%"
        params.extend([like_user, like_user])
    if data_inicio:
        where_parts.append("date(a.created_at) >= date(?)")
        params.append(data_inicio)
    if data_fim:
        where_parts.append("date(a.created_at) <= date(?)")
        params.append(data_fim)
    where = "WHERE " + " AND ".join(where_parts) if where_parts else ""
    logs = query_all(
        f"""
        SELECT a.*, u.nome AS usuario_nome, u.username AS usuario_login
        FROM audit_logs a
        LEFT JOIN users u ON u.id = a.user_id
        {where}
        ORDER BY a.id DESC
        LIMIT 500
        """,
        params,
    )
    actions = query_all("SELECT DISTINCT action FROM audit_logs ORDER BY action ASC LIMIT 150")
    return render_template(
        "auditoria.html",
        title="Auditoria",
        logs=logs,
        actions=actions,
        q=q,
        action=action,
        usuario=usuario,
        data_inicio=data_inicio,
        data_fim=data_fim,
    )


@app.post("/auditoria/excluir")
@login_required
@module_required("auditoria")
@roles_required("admin")
def excluir_auditoria() -> Response:
    ids = _parse_ids(request.form.getlist("ids"))
    senha_admin = request.form.get("senha_admin", "")
    motivo = str(request.form.get("motivo") or "").strip()[:500]

    if not ids:
        flash("Selecione pelo menos um registro de auditoria para excluir.", "error")
        return redirect(url_for("auditoria"))

    if not check_password_hash(g.user["password_hash"], senha_admin):
        flash("Senha do admin incorreta. Nenhum registro foi excluído.", "error")
        return redirect(url_for("auditoria"))

    placeholders = ",".join("?" for _ in ids)
    with closing(get_conn()) as conn:
        rows = conn.execute(
            f"""
            SELECT a.*, u.nome AS usuario_nome, u.username AS usuario_login
            FROM audit_logs a
            LEFT JOIN users u ON u.id = a.user_id
            WHERE a.id IN ({placeholders})
            ORDER BY a.id ASC
            """,
            ids,
        ).fetchall()

        if not rows:
            flash("Nenhum registro válido foi encontrado para excluir.", "error")
            return redirect(url_for("auditoria"))

        backup_path = _backup_auditoria_rows(rows, motivo)
        conn.execute(f"DELETE FROM audit_logs WHERE id IN ({placeholders})", ids)
        conn.commit()

    registrar_auditoria(
        "AUDITORIA_REMOVIDA",
        "audit_logs",
        ",".join(str(row["id"]) for row in rows),
        {
            "total_removido": len(rows),
            "motivo": motivo,
            "backup": backup_path,
        },
    )
    flash(f"{len(rows)} registro(s) de auditoria removido(s). Backup criado em {backup_path}.", "success")
    return redirect(url_for("auditoria"))


@app.route("/configuracoes", methods=["GET", "POST"])
@login_required
@module_required("configuracoes")
@roles_required("admin")
def configuracoes() -> str | Response:
    if request.method == "POST":
        set_setting("vincular_estoque", "1" if request.form.get("vincular_estoque") == "1" else "0")
        set_setting("usar_conferente", "1" if request.form.get("usar_conferente") == "1" else "0")
        set_setting("maintenance_mode", "1" if request.form.get("maintenance_mode") == "1" else "0")
        set_setting("code_editor_extra_password", "1" if request.form.get("code_editor_extra_password") == "1" else "0")
        registrar_auditoria("salvar_configuracoes", "settings", "geral", {"maintenance_mode": request.form.get("maintenance_mode") == "1"})
        flash("Configuração salva com sucesso.", "success")
        return redirect(url_for("configuracoes"))
    return render_template(
        "configuracoes.html",
        title="Configurações",
        vincular_estoque=get_setting("vincular_estoque", "1") == "1",
        usar_conferente=get_setting("usar_conferente", "1") == "1",
        maintenance_mode=get_setting("maintenance_mode", "0") == "1",
        code_editor_extra_password=get_setting("code_editor_extra_password", "1") == "1",
        mcp_db_path=DB_PATH,
        mcp_command="python -m controle_separacao.mcp_server",
    )


@app.route("/mcp", methods=["GET"])
@login_required
@module_required("mcp_teste")
def mcp_teste() -> str:
    contexto = str(request.args.get("contexto") or "mcp_teste").strip().lower()
    return render_template(
        "mcp_teste.html",
        title="MCP/IA",
        contexto=contexto,
        consultas_salvas=_mcp_lista_consultas_salvas(contexto),
        historico_mcp=_mcp_lista_historico(contexto, limite=25),
        exemplos=[
            "calcule 10*3",
            "quanto é 150 + 20%",
            "status do sistema",
            "resumo por linha limite 50",
            "listar categorias limite 50",
            "pesquisar farinha limite 50",
            "estoque baixo até 10",
            "listar produtos linha pão de alho limite 30",
            "buscar produtos da linha linguiça com saldo entre 1 e 20",
            "consultar produto 123",
            "listar lojas",
            "lotes abertos",
            "movimentações do estoque",
        ],
    )


@app.route("/api/mcp/perguntar", methods=["POST"])
@login_required
def api_mcp_perguntar() -> Response:
    payload = request.get_json(silent=True) or {}
    mensagem = payload.get("mensagem") or request.form.get("mensagem") or ""
    contexto = str(payload.get("contexto") or request.form.get("contexto") or "mcp_teste").strip().lower()
    linha = str(payload.get("linha") or request.form.get("linha") or "").strip()
    filtros = payload.get("filtros") if isinstance(payload.get("filtros"), dict) else {}
    if not _mcp_contexto_permitido(contexto):
        return jsonify({"error": "Você não tem permissão para consultar este contexto."}), 403
    resposta = _executar_pergunta_mcp(str(mensagem), linha=linha, filtros=filtros)
    _mcp_registrar_historico(contexto=contexto, pergunta=str(mensagem), linha=linha, filtros=filtros, resposta=resposta)
    return jsonify(resposta)


@app.route("/api/mcp/preparar-acao", methods=["POST"])
@login_required
def api_mcp_preparar_acao() -> Response:
    payload = request.get_json(silent=True) or {}
    contexto = str(payload.get("contexto") or "mcp_teste").strip().lower()
    action_type = str(payload.get("action_type") or "").strip().lower()
    pergunta = str(payload.get("pergunta") or "").strip()
    linha = str(payload.get("linha") or "").strip()
    filtros = payload.get("filtros") if isinstance(payload.get("filtros"), dict) else {}
    if not _mcp_contexto_permitido(contexto):
        return jsonify({"error": "Você não tem permissão para preparar ação neste contexto."}), 403
    allowed = {"criar_lista_balanco", "gerar_relatorio_gerencial", "revisar_lote", "exportar_excel", "exportar_pdf"}
    if action_type not in allowed:
        return jsonify({"error": "Tipo de ação inválido."}), 400
    resposta = _executar_pergunta_mcp(pergunta or "status do sistema", linha=linha, filtros=filtros)
    title_map = {
        "criar_lista_balanco": "Lista de balanço preparada",
        "gerar_relatorio_gerencial": "Relatório gerencial preparado",
        "revisar_lote": "Revisão de lote preparada",
        "exportar_excel": "Exportação Excel preparada",
        "exportar_pdf": "Exportação PDF preparada",
    }
    payload_db = {
        "linha": linha,
        "filtros": filtros,
        "tool": resposta.get("tool"),
        "total_registros": _mcp_total_registros_resposta(resposta),
        "summary": (resposta.get("table") or {}).get("summary") or {},
    }
    with closing(get_conn()) as conn:
        cur = conn.execute(
            """
            INSERT INTO mcp_prepared_actions (user_id, contexto, titulo, action_type, pergunta, payload, status, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, 'rascunho', ?)
            """,
            (g.user["id"], contexto, title_map[action_type], action_type, pergunta[:800], json.dumps(payload_db, ensure_ascii=False, default=str), agora_iso()),
        )
        conn.commit()
        action_id = int(cur.lastrowid)
    registrar_auditoria("mcp_preparar_acao", "mcp_action", str(action_id), payload_db)
    return jsonify({
        "ok": True,
        "id": action_id,
        "status": "rascunho",
        "mensagem": f"{title_map[action_type]} em modo rascunho. Nada foi alterado no estoque. Revise antes de executar qualquer ação real.",
        "resumo": payload_db,
    })


@app.route("/api/mcp/executar-acao", methods=["POST"])
@login_required
def api_mcp_executar_acao() -> Response:
    payload = request.get_json(silent=True) or {}
    contexto = str(payload.get("contexto") or "mcp_teste").strip().lower()
    action_type = str(payload.get("action_type") or "").strip().lower()
    pergunta = str(payload.get("pergunta") or "").strip()
    linha = str(payload.get("linha") or "").strip()
    filtros = payload.get("filtros") if isinstance(payload.get("filtros"), dict) else {}
    password = str(payload.get("password") or "")
    if not _mcp_contexto_permitido(contexto):
        return jsonify({"error": "Você não tem permissão para executar ação neste contexto."}), 403
    if _mcp_acao_sensivel(action_type) and not _mcp_senha_atual_valida(password):
        return jsonify({
            "requires_password": True,
            "error": "Essa ação mexe em informação sensível. Confirme sua senha para continuar.",
        }), 401
    try:
        resultado = _mcp_executar_acao_real(action_type, pergunta, linha=linha, filtros=filtros)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify(resultado)


@app.route("/api/mcp/categorias", methods=["GET"])
@login_required
def api_mcp_categorias() -> Response:
    termo = str(request.args.get("termo") or "").strip()
    limite = _mcp_extrair_limite(str(request.args.get("limite") or "80"), padrao=80, maximo=120)
    try:
        from . import mcp_server as mcp_tools

        categorias = mcp_tools.listar_categorias_linha(termo=termo, limite=limite)
    except Exception as exc:
        return jsonify({"categorias": [], "error": str(exc)}), 500
    return jsonify({"categorias": categorias})


@app.route("/api/mcp/sugestoes", methods=["GET"])
@login_required
def api_mcp_sugestoes() -> Response:
    termo = str(request.args.get("termo") or "").strip()
    linha = str(request.args.get("linha") or "").strip()
    limite = _mcp_extrair_limite(str(request.args.get("limite") or "12"), padrao=12, maximo=20)
    try:
        from . import mcp_server as mcp_tools

        sugestoes = mcp_tools.sugerir_produtos(termo=termo, linha=linha, limite=limite)
    except Exception as exc:
        return jsonify({"sugestoes": [], "error": str(exc)}), 500
    return jsonify({"sugestoes": sugestoes})


@app.route("/api/mcp/exportar/<formato>", methods=["POST"])
@login_required
def api_mcp_exportar(formato: str) -> Response:
    payload = request.get_json(silent=True) or {}
    mensagem = payload.get("mensagem") or request.form.get("mensagem") or ""
    contexto = str(payload.get("contexto") or request.form.get("contexto") or "mcp_teste").strip().lower()
    linha = str(payload.get("linha") or request.form.get("linha") or "").strip()
    filtros = payload.get("filtros") if isinstance(payload.get("filtros"), dict) else {}
    if not _mcp_contexto_permitido(contexto):
        return jsonify({"error": "Você não tem permissão para exportar este contexto."}), 403
    resposta = _executar_pergunta_mcp(str(mensagem), linha=linha, filtros=filtros)
    formato_normalizado = str(formato or "").strip().lower()
    _mcp_registrar_historico(contexto=contexto, pergunta=str(mensagem), linha=linha, filtros=filtros, resposta=resposta, exportado_formato=formato_normalizado)
    registrar_auditoria("mcp_exportar", "mcp", formato_normalizado, {"contexto": contexto, "pergunta": str(mensagem)[:300], "linha": linha})

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if formato_normalizado in {"excel", "xlsx"}:
        buffer = _mcp_criar_excel(resposta)
        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"consulta_mcp_{timestamp}.xlsx",
        )

    if formato_normalizado == "pdf":
        buffer = _mcp_criar_pdf(resposta)
        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"consulta_mcp_{timestamp}.pdf",
        )

    return jsonify({"error": "Formato inválido. Use excel ou pdf."}), 400




ERP_IMPORT_MODOS = [
    ("atualizar_saldo", "Atualizar saldo pelo ERP e cadastrar novos"),
    ("somente_cadastrar", "Cadastrar novos, sem alterar saldo dos existentes"),
    ("dados_cadastrais", "Atualizar cadastro/categoria, sem mexer no saldo"),
]


def _erp_resumo_from_json(row: sqlite3.Row | None) -> dict[str, Any]:
    if row is None:
        return {}
    try:
        return json.loads(row["resumo_json"] or "{}")
    except json.JSONDecodeError:
        return {}


def _erp_importacao_row(import_id: int) -> sqlite3.Row | None:
    return query_one(
        """
        SELECT imp.*, u.nome AS usuario_nome, u.username AS usuario_login
        FROM erp_stock_imports imp
        LEFT JOIN users u ON u.id = imp.criado_por
        WHERE imp.id = ?
        """,
        (import_id,),
    )


def _erp_importacoes_recentes(limit: int = 15) -> list[sqlite3.Row]:
    return query_all(
        """
        SELECT imp.*, u.nome AS usuario_nome, u.username AS usuario_login
        FROM erp_stock_imports imp
        LEFT JOIN users u ON u.id = imp.criado_por
        ORDER BY imp.id DESC
        LIMIT ?
        """,
        (max(1, min(int(limit or 15), 50)),),
    )



def _erp_backup_dir() -> Path:
    """Pasta segura para backup das importações ERP removidas pelo admin."""
    folder = Path(BASE_DIR) / "backups" / "importacoes_erp"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def _backup_erp_importacao(importacao: sqlite3.Row, itens: list[sqlite3.Row], motivo: str = "") -> str:
    """Salva a importação e seus itens antes de remover do banco."""
    payload = {
        "gerado_em": agora_iso(),
        "admin_id": g.user["id"] if getattr(g, "user", None) is not None else None,
        "admin_username": g.user["username"] if getattr(g, "user", None) is not None else None,
        "motivo": motivo,
        "importacao": dict(importacao),
        "total_itens": len(itens),
        "itens": [dict(item) for item in itens],
        "observacao": "Backup criado antes de remover o registro da importação ERP. Remover o histórico da importação não desfaz alterações de estoque já aplicadas.",
    }
    filename = f"importacao_erp_removida_{importacao['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.json"
    path = _erp_backup_dir() / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return str(path.relative_to(BASE_DIR))


def _erp_import_items(import_id: int, filtro: str = "", limit: int = 350) -> list[sqlite3.Row]:
    where = ["import_id = ?"]
    params: list[Any] = [import_id]
    filtro_norm = str(filtro or "").strip().upper()
    if filtro_norm in {"NOVO", "ALTERAR_SALDO", "SEM_ALTERACAO", "APLICADO"}:
        where.append("status = ?")
        params.append(filtro_norm)
    elif filtro_norm:
        like = f"%{filtro_norm}%"
        where.append("(UPPER(codigo) LIKE ? OR UPPER(descricao) LIKE ? OR UPPER(linha) LIKE ? OR UPPER(codigo_barras) LIKE ?)")
        params.extend([like, like, like, like])
    params.append(max(1, min(int(limit or 350), 1000)))
    return query_all(
        f"""
        SELECT *
        FROM erp_stock_import_items
        WHERE {' AND '.join(where)}
        ORDER BY
            CASE status
                WHEN 'NOVO' THEN 0
                WHEN 'ALTERAR_SALDO' THEN 1
                WHEN 'SEM_ALTERACAO' THEN 2
                ELSE 3
            END,
            linha COLLATE NOCASE ASC,
            descricao COLLATE NOCASE ASC
        LIMIT ?
        """,
        params,
    )


def _erp_calcular_resumo(produtos: list[dict[str, Any]], conn: sqlite3.Connection) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    codigos = [str(item.get("codigo") or "").strip() for item in produtos if str(item.get("codigo") or "").strip()]
    existentes: dict[str, sqlite3.Row] = {}
    if codigos:
        for i in range(0, len(codigos), 800):
            parte = codigos[i:i + 800]
            placeholders = ",".join("?" for _ in parte)
            for row in conn.execute(f"SELECT id, codigo, quantidade_atual FROM stock_items WHERE codigo IN ({placeholders})", parte).fetchall():
                existentes[str(row["codigo"])] = row

    resumo = {
        "novos": 0,
        "alterar_saldo": 0,
        "sem_alteracao": 0,
        "saldo_total_erp": 0.0,
        "saldo_total_anterior": 0.0,
        "delta_total": 0.0,
        "linhas": {},
    }
    preparados: list[dict[str, Any]] = []
    for item in produtos:
        codigo = str(item.get("codigo") or "").strip()
        saldo_erp = float(item.get("saldo_qtd") or 0)
        atual = existentes.get(codigo)
        saldo_anterior = float(atual["quantidade_atual"] or 0) if atual else None
        delta = saldo_erp - float(saldo_anterior or 0)
        if atual is None:
            status = "NOVO"
            resumo["novos"] += 1
            motivo = "Produto ainda não existe no estoque do sistema."
        elif abs(delta) > 0.000001:
            status = "ALTERAR_SALDO"
            resumo["alterar_saldo"] += 1
            motivo = "Saldo do ERP diferente do saldo atual do sistema."
        else:
            status = "SEM_ALTERACAO"
            resumo["sem_alteracao"] += 1
            motivo = "Saldo igual ao saldo atual do sistema."
        linha = str(item.get("linha") or "Sem categoria")
        resumo["saldo_total_erp"] += saldo_erp
        resumo["saldo_total_anterior"] += float(saldo_anterior or 0)
        resumo["delta_total"] += delta
        resumo["linhas"][linha] = resumo["linhas"].get(linha, 0) + 1
        preparado = dict(item)
        preparado.update({
            "stock_item_id": atual["id"] if atual else None,
            "saldo_anterior": saldo_anterior,
            "delta": delta,
            "status": status,
            "motivo": motivo,
        })
        preparados.append(preparado)
    resumo["linhas"] = dict(sorted(resumo["linhas"].items(), key=lambda kv: (-kv[1], kv[0]))[:20])
    return resumo, preparados


@app.route("/estoque/importar-erp", methods=["GET"])
@login_required
@module_required("estoque")
def importar_estoque_erp() -> str | Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode importar arquivo do ERP para atualizar o estoque.")
    return render_template(
        "importar_erp_estoque.html",
        title="Importar estoque do ERP",
        importacoes=_erp_importacoes_recentes(),
        importacao=None,
        resumo={},
        itens=[],
        filtro="",
        modos=ERP_IMPORT_MODOS,
    )


@app.route("/estoque/importar-erp/preview", methods=["POST"])
@login_required
@module_required("estoque")
def importar_estoque_erp_preview() -> Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode importar arquivo do ERP para atualizar o estoque.")
    arquivo = request.files.get("arquivo_erp")
    if arquivo is None or not arquivo.filename:
        flash("Selecione o arquivo do ERP antes de importar.", "error")
        return redirect(url_for("importar_estoque_erp"))
    filename = os.path.basename(arquivo.filename)
    try:
        from .erp_importer import parse_erp_stock_file
        payload = parse_erp_stock_file(arquivo.read(), filename=filename)
    except Exception as exc:
        flash(f"Não consegui ler esse arquivo do ERP: {exc}", "error")
        registrar_auditoria("erro_importacao_erp", "estoque", filename, {"erro": str(exc)})
        return redirect(url_for("importar_estoque_erp"))

    produtos = payload.get("produtos") or []
    if not produtos:
        flash("O arquivo foi lido, mas nenhum produto válido foi encontrado.", "error")
        return redirect(url_for("importar_estoque_erp"))

    with closing(get_conn()) as conn:
        resumo, preparados = _erp_calcular_resumo(produtos, conn)
        cur = conn.execute(
            """
            INSERT INTO erp_stock_imports
            (filename, loja, data_base, total_linhas, total_produtos, total_grupos, duplicados, status, resumo_json, criado_por, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'PREVIEW', ?, ?, ?)
            """,
            (
                filename,
                payload.get("loja") or "",
                payload.get("data_base") or "",
                int(payload.get("total_linhas") or 0),
                int(payload.get("total_produtos") or len(preparados)),
                int(payload.get("total_grupos") or 0),
                int(payload.get("duplicados") or 0),
                json.dumps(resumo, ensure_ascii=False),
                g.user["id"],
                agora_iso(),
            ),
        )
        import_id = int(cur.lastrowid)
        rows = []
        for item in preparados:
            rows.append((
                import_id,
                int(item.get("row_number") or 0),
                item.get("codigo") or "",
                item.get("codigo_barras") or "",
                item.get("nivel") or "",
                item.get("descricao") or "",
                item.get("linha") or "",
                item.get("caminho_linha") or "",
                float(item.get("preco_custo") or 0),
                float(item.get("preco_venda") or 0),
                float(item.get("saldo_qtd") or 0),
                float(item.get("saldo_custo") or 0),
                float(item.get("saldo_venda") or 0),
                float(item.get("dias") or 0),
                float(item.get("sugestao") or 0),
                float(item.get("estoque_ideal") or 0),
                item.get("stock_item_id"),
                item.get("saldo_anterior"),
                float(item.get("delta") or 0),
                item.get("status") or "PENDENTE",
                item.get("motivo") or "",
            ))
        conn.executemany(
            """
            INSERT INTO erp_stock_import_items
            (import_id, row_number, codigo, codigo_barras, nivel, descricao, linha, caminho_linha, preco_custo, preco_venda,
             saldo_qtd, saldo_custo, saldo_venda, dias, sugestao, estoque_ideal, stock_item_id, saldo_anterior, delta, status, motivo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
    # Não registra auditoria/movimentação apenas por importar a prévia do arquivo do estoque.
    # A tela deve ficar limpa; movimentações devem aparecer somente para balanço e alterações manuais do estoque.
    flash("Arquivo do ERP lido com sucesso. Confira o resumo antes de aplicar no estoque.", "success")
    return redirect(url_for("detalhe_importacao_erp", import_id=import_id))


@app.get("/estoque/importar-erp/<int:import_id>")
@login_required
@module_required("estoque")
def detalhe_importacao_erp(import_id: int) -> str | Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode ver importações do ERP.")
    importacao = _erp_importacao_row(import_id)
    if importacao is None:
        flash("Importação não encontrada.", "error")
        return redirect(url_for("importar_estoque_erp"))
    filtro = request.args.get("filtro", "").strip()
    return render_template(
        "importar_erp_estoque.html",
        title="Importação ERP",
        importacoes=_erp_importacoes_recentes(),
        importacao=importacao,
        resumo=_erp_resumo_from_json(importacao),
        itens=_erp_import_items(import_id, filtro=filtro, limit=500),
        filtro=filtro,
        modos=ERP_IMPORT_MODOS,
    )


@app.post("/estoque/importar-erp/<int:import_id>/aplicar")
@login_required
@module_required("estoque")
def aplicar_importacao_erp(import_id: int) -> Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode aplicar importação do ERP.")
    modo = request.form.get("modo", "atualizar_saldo").strip()
    if modo not in {key for key, _ in ERP_IMPORT_MODOS}:
        modo = "atualizar_saldo"
    importacao = _erp_importacao_row(import_id)
    if importacao is None:
        flash("Importação não encontrada.", "error")
        return redirect(url_for("importar_estoque_erp"))
    if importacao["status"] == "APLICADO":
        flash("Essa importação já foi aplicada antes. Para evitar duplicidade, gere uma nova prévia com o arquivo atualizado.", "error")
        return redirect(url_for("detalhe_importacao_erp", import_id=import_id))

    itens = query_all("SELECT * FROM erp_stock_import_items WHERE import_id = ? ORDER BY id ASC", (import_id,))
    if not itens:
        flash("Nenhum item encontrado para aplicar.", "error")
        return redirect(url_for("detalhe_importacao_erp", import_id=import_id))

    atualizados = 0
    cadastrados = 0
    movimentos = 0
    agora = agora_iso()
    with closing(get_conn()) as conn:
        for item in itens:
            existente = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item["codigo"],)).fetchone()
            saldo_erp = float(item["saldo_qtd"] or 0)
            custo = float(item["preco_custo"] or 0)
            if existente is None:
                cur = conn.execute(
                    """
                    INSERT INTO stock_items
                    (codigo, codigo_barras, descricao, fator_embalagem, quantidade_atual, custo_unitario, linha_erp, erp_loja, erp_nivel, linha_caminho_erp, erp_data_base, erp_atualizado_em, ativo, atualizado_em)
                    VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (
                        item["codigo"], item["codigo_barras"] or "", item["descricao"], saldo_erp,
                        custo, item["linha"] or "", importacao["loja"] or "", item["nivel"] or "", item["caminho_linha"] or "", importacao["data_base"] or "", agora, agora,
                    ),
                )
                stock_item_id = int(cur.lastrowid)
                cadastrados += 1
                delta = saldo_erp
                # Importação de arquivo do estoque não gera movimentação para não poluir o histórico.
                # O saldo inicial do produto fica salvo no cadastro do estoque.
                if abs(delta) > 0.000001:
                    pass
            else:
                stock_item_id = int(existente["id"])
                saldo_anterior = float(existente["quantidade_atual"] or 0)
                nova_quantidade = saldo_anterior if modo in {"somente_cadastrar", "dados_cadastrais"} else saldo_erp
                delta = nova_quantidade - saldo_anterior
                if modo != "somente_cadastrar":
                    conn.execute(
                        """
                        UPDATE stock_items
                        SET codigo_barras = COALESCE(NULLIF(?, ''), codigo_barras),
                            descricao = ?,
                            quantidade_atual = ?,
                            custo_unitario = CASE WHEN ? > 0 THEN ? ELSE custo_unitario END,
                            linha_erp = ?,
                            erp_loja = ?,
                            erp_nivel = ?,
                            linha_caminho_erp = ?,
                            erp_data_base = ?,
                            erp_atualizado_em = ?,
                            ativo = 1,
                            atualizado_em = ?
                        WHERE id = ?
                        """,
                        (
                            item["codigo_barras"] or "", item["descricao"], nova_quantidade,
                            custo, custo, item["linha"] or "", importacao["loja"] or "", item["nivel"] or "", item["caminho_linha"] or "", importacao["data_base"] or "", agora, agora, stock_item_id,
                        ),
                    )
                    atualizados += 1
                # Importação de arquivo do estoque não gera movimentação para não poluir o histórico.
                # Registros de movimentação ficam reservados para balanço e alterações manuais do estoque.
                if modo == "atualizar_saldo" and abs(delta) > 0.000001:
                    pass
            conn.execute(
                "UPDATE erp_stock_import_items SET stock_item_id = ?, saldo_anterior = COALESCE(saldo_anterior, ?), delta = ?, status = 'APLICADO' WHERE id = ?",
                (stock_item_id, item["saldo_anterior"], delta, item["id"]),
            )
        resumo = _erp_resumo_from_json(importacao)
        resumo.update({"modo_aplicado": modo, "atualizados": atualizados, "cadastrados": cadastrados, "movimentos": 0, "observacao_movimentos": "Importação de estoque não gera movimentações no histórico."})
        conn.execute(
            "UPDATE erp_stock_imports SET status = 'APLICADO', modo = ?, resumo_json = ?, aplicado_em = ? WHERE id = ?",
            (modo, json.dumps(resumo, ensure_ascii=False), agora, import_id),
        )
        conn.commit()
    # Não registra auditoria/movimentação da importação em massa para evitar excesso de registros na tela.
    flash(f"Importação aplicada: {cadastrados} novo(s), {atualizados} atualizado(s). Nenhuma movimentação foi registrada no histórico.", "success")
    return redirect(url_for("detalhe_importacao_erp", import_id=import_id))



@app.get("/estoque/importar-erp/<int:import_id>/remover")
@login_required
@module_required("estoque")
@roles_required("admin")
def remover_importacao_erp_confirmar(import_id: int) -> str | Response:
    """Tela de confirmação para remover uma importação ERP da lista de importações recentes."""
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode remover importações do ERP.")
    importacao = _erp_importacao_row(import_id)
    if importacao is None:
        flash("Importação ERP não encontrada.", "error")
        return redirect(url_for("importar_estoque_erp"))
    resumo = _erp_resumo_from_json(importacao)
    total_itens = query_one("SELECT COUNT(*) AS total FROM erp_stock_import_items WHERE import_id = ?", (import_id,))
    return render_template(
        "importacao_erp_remover.html",
        title=f"Remover importação ERP #{import_id}",
        importacao=importacao,
        resumo=resumo,
        total_itens=int(total_itens["total"] or 0) if total_itens else 0,
    )


@app.post("/estoque/importar-erp/<int:import_id>/remover")
@login_required
@module_required("estoque")
@roles_required("admin")
def remover_importacao_erp(import_id: int) -> Response:
    """Remove o histórico de uma importação ERP, com senha, backup e auditoria.

    Se a importação já foi aplicada, esta ação remove apenas o registro
    da importação e seus itens de prévia. Ela NÃO desfaz movimentações nem saldo do estoque.
    """
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode remover importações do ERP.")

    importacao = _erp_importacao_row(import_id)
    if importacao is None:
        flash("Importação ERP não encontrada.", "error")
        return redirect(url_for("importar_estoque_erp"))

    senha_admin = request.form.get("senha_admin", "")
    motivo = str(request.form.get("motivo") or "").strip()[:500]
    if not check_password_hash(g.user["password_hash"], senha_admin):
        flash("Senha do admin incorreta. A importação não foi removida.", "error")
        return redirect(f"/estoque/importar-erp/{import_id}/remover")

    with closing(get_conn()) as conn:
        itens = conn.execute("SELECT * FROM erp_stock_import_items WHERE import_id = ? ORDER BY id ASC", (import_id,)).fetchall()
        backup_path = _backup_erp_importacao(importacao, itens, motivo)
        conn.execute("DELETE FROM erp_stock_import_items WHERE import_id = ?", (import_id,))
        conn.execute("DELETE FROM erp_stock_imports WHERE id = ?", (import_id,))
        conn.commit()

    registrar_auditoria(
        "remover_importacao_erp",
        "erp_stock_imports",
        str(import_id),
        {
            "filename": importacao["filename"],
            "status": importacao["status"],
            "total_itens": len(itens),
            "motivo": motivo,
            "backup": backup_path,
            "observacao": "Histórico removido. Se já estava aplicado, o estoque não foi desfeito automaticamente.",
        },
    )
    flash(f"Importação ERP #{import_id} removida com sucesso. Backup criado em {backup_path}.", "success")
    return redirect(url_for("importar_estoque_erp"))
@app.get("/estoque/importar-erp/<int:import_id>/exportar.xlsx")
@login_required
@module_required("estoque")
def exportar_importacao_erp_excel(import_id: int) -> Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode exportar importação do ERP.")
    importacao = _erp_importacao_row(import_id)
    if importacao is None:
        flash("Importação não encontrada.", "error")
        return redirect(url_for("importar_estoque_erp"))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        flash("Para exportar em Excel, instale a dependência openpyxl.", "error")
        return redirect(url_for("detalhe_importacao_erp", import_id=import_id))
    itens = query_all("SELECT * FROM erp_stock_import_items WHERE import_id = ? ORDER BY linha, descricao", (import_id,))
    wb = Workbook()
    ws = wb.active
    ws.title = "Importação ERP"
    ws.append(["Importação ERP", f"#{import_id}"])
    ws.append(["Arquivo", importacao["filename"]])
    ws.append(["Loja", importacao["loja"] or "-"])
    ws.append(["Data base", importacao["data_base"] or "-"])
    ws.append(["Status", importacao["status"]])
    ws.append([])
    headers = ["Status", "Linha", "Código", "Código barras", "Descrição", "Saldo anterior", "Saldo ERP", "Delta", "Custo", "Nível"]
    ws.append(headers)
    for cell in ws[7]:
        cell.font = Font(bold=True)
    for item in itens:
        ws.append([item["status"], item["linha"], item["codigo"], item["codigo_barras"], item["descricao"], item["saldo_anterior"], item["saldo_qtd"], item["delta"], item["preco_custo"], item["nivel"]])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 42)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    registrar_auditoria("exportar_importacao_erp_excel", "erp_stock_imports", str(import_id), {})
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"importacao_erp_{import_id}.xlsx")


@app.get("/estoque/importar-erp/<int:import_id>/exportar.pdf")
@login_required
@module_required("estoque")
def exportar_importacao_erp_pdf(import_id: int) -> Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode exportar importação do ERP.")
    importacao = _erp_importacao_row(import_id)
    if importacao is None:
        flash("Importação não encontrada.", "error")
        return redirect(url_for("importar_estoque_erp"))
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    resumo = _erp_resumo_from_json(importacao)
    itens = query_all("SELECT * FROM erp_stock_import_items WHERE import_id = ? ORDER BY status, linha, descricao LIMIT 120", (import_id,))
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=.8*cm, leftMargin=.8*cm, topMargin=.8*cm, bottomMargin=.8*cm)
    styles = getSampleStyleSheet()
    story: list[Any] = [Paragraph(f"Importação ERP #{import_id}", styles["Title"]), Paragraph(f"Arquivo: {importacao['filename']} • Loja: {importacao['loja'] or '-'} • Data base: {importacao['data_base'] or '-'}", styles["Normal"]), Spacer(1, .25*cm)]
    cards = [["Indicador", "Valor"], ["Produtos", str(importacao["total_produtos"] or 0)], ["Novos", str(resumo.get("novos", 0))], ["Com saldo diferente", str(resumo.get("alterar_saldo", 0))], ["Sem alteração", str(resumo.get("sem_alteracao", 0))], ["Delta total", fmt_num(resumo.get("delta_total", 0))], ["Status", importacao["status"]]]
    table = Table(cards, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#3f7e33")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.grey), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")]))
    story.append(table); story.append(Spacer(1, .25*cm))
    rows = [["Status", "Linha", "Código", "Descrição", "Anterior", "ERP", "Delta"]]
    rows.extend([[i["status"], i["linha"] or "-", i["codigo"], (i["descricao"] or "")[:70], fmt_num(i["saldo_anterior"] or 0), fmt_num(i["saldo_qtd"] or 0), fmt_num(i["delta"] or 0)] for i in itens])
    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#3f7e33")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .2, colors.grey), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 7)]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    registrar_auditoria("exportar_importacao_erp_pdf", "erp_stock_imports", str(import_id), {})
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=f"importacao_erp_{import_id}.pdf")



@app.route("/recebimentos", methods=["GET", "POST"])
@login_required
@module_required("recebimentos")
def recebimentos() -> str | Response:
    if request.method == "POST":
        nota_numero = request.form.get("nota_numero", "").strip()
        observacao = request.form.get("observacao", "").strip()
        itens_json = request.form.get("itens_json", "[]")
        if not nota_numero:
            flash("Informe o número da nota.", "error")
            return redirect(url_for("recebimentos"))
        try:
            itens = json.loads(itens_json)
        except json.JSONDecodeError:
            itens = []
        if not isinstance(itens, list) or not itens:
            flash("Adicione pelo menos um produto na conferência.", "error")
            return redirect(url_for("recebimentos"))

        agora = agora_iso()
        with closing(get_conn()) as conn:
            cur = conn.execute(
                """
                INSERT INTO receipts (nota_numero, status, observacao, conferente_id, criado_em, finalizado_em)
                VALUES (?, 'FINALIZADO', ?, ?, ?, ?)
                """,
                (nota_numero, observacao, g.user["id"], agora, agora),
            )
            receipt_id = cur.lastrowid
            total_itens = 0
            for raw in itens:
                if not isinstance(raw, dict):
                    continue
                codigo_digitado = str(raw.get("codigo") or "").strip()
                validade = str(raw.get("validade") or "").strip()
                try:
                    quantidade = parse_float(str(raw.get("quantidade") or ""), "Quantidade")
                except ValueError:
                    continue
                if not codigo_digitado or quantidade <= 0:
                    continue
                produto = conn.execute(
                    """
                    SELECT * FROM stock_items
                    WHERE ativo = 1 AND (codigo = ? OR codigo_barras = ?)
                    ORDER BY CASE WHEN codigo = ? THEN 0 ELSE 1 END
                    LIMIT 1
                    """,
                    (codigo_digitado, codigo_digitado, codigo_digitado),
                ).fetchone()
                if produto:
                    stock_item_id = produto["id"]
                    codigo = produto["codigo"]
                    codigo_barras = produto["codigo_barras"] or ""
                    descricao = produto["descricao"]
                    conn.execute(
                        "UPDATE stock_items SET quantidade_atual = quantidade_atual + ?, atualizado_em = ? WHERE id = ?",
                        (quantidade, agora, stock_item_id),
                    )
                    conn.execute(
                        """
                        INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em)
                        VALUES (?, 'RECEBIMENTO_MERCADORIA', ?, ?, 'RECEBIMENTO', ?, ?, ?)
                        """,
                        (stock_item_id, quantidade, f"Recebimento NF {nota_numero}" + (f" | Validade {validade}" if validade else ""), receipt_id, g.user["id"], agora),
                    )
                else:
                    stock_item_id = None
                    codigo = codigo_digitado
                    codigo_barras = ""
                    descricao = str(raw.get("descricao") or "Produto não cadastrado").strip() or "Produto não cadastrado"
                conn.execute(
                    """
                    INSERT INTO receipt_items (receipt_id, stock_item_id, codigo, codigo_barras, descricao, validade, quantidade, criado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (receipt_id, stock_item_id, codigo, codigo_barras, descricao, validade, quantidade, agora),
                )
                total_itens += 1
            if total_itens == 0:
                conn.rollback()
                flash("Nenhum item válido foi informado.", "error")
                return redirect(url_for("recebimentos"))
            conn.commit()
        registrar_auditoria("registrar_recebimento", "receipts", str(receipt_id), {"nota": nota_numero, "itens": total_itens})
        flash(f"Recebimento da NF {nota_numero} registrado com {total_itens} item(ns).", "success")
        return redirect(url_for("recebimentos"))

    recebimentos_lista = query_all(
        """
        SELECT r.*, u.nome AS conferente_nome, COUNT(ri.id) AS total_itens, COALESCE(SUM(ri.quantidade), 0) AS total_quantidade
        FROM receipts r
        LEFT JOIN users u ON u.id = r.conferente_id
        LEFT JOIN receipt_items ri ON ri.receipt_id = r.id
        GROUP BY r.id
        ORDER BY r.id DESC
        LIMIT 40
        """
    )
    return render_template("recebimentos.html", title="Recebimentos", recebimentos=recebimentos_lista)


@app.get("/api/estoque/produto")
@login_required
@module_required("recebimentos")
def api_estoque_produto() -> Response:
    codigo = request.args.get("codigo", "").strip()
    if not codigo:
        return jsonify({"ok": False, "error": "Código vazio."}), 400
    item = query_one(
        """
        SELECT id, codigo, codigo_barras, descricao, fator_embalagem, quantidade_atual
        FROM stock_items
        WHERE ativo = 1 AND (codigo = ? OR codigo_barras = ?)
        ORDER BY CASE WHEN codigo = ? THEN 0 ELSE 1 END
        LIMIT 1
        """,
        (codigo, codigo, codigo),
    )
    if not item:
        return jsonify({"ok": False, "error": "Produto não encontrado."}), 404
    return jsonify({"ok": True, "produto": dict(item)})
@app.route("/estoque", methods=["GET", "POST"])
@login_required
@module_required("estoque")
def estoque() -> str | Response:
    if request.method == "POST":
        if not can_edit_stock_registration(g.user):
            return forbidden_redirect("Somente usuários com permissão de admin podem cadastrar ou editar itens do estoque.")

        codigo = request.form.get("codigo", "").strip()
        descricao = request.form.get("descricao", "").strip()
        codigo_barras = request.form.get("codigo_barras", "").strip()
        try:
            fator_embalagem = parse_fator_embalagem(request.form.get("fator_embalagem", "1"))
            quantidade = parse_float(request.form.get("quantidade_atual", ""), "Quantidade")
            custo = parse_float(request.form.get("custo_unitario", "0") or "0", "Custo unitário")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("estoque"))

        if not codigo or not descricao:
            flash("Informe código e descrição.", "error")
            return redirect(url_for("estoque"))

        redirect_q = request.form.get("redirect_q", "").strip()
        redirect_somente = "1" if request.form.get("redirect_somente_com_saldo") == "1" else "0"

        with closing(get_conn()) as conn:
            existente = conn.execute(
                "SELECT * FROM stock_items WHERE codigo = ? OR (codigo_barras IS NOT NULL AND codigo_barras <> '' AND codigo_barras = ?)",
                (codigo, codigo),
            ).fetchone()
            if existente:
                delta = quantidade - float(existente["quantidade_atual"])
                conn.execute(
                    "UPDATE stock_items SET codigo = ?, codigo_barras = ?, descricao = ?, fator_embalagem = ?, quantidade_atual = ?, custo_unitario = ?, ativo = 1, atualizado_em = ? WHERE id = ?",
                    (codigo, codigo_barras, descricao, fator_embalagem, quantidade, custo, agora_iso(), existente["id"]),
                )
                conn.execute(
                    "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (existente["id"], "AJUSTE_MANUAL", delta, "Ajuste manual do cadastro", "ESTOQUE", existente["id"], g.user["id"], agora_iso()),
                )
                flash("Produto atualizado no estoque.", "success")
            else:
                cursor = conn.execute(
                    "INSERT INTO stock_items (codigo, codigo_barras, descricao, fator_embalagem, quantidade_atual, custo_unitario, ativo, atualizado_em) VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
                    (codigo, codigo_barras, descricao, fator_embalagem, quantidade, custo, agora_iso()),
                )
                conn.execute(
                    "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (cursor.lastrowid, "ENTRADA_INICIAL", quantidade, "Cadastro inicial do produto", "ESTOQUE", cursor.lastrowid, g.user["id"], agora_iso()),
                )
                flash("Produto cadastrado no estoque.", "success")
            conn.commit()
        return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))

    filters = sanitize_stock_history_filters(request.args)
    termo = filters["q"]
    somente_com_saldo = request.args.get("somente_com_saldo", "0") == "1"
    hist_usuario = filters["hist_usuario"]
    hist_tipo = filters["hist_tipo"]
    hist_data_inicial = filters["hist_data_inicial"]
    hist_data_final = filters["hist_data_final"]
    stock_items: list[sqlite3.Row] = []
    busca_realizada = bool(termo)

    if termo:
        filtros = ["ativo = 1"]
        params: list[Any] = []
        filtros.append("(codigo = ? OR codigo_barras = ? OR codigo LIKE ? OR descricao LIKE ? OR codigo_barras LIKE ?)")
        like = f"%{termo}%"
        params.extend([termo, termo, like, like, like])
        if somente_com_saldo:
            filtros.append("quantidade_atual > 0")
        where_sql = " AND ".join(filtros)
        stock_items = query_all(
            f"""
            SELECT *
            FROM stock_items
            WHERE {where_sql}
            ORDER BY
                CASE
                    WHEN codigo = ? THEN 0
                    WHEN codigo_barras = ? THEN 1
                    WHEN codigo LIKE ? THEN 2
                    ELSE 3
                END,
                descricao COLLATE NOCASE ASC,
                codigo ASC
            LIMIT 80
            """,
            params + [termo, termo, like],
        )

    stock_movements = fetch_stock_movements(filters, limit=80)
    history_user_options = query_all(
        "SELECT id, nome, username FROM users WHERE ativo = 1 ORDER BY nome COLLATE NOCASE ASC, username COLLATE NOCASE ASC"
    )

    return render_template(
        "estoque.html",
        title="Estoque",
        stock_items=stock_items,
        stock_movements=stock_movements,
        termo_busca=termo,
        somente_com_saldo=somente_com_saldo,
        busca_realizada=busca_realizada,
        hist_usuario=hist_usuario,
        hist_tipo=hist_tipo,
        hist_data_inicial=hist_data_inicial,
        hist_data_final=hist_data_final,
        history_user_options=history_user_options,
        stock_movement_type_options=STOCK_MOVEMENT_TYPE_OPTIONS,
    )


@app.post("/estoque/<int:stock_item_id>/editar")
@login_required
@module_required("estoque")
def editar_item_estoque(stock_item_id: int) -> Response:
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente usuários com permissão de admin podem editar embalagem ou valor do estoque.")

    redirect_q = request.form.get("redirect_q", "").strip()
    redirect_somente = "1" if request.form.get("redirect_somente_com_saldo") == "1" else "0"

    try:
        fator_embalagem = parse_fator_embalagem(request.form.get("fator_embalagem", "1"))
        custo = parse_float(request.form.get("custo_unitario", "0") or "0", "Custo unitário")
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))

    with closing(get_conn()) as conn:
        item = conn.execute("SELECT * FROM stock_items WHERE id = ?", (stock_item_id,)).fetchone()
        if item is None:
            flash("Produto não encontrado.", "error")
            return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))

        conn.execute(
            "UPDATE stock_items SET fator_embalagem = ?, custo_unitario = ?, atualizado_em = ? WHERE id = ?",
            (fator_embalagem, custo, agora_iso(), stock_item_id),
        )
        conn.commit()

    flash("Embalagem e valor atualizados.", "success")
    return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))


@app.post("/estoque/<int:stock_item_id>/ajustar")
@login_required
@module_required("estoque")
def ajustar_estoque(stock_item_id: int) -> Response:
    if not can_adjust_stock(g.user):
        return forbidden_redirect("Sem permissão para ajustar estoque.")

    redirect_q = request.form.get("redirect_q", "").strip()
    redirect_somente = "1" if request.form.get("redirect_somente_com_saldo") == "1" else "0"
    try:
        nova_qtd = parse_float(request.form.get("nova_quantidade", ""), "Quantidade")
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))

    with closing(get_conn()) as conn:
        item = conn.execute("SELECT * FROM stock_items WHERE id = ?", (stock_item_id,)).fetchone()
        if item is None:
            flash("Produto não encontrado.", "error")
            return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))
        delta = nova_qtd - float(item["quantidade_atual"])
        conn.execute(
            "UPDATE stock_items SET quantidade_atual = ?, atualizado_em = ? WHERE id = ?",
            (nova_qtd, agora_iso(), stock_item_id),
        )
        conn.execute(
            "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (stock_item_id, "RECONTAGEM", delta, "Recontagem manual do estoque", "ESTOQUE", stock_item_id, g.user["id"], agora_iso()),
        )
        conn.commit()

    flash("Estoque ajustado.", "success")
    return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))


@app.post("/estoque/<int:stock_item_id>/remover")
@login_required
@module_required("estoque")
def remover_item_estoque(stock_item_id: int) -> Response:
    if not user_is_admin(g.user):
        flash("Apenas o admin pode remover item do estoque.", "error")
        return redirect(url_for("estoque"))

    redirect_q = request.form.get("redirect_q", "").strip()
    redirect_somente = "1" if request.form.get("redirect_somente_com_saldo") == "1" else "0"

    with closing(get_conn()) as conn:
        item = conn.execute("SELECT * FROM stock_items WHERE id = ?", (stock_item_id,)).fetchone()
        if item is None or int(item["ativo"] or 0) != 1:
            flash("Item de estoque não encontrado.", "error")
            return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))

        quantidade_atual = float(item["quantidade_atual"] or 0)
        conn.execute(
            "UPDATE stock_items SET ativo = 0, quantidade_atual = 0, atualizado_em = ? WHERE id = ?",
            (agora_iso(), stock_item_id),
        )
        conn.execute(
            "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                stock_item_id,
                "REMOVIDO_ESTOQUE",
                -quantidade_atual,
                "Item removido da lista ativa do estoque",
                "ESTOQUE",
                stock_item_id,
                g.user["id"],
                agora_iso(),
            ),
        )
        conn.commit()

    flash("Item removido do estoque.", "success")
    return redirect(url_for("estoque", q=redirect_q, somente_com_saldo=redirect_somente))



def usuarios_por_role(role: str | None = None) -> list[sqlite3.Row]:
    if role:
        return query_all("SELECT * FROM users WHERE ativo = 1 AND role = ? ORDER BY nome", (role,))
    return query_all("SELECT * FROM users WHERE ativo = 1 ORDER BY nome")


def copiar_pendencias_para_nova_separacao(conn: sqlite3.Connection, separation_id: int, store_id: int) -> int:
    pendencias = conn.execute(
        """
        SELECT si.*
        FROM separation_items si
        JOIN separations s ON s.id = si.separation_id
        WHERE s.store_id = ?
          AND s.status = 'FINALIZADA'
          AND COALESCE(si.carryover_copied, 0) = 0
          AND (COALESCE(si.quantidade_pedida, 0) - COALESCE(si.quantidade_separada, 0)) > 0
        ORDER BY COALESCE(s.finalizado_em, s.criado_em) ASC, si.id ASC
        """,
        (store_id,),
    ).fetchall()

    copiados = 0
    for item in pendencias:
        restante = float(item["quantidade_pedida"] or 0) - float(item["quantidade_separada"] or 0)
        if restante <= 0:
            continue

        existente = conn.execute(
            "SELECT id, quantidade_pedida FROM separation_items WHERE separation_id = ? AND codigo = ?",
            (separation_id, item["codigo"]),
        ).fetchone()
        if existente is None:
            conn.execute(
                """
                INSERT INTO separation_items (
                    separation_id, codigo, descricao, fator_embalagem, quantidade_pedida, quantidade_separada, status,
                    custo_unitario_ref, carryover_source_item_id, carryover_copied, criado_em, atualizado_em
                ) VALUES (?, ?, ?, ?, ?, 0, 'PENDENTE', ?, ?, 0, ?, ?)
                """,
                (
                    separation_id,
                    item["codigo"],
                    item["descricao"],
                    float(item["fator_embalagem"] or 1),
                    restante,
                    item["custo_unitario_ref"],
                    item["id"],
                    agora_iso(),
                    agora_iso(),
                ),
            )
        else:
            nova_quantidade = float(existente["quantidade_pedida"] or 0) + restante
            conn.execute(
                "UPDATE separation_items SET descricao = ?, fator_embalagem = ?, quantidade_pedida = ?, atualizado_em = ? WHERE id = ?",
                (item["descricao"], float(item["fator_embalagem"] or 1), nova_quantidade, agora_iso(), existente["id"]),
            )
        conn.execute(
            "UPDATE separation_items SET carryover_copied = 1, atualizado_em = ? WHERE id = ?",
            (agora_iso(), item["id"]),
        )
        copiados += 1

    return copiados


def desfazer_pendencias_transferidas(conn: sqlite3.Connection, separation_id: int) -> None:
    origem_ids = [
        row["carryover_source_item_id"]
        for row in conn.execute(
            "SELECT DISTINCT carryover_source_item_id FROM separation_items WHERE separation_id = ? AND carryover_source_item_id IS NOT NULL",
            (separation_id,),
        ).fetchall()
    ]
    if origem_ids:
        conn.execute(
            f"UPDATE separation_items SET carryover_copied = 0, atualizado_em = ? WHERE id IN ({','.join('?' for _ in origem_ids)})",
            (agora_iso(), *origem_ids),
        )


@app.route("/separacoes/nova", methods=["GET", "POST"])
@login_required
@module_required("pedidos")
def nova_separacao() -> str | Response:
    if request.method == "POST":
        lote_nome = request.form.get("lote_nome", "").strip()
        data_referencia = request.form.get("data_referencia", "").strip()
        responsavel_id = request.form.get("responsavel_id", "").strip()
        usar_conferente = get_setting("usar_conferente", "1") == "1"
        conferente_id = request.form.get("conferente_id", "").strip() or None
        if not usar_conferente:
            conferente_id = None
        stores = request.form.getlist("stores")
        usar_estoque = 1 if request.form.get("usar_estoque") == "1" else 0
        trazer_pendencias = request.form.get("trazer_pendencias") == "1"
        observacao = request.form.get("observacao", "").strip()

        if not lote_nome or not data_referencia or not responsavel_id.isdigit() or not stores:
            flash("Preencha lote, data, responsável e selecione ao menos uma loja.", "error")
            return redirect(url_for("nova_separacao"))

        lote_codigo = novo_lote_codigo()
        pendencias_copiadas = 0
        with closing(get_conn()) as conn:
            for store_id in stores:
                if not str(store_id).isdigit():
                    continue
                cursor = conn.execute(
                    """
                    INSERT INTO separations (
                        lote_codigo, lote_nome, data_referencia, store_id, responsavel_id, conferente_id,
                        status, usar_estoque, observacao, criado_por, criado_em
                    ) VALUES (?, ?, ?, ?, ?, ?, 'ABERTA', ?, ?, ?, ?)
                    """,
                    (
                        lote_codigo,
                        lote_nome,
                        data_referencia,
                        int(store_id),
                        int(responsavel_id),
                        int(conferente_id) if conferente_id and str(conferente_id).isdigit() else None,
                        usar_estoque,
                        observacao,
                        g.user["id"],
                        agora_iso(),
                    ),
                )
                if trazer_pendencias:
                    pendencias_copiadas += copiar_pendencias_para_nova_separacao(conn, cursor.lastrowid, int(store_id))
            conn.commit()

        mensagem = "Separações criadas. Agora você pode lançar os itens do lote em uma tela única, com quantidade diferente para cada loja, sem entrar uma por uma."
        if trazer_pendencias and pendencias_copiadas:
            mensagem += f" Também trouxe {pendencias_copiadas} pendência(s) parcial(is) de dias anteriores para completar o restante."
        flash(mensagem, "success")
        return redirect(url_for("grade_lote", lote_codigo=lote_codigo))

    return render_template(
        "nova_separacao.html",
        title="Criar separações",
        hoje=datetime.now().strftime("%Y-%m-%d"),
        stores=query_all("SELECT * FROM stores WHERE ativo = 1 ORDER BY nome"),
        separadores=usuarios_por_role("separador"),
        conferentes=usuarios_por_role("conferente"),
        usar_conferente=get_setting("usar_conferente", "1") == "1",
        trazer_pendencias_padrao=True,
    )


def listar_lotes_em_aberto() -> list[sqlite3.Row]:
    chave_expr = lote_operacao_chave_expr("s")
    return query_all(
        f"""
        SELECT {chave_expr} AS operacao_chave,
               MAX(s.lote_codigo) AS lote_codigo,
               s.lote_nome,
               s.data_referencia,
               MAX(r.nome) AS responsavel_nome,
               MAX(c.nome) AS conferente_nome,
               COUNT(*) AS total_lojas,
               SUM(CASE WHEN s.status = 'FINALIZADA' THEN 1 ELSE 0 END) AS lojas_finalizadas,
               GROUP_CONCAT(st.nome, ' • ') AS lojas
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        GROUP BY operacao_chave, s.lote_nome, s.data_referencia
        HAVING SUM(CASE WHEN s.status NOT IN ('FINALIZADA', 'CANCELADA') THEN 1 ELSE 0 END) > 0
        ORDER BY MAX(s.id) DESC
        """
    )


def carregar_lote(operacao_chave: str) -> list[sqlite3.Row]:
    chave_expr = lote_operacao_chave_expr("s")
    rows = query_all(
        f"""
        SELECT s.*, st.nome AS store_nome,
               r.nome AS responsavel_nome,
               c.nome AS conferente_nome,
               {chave_expr} AS operacao_chave
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        WHERE {chave_expr} = ?
          AND s.status <> 'CANCELADA'
        ORDER BY s.id ASC
        """,
        (operacao_chave,),
    )
    return sort_store_rows(rows)


def produtos_do_lote(operacao_chave: str, separacoes: list[sqlite3.Row]) -> list[dict[str, Any]]:
    store_ids = [row["store_id"] for row in separacoes]
    store_names = {row["store_id"]: row["store_nome"] for row in separacoes}
    chave_expr = lote_operacao_chave_expr("s")
    rows = query_all(
        f"""
        SELECT si.codigo, si.descricao, si.fator_embalagem, si.quantidade_pedida, s.store_id
        FROM separation_items si
        JOIN separations s ON s.id = si.separation_id
        WHERE {chave_expr} = ?
        ORDER BY si.descricao COLLATE NOCASE ASC, si.codigo ASC, s.store_id ASC
        """
        ,
        (operacao_chave,),
    )
    produtos: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["codigo"], row["descricao"])
        if key not in produtos:
            produtos[key] = {
                "codigo": row["codigo"],
                "descricao": row["descricao"],
                "fator_embalagem": float(row["fator_embalagem"] or 1),
                "total": 0.0,
                "quantidades": {},
            }
        quantidade = float(row["quantidade_pedida"] or 0)
        produtos[key]["total"] += quantidade
        produtos[key]["quantidades"][row["store_id"]] = quantidade

    resultado: list[dict[str, Any]] = []
    for produto in produtos.values():
        produto["linhas"] = [
            {
                "store_id": store_id,
                "store_nome": store_names[store_id],
                "quantidade": produto["quantidades"].get(store_id, 0),
            }
            for store_id in store_ids
        ]
        resultado.append(produto)
    return resultado





def lotes_visiveis_para_usuario(user: sqlite3.Row | None) -> list[sqlite3.Row]:
    todos = listar_lotes_em_aberto()
    if user is None or user_is_admin(user):
        return todos
    resultado: list[sqlite3.Row] = []
    for lote in todos:
        separacoes = carregar_lote(lote["operacao_chave"])
        if normalize_role(user["role"]) == "separador" and any(sep["responsavel_id"] == user["id"] for sep in separacoes):
            resultado.append(lote)
        elif normalize_role(user["role"]) == "conferente" and any(sep["conferente_id"] == user["id"] for sep in separacoes):
            resultado.append(lote)
    return resultado


def pode_acessar_lote_operacao(separacoes: list[sqlite3.Row], modo: str) -> bool:
    if g.user is None:
        return False
    if user_is_admin(g.user):
        return True
    if modo == "separacao":
        return normalize_role(g.user["role"]) == "separador" and any(sep["responsavel_id"] == g.user["id"] for sep in separacoes)
    if modo == "conferencia":
        return normalize_role(g.user["role"]) == "conferente" and any(sep["conferente_id"] == g.user["id"] for sep in separacoes)
    return False


def itens_do_lote_para_fluxo(operacao_chave: str, separacoes: list[sqlite3.Row]) -> list[dict[str, Any]]:
    store_ids = [row["store_id"] for row in separacoes]
    store_names = {row["store_id"]: row["store_nome"] for row in separacoes}
    separation_ids = {row["store_id"]: row["id"] for row in separacoes}
    chave_expr = lote_operacao_chave_expr("s")
    rows = query_all(
        f"""
        SELECT si.id, si.codigo, si.descricao, si.fator_embalagem, si.quantidade_pedida, si.quantidade_separada,
               COALESCE(si.quantidade_conferida, si.quantidade_separada) AS quantidade_conferencia_visivel,
               si.status, s.id AS separation_id, s.store_id
        FROM separation_items si
        JOIN separations s ON s.id = si.separation_id
        WHERE {chave_expr} = ?
        ORDER BY si.descricao COLLATE NOCASE ASC, si.codigo ASC, s.store_id ASC
        """,
        (operacao_chave,),
    )
    produtos: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["codigo"], row["descricao"])
        if key not in produtos:
            produtos[key] = {
                "codigo": row["codigo"],
                "descricao": row["descricao"],
                "fator_embalagem": float(row["fator_embalagem"] or 1),
                "total_pedido": 0.0,
                "total_separado": 0.0,
                "total_conferido": 0.0,
                "stores": {},
            }
        pedido = float(row["quantidade_pedida"] or 0)
        separado = float(row["quantidade_separada"] or 0)
        conferido = float(row["quantidade_conferencia_visivel"] or 0)
        produtos[key]["total_pedido"] += pedido
        produtos[key]["total_separado"] += separado
        produtos[key]["total_conferido"] += conferido
        produtos[key]["stores"][row["store_id"]] = {
            "item_id": row["id"],
            "separation_id": row["separation_id"],
            "store_id": row["store_id"],
            "store_nome": store_names[row["store_id"]],
            "quantidade_pedida": pedido,
            "quantidade_separada": separado,
            "quantidade_conferida": conferido,
            "fator_embalagem": float(row["fator_embalagem"] or 1),
            "status": row["status"],
        }

    resultado: list[dict[str, Any]] = []
    for produto in produtos.values():
        linhas = []
        for store_id in store_ids:
            linha = produto["stores"].get(store_id)
            if linha is None:
                linha = {
                    "item_id": None,
                    "separation_id": separation_ids[store_id],
                    "store_id": store_id,
                    "store_nome": store_names[store_id],
                    "quantidade_pedida": 0.0,
                    "quantidade_separada": 0.0,
                    "quantidade_conferida": 0.0,
                    "fator_embalagem": float(produto["fator_embalagem"] or 1),
                    "status": "PENDENTE",
                }
            linhas.append(linha)
        produto["linhas"] = linhas
        resultado.append(produto)
    return resultado


def distribuir_quantidades_racionalizadas(total_real: float, linhas: list[dict[str, Any]], lojas_fixas: set[int]) -> dict[int, float]:
    pedido_por_loja = {linha["store_id"]: float(linha["quantidade_pedida"] or 0) for linha in linhas}
    distribuicao = {store_id: 0.0 for store_id in pedido_por_loja}
    total_fixo = 0.0
    for store_id in lojas_fixas:
        valor = min(pedido_por_loja.get(store_id, 0.0), max(total_real, 0.0))
        distribuicao[store_id] = valor
        total_fixo += valor
    restante = max(total_real - total_fixo, 0.0)
    livres = [store_id for store_id in pedido_por_loja if store_id not in lojas_fixas]
    total_pedido_livre = sum(pedido_por_loja[store_id] for store_id in livres)
    if total_pedido_livre <= 0:
        return distribuicao
    base = {}
    for store_id in livres:
        exato = restante * (pedido_por_loja[store_id] / total_pedido_livre)
        base[store_id] = exato
        distribuicao[store_id] = float(int(exato))
    usado = sum(distribuicao.values())
    sobra = int(round(total_real - usado))
    ordem = sorted(livres, key=lambda sid: (-(base[sid] - int(base[sid])), natural_store_sort_key(str(sid))))
    idx = 0
    while sobra > 0 and ordem:
        store_id = ordem[idx % len(ordem)]
        if distribuicao[store_id] < pedido_por_loja[store_id]:
            distribuicao[store_id] += 1
            sobra -= 1
        idx += 1
        if idx > 10000:
            break
    for store_id, pedido in pedido_por_loja.items():
        if distribuicao[store_id] > pedido:
            distribuicao[store_id] = pedido
    return distribuicao


def atualizar_status_item(qtd_pedida: float, qtd_real: float, conferido: bool = False) -> str:
    if qtd_real <= 0:
        return "PENDENTE"
    if qtd_real < qtd_pedida:
        return "CONFERIDO" if conferido else "PARCIAL"
    return "CONFERIDO" if conferido else "SEPARADO"


def itens_pendentes_lote(operacao_chave: str, modo: str) -> list[dict[str, str]]:
    produtos = itens_do_lote_para_fluxo(operacao_chave, carregar_lote(operacao_chave))
    resultado: list[dict[str, str]] = []
    for produto in produtos:
        if modo == "separacao":
            pendente = any(float(linha["quantidade_separada"] or 0) < float(linha["quantidade_pedida"] or 0) for linha in produto["linhas"])
        else:
            pendente = any(str(linha["status"]) != "CONFERIDO" and float(linha["quantidade_separada"] or 0) > 0 for linha in produto["linhas"])
        if pendente or not resultado:
            resultado.append({"codigo": produto["codigo"], "descricao": produto["descricao"]})
    return resultado

def separation_visibility_clause() -> tuple[str, list[Any]]:
    user = g.user
    if user_is_admin(user):
        return "WHERE s.status <> 'CANCELADA'", []
    if normalize_role(user["role"]) == "separador":
        return "WHERE s.status <> 'CANCELADA' AND s.responsavel_id = ?", [user["id"]]
    return "WHERE s.status <> 'CANCELADA' AND (s.conferente_id = ? OR s.responsavel_id = ?)", [user["id"], user["id"]]




@app.get("/lotes")
@login_required
@module_required("lotes")
def listar_lotes() -> str:
    return render_template("lotes.html", title="Lotes", lotes=listar_lotes_em_aberto())




@app.route("/lotes/<lote_codigo>/grade", methods=["GET", "POST"])
@login_required
@module_required("lotes")
def grade_lote(lote_codigo: str) -> str | Response:
    separacoes = carregar_lote(lote_codigo)
    if not separacoes:
        flash("Lote não encontrado.", "error")
        return redirect(url_for("listar_lotes"))

    # Regra de permissão: separador apenas separa itens já lançados.
    # Ele não pode abrir a grade de montagem nem adicionar produtos ao pedido criado.
    if not user_is_admin(g.user) and normalize_role(g.user["role"]) == "separador":
        flash("Seu acesso é apenas para separar itens. A montagem/adição de produtos fica com o responsável/admin.", "error")
        return redirect(url_for("separar_itens_lote", lote_codigo=lote_codigo))

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        descricao = request.form.get("descricao", "").strip()
        try:
            fator_embalagem = parse_fator_embalagem(request.form.get("fator_embalagem", "1"))
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("grade_lote", lote_codigo=lote_codigo))
        if not codigo or not descricao:
            flash("Informe código e descrição.", "error")
            return redirect(url_for("grade_lote", lote_codigo=lote_codigo))

        quantidades: list[tuple[int, float, float]] = []
        for sep in separacoes:
            if sep["status"] == "FINALIZADA":
                continue
            raw = request.form.get(f"qty_{sep['id']}", "").strip()
            if not raw:
                continue
            try:
                quantidade_emb = parse_float(raw, f"Quantidade da loja {sep['store_nome']}")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("grade_lote", lote_codigo=lote_codigo))
            if quantidade_emb > 0:
                quantidades.append((sep["id"], quantidade_emb, quantidade_emb * fator_embalagem))

        if not quantidades:
            flash("Preencha ao menos uma loja com quantidade maior que zero.", "error")
            return redirect(url_for("grade_lote", lote_codigo=lote_codigo))

        stock = query_one("SELECT custo_unitario FROM stock_items WHERE codigo = ?", (codigo,))
        custo_ref = stock["custo_unitario"] if stock else 0

        with closing(get_conn()) as conn:
            for separation_id, quantidade_emb, quantidade in quantidades:
                existente = conn.execute(
                    "SELECT id, quantidade_pedida FROM separation_items WHERE separation_id = ? AND codigo = ?",
                    (separation_id, codigo),
                ).fetchone()
                if existente is None:
                    conn.execute(
                        """
                        INSERT INTO separation_items (separation_id, codigo, descricao, fator_embalagem, quantidade_pedida, quantidade_separada, status, custo_unitario_ref, criado_em, atualizado_em)
                        VALUES (?, ?, ?, ?, ?, 0, 'PENDENTE', ?, ?, ?)
                        """,
                        (separation_id, codigo, descricao, fator_embalagem, quantidade, custo_ref, agora_iso(), agora_iso()),
                    )
                else:
                    nova_quantidade = float(existente["quantidade_pedida"]) + quantidade
                    conn.execute(
                        "UPDATE separation_items SET descricao = ?, fator_embalagem = ?, quantidade_pedida = ?, atualizado_em = ? WHERE id = ?",
                        (descricao, fator_embalagem, nova_quantidade, agora_iso(), existente["id"]),
                    )
            conn.commit()

        flash(f"Produto lançado em {len(quantidades)} loja(s) com quantidades individuais.", "success")
        return redirect(url_for("grade_lote", lote_codigo=lote_codigo))

    primeira = separacoes[0]
    return render_template(
        "grade_lote.html",
        title="Lançamento por lote",
        lote_codigo=lote_codigo,
        lote_nome=primeira["lote_nome"],
        data_referencia=primeira["data_referencia"],
        responsavel_nome=primeira["responsavel_nome"],
        conferente_nome=primeira["conferente_nome"],
        separacoes=separacoes,
        produtos=produtos_do_lote(lote_codigo, separacoes),
    )







@app.route("/lotes/<lote_codigo>/separar-itens", methods=["GET", "POST"])
@login_required
@module_required("separacoes")
def separar_itens_lote(lote_codigo: str) -> str | Response:
    separacoes = carregar_lote(lote_codigo)
    if not separacoes or not pode_acessar_lote_operacao(separacoes, "separacao"):
        flash("Lote não encontrado ou sem permissão para separar.", "error")
        return redirect(url_for("listar_separacoes"))

    produtos = itens_do_lote_para_fluxo(lote_codigo, separacoes)
    if not produtos:
        flash("Esse lote ainda não possui itens para separar.", "error")
        return redirect(url_for("grade_lote", lote_codigo=lote_codigo) if user_is_admin(g.user) else url_for("listar_separacoes"))

    codigo_atual = request.values.get("codigo", "").strip()
    produto_atual = next((p for p in produtos if p["codigo"] == codigo_atual), None) if codigo_atual else None
    if produto_atual is None:
        pendentes = itens_pendentes_lote(lote_codigo, "separacao")
        if pendentes:
            codigo_atual = pendentes[0]["codigo"]
            produto_atual = next((p for p in produtos if p["codigo"] == codigo_atual), produtos[0])
        else:
            produto_atual = produtos[0]
            codigo_atual = produto_atual["codigo"]

    if request.method == "POST":
        try:
            quantidade_real = parse_float(request.form.get("quantidade_real", "0"), "Quantidade real")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("separar_itens_lote", lote_codigo=lote_codigo, codigo=codigo_atual))

        linhas = produto_atual["linhas"]
        if request.form.get("aplicar_racionalizacao") == "1":
            fixas: set[int] = set()
            for linha in linhas:
                if request.form.get(f"fixo_{linha['store_id']}") == "1":
                    fixas.add(int(linha["store_id"]))
            sugestao = distribuir_quantidades_racionalizadas(quantidade_real, linhas, fixas)
        else:
            sugestao = {}

        valores_salvar: dict[int, float] = {}
        for linha in linhas:
            campo = request.form.get(f"quantidade_loja_{linha['store_id']}", "").strip()
            if campo:
                try:
                    valores_salvar[int(linha["store_id"])] = parse_float(campo, f"Quantidade da {linha['store_nome']}")
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("separar_itens_lote", lote_codigo=lote_codigo, codigo=codigo_atual))
            else:
                valores_salvar[int(linha["store_id"])] = float(sugestao.get(int(linha["store_id"]), linha["quantidade_separada"] or 0))

        with closing(get_conn()) as conn:
            for linha in linhas:
                item_id = linha["item_id"]
                if not item_id:
                    continue
                qtd = float(valores_salvar.get(int(linha["store_id"]), 0))
                status = atualizar_status_item(float(linha["quantidade_pedida"] or 0), qtd, conferido=False)
                conn.execute(
                    "UPDATE separation_items SET quantidade_separada = ?, status = ?, atualizado_em = ?, quantidade_conferida = NULL, conferido_em = NULL WHERE id = ?",
                    (qtd, status, agora_iso(), item_id),
                )
                conn.execute(
                    "UPDATE separations SET status = CASE WHEN status = 'ABERTA' THEN 'EM_SEPARACAO' ELSE status END WHERE id = ? AND status <> 'FINALIZADA'",
                    (linha["separation_id"],),
                )
            conn.commit()

        produto_indices = [p["codigo"] for p in produtos]
        idx = produto_indices.index(codigo_atual)
        proximo_codigo = produto_indices[idx + 1] if idx + 1 < len(produto_indices) else codigo_atual
        flash("Separação do item salva com sucesso.", "success")
        return redirect(url_for("separar_itens_lote", lote_codigo=lote_codigo, codigo=proximo_codigo))

    lotes_visiveis = lotes_visiveis_para_usuario(g.user)
    return render_template(
        "operacao_item_lote.html",
        title="Separar itens do lote",
        modo="separacao",
        lotes=lotes_visiveis,
        lote_codigo=lote_codigo,
        separacoes=separacoes,
        produtos=produtos,
        produto_atual=produto_atual,
        responsavel_nome=separacoes[0]["responsavel_nome"],
        conferente_nome=separacoes[0]["conferente_nome"],
    )


@app.route("/lotes/<lote_codigo>/conferir-itens", methods=["GET", "POST"])
@login_required
@module_required("separacoes")
def conferir_itens_lote(lote_codigo: str) -> str | Response:
    separacoes = carregar_lote(lote_codigo)
    if not separacoes or not pode_acessar_lote_operacao(separacoes, "conferencia"):
        flash("Lote não encontrado ou sem permissão para conferir.", "error")
        return redirect(url_for("listar_separacoes"))

    produtos = itens_do_lote_para_fluxo(lote_codigo, separacoes)
    produtos_com_separacao = [p for p in produtos if any(float(l["quantidade_separada"] or 0) > 0 for l in p["linhas"]) ]
    if not produtos_com_separacao:
        flash("Esse lote ainda não possui itens separados para conferência.", "error")
        return redirect(url_for("listar_separacoes"))

    codigo_atual = request.values.get("codigo", "").strip()
    produto_atual = next((p for p in produtos_com_separacao if p["codigo"] == codigo_atual), None) if codigo_atual else None
    if produto_atual is None:
        pendentes = itens_pendentes_lote(lote_codigo, "conferencia")
        if pendentes:
            codigo_atual = pendentes[0]["codigo"]
            produto_atual = next((p for p in produtos_com_separacao if p["codigo"] == codigo_atual), produtos_com_separacao[0])
        else:
            produto_atual = produtos_com_separacao[0]
            codigo_atual = produto_atual["codigo"]

    if request.method == "POST":
        with closing(get_conn()) as conn:
            for linha in produto_atual["linhas"]:
                item_id = linha["item_id"]
                if not item_id:
                    continue
                try:
                    qtd_conf = parse_float(request.form.get(f"confirmada_loja_{linha['store_id']}", linha["quantidade_separada"]), f"Conferência da {linha['store_nome']}")
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("conferir_itens_lote", lote_codigo=lote_codigo, codigo=codigo_atual))
                status = atualizar_status_item(float(linha["quantidade_pedida"] or 0), qtd_conf, conferido=True)
                conn.execute(
                    "UPDATE separation_items SET quantidade_conferida = ?, quantidade_separada = ?, status = ?, conferido_em = ?, atualizado_em = ? WHERE id = ?",
                    (qtd_conf, qtd_conf, status, agora_iso(), agora_iso(), item_id),
                )
            for sep in separacoes:
                pend = conn.execute("SELECT COUNT(*) AS c FROM separation_items WHERE separation_id = ? AND status NOT IN ('CONFERIDO', 'PENDENTE')", (sep['id'],)).fetchone()["c"]
                tem_separado = conn.execute("SELECT COUNT(*) AS c FROM separation_items WHERE separation_id = ? AND quantidade_separada > 0", (sep['id'],)).fetchone()["c"]
                if tem_separado > 0 and pend == 0 and sep['status'] != 'FINALIZADA':
                    conn.execute("UPDATE separations SET status = 'AGUARDANDO_CONFERENCIA', enviado_conferencia_em = COALESCE(enviado_conferencia_em, ?) WHERE id = ?", (agora_iso(), sep['id']))
            conn.commit()
        codigos = [p["codigo"] for p in produtos_com_separacao]
        idx = codigos.index(codigo_atual)
        proximo_codigo = codigos[idx + 1] if idx + 1 < len(codigos) else codigo_atual
        flash("Conferência do item registrada com sucesso.", "success")
        return redirect(url_for("conferir_itens_lote", lote_codigo=lote_codigo, codigo=proximo_codigo))

    lotes_visiveis = lotes_visiveis_para_usuario(g.user)
    return render_template(
        "operacao_item_lote.html",
        title="Conferir itens do lote",
        modo="conferencia",
        lotes=lotes_visiveis,
        lote_codigo=lote_codigo,
        separacoes=separacoes,
        produtos=produtos_com_separacao,
        produto_atual=produto_atual,
        responsavel_nome=separacoes[0]["responsavel_nome"],
        conferente_nome=separacoes[0]["conferente_nome"],
    )


def finalizar_separacao_no_conn(conn: sqlite3.Connection, separation: sqlite3.Row, actor_id: int) -> float:
    itens = conn.execute("SELECT * FROM separation_items WHERE separation_id = ?", (separation['id'],)).fetchall()
    if not itens:
        return 0.0
    usar_controle_global = get_setting("vincular_estoque", "1") == "1"
    precisa_abater = usar_controle_global and bool(separation["usar_estoque"])
    if precisa_abater:
        problemas = validar_estoque_para_finalizacao(conn, separation['id'])
        if problemas:
            raise ValueError("Saldo insuficiente para finalizar com controle de estoque: " + "; ".join(problemas))
        for item in itens:
            stock = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item['codigo'],)).fetchone()
            if stock is None:
                conn.execute(
                    "INSERT INTO stock_items (codigo, descricao, quantidade_atual, custo_unitario, atualizado_em) VALUES (?, ?, 0, ?, ?)",
                    (item['codigo'], item['descricao'], item['custo_unitario_ref'], agora_iso()),
                )
                stock = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item['codigo'],)).fetchone()
            novo_saldo = float(stock['quantidade_atual']) - float(item['quantidade_separada'])
            conn.execute(
                "UPDATE stock_items SET quantidade_atual = ?, atualizado_em = ? WHERE id = ?",
                (novo_saldo, agora_iso(), stock['id']),
            )
            conn.execute(
                "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, 'SAIDA_SEPARACAO', ?, ?, 'SEPARACAO', ?, ?, ?)",
                (stock['id'], -float(item['quantidade_separada']), f"Saída da separação {separation['lote_nome']} - {separation['store_nome']}", separation['id'], actor_id, agora_iso()),
            )
    pendencias_restantes = sum(max(float(item['quantidade_pedida'] or 0) - float(item['quantidade_separada'] or 0), 0) for item in itens)
    conn.execute("UPDATE separations SET status = 'FINALIZADA', finalizado_em = ? WHERE id = ?", (agora_iso(), separation['id']))
    return pendencias_restantes


@app.post("/lotes/<lote_codigo>/finalizar-conferencia")
@login_required
@module_required("separacoes")
def finalizar_conferencia_lote(lote_codigo: str) -> Response:
    separacoes = carregar_lote(lote_codigo)
    if not separacoes or not pode_acessar_lote_operacao(separacoes, "conferencia"):
        flash("Lote não encontrado ou sem permissão para finalizar a conferência.", "error")
        return redirect(url_for("listar_separacoes"))
    try:
        pendencias = 0.0
        with closing(get_conn()) as conn:
            for sep in separacoes:
                if sep['status'] == 'FINALIZADA':
                    continue
                itens = conn.execute("SELECT COUNT(*) AS c FROM separation_items WHERE separation_id = ?", (sep['id'],)).fetchone()['c']
                if itens == 0:
                    continue
                pend_nao_conferidos = conn.execute("SELECT COUNT(*) AS c FROM separation_items WHERE separation_id = ? AND quantidade_separada > 0 AND status NOT IN ('CONFERIDO')", (sep['id'],)).fetchone()['c']
                if pend_nao_conferidos > 0:
                    raise ValueError(f"A loja {sep['store_nome']} ainda possui itens sem conferência.")
                pendencias += finalizar_separacao_no_conn(conn, sep, g.user['id'])
            conn.commit()
    except ValueError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('conferir_itens_lote', lote_codigo=lote_codigo))
    if pendencias > 0:
        flash('Conferência concluída. O lote foi finalizado com pendências parciais registradas para reaproveitamento futuro.', 'success')
    else:
        flash('Conferência concluída e lote finalizado.', 'success')
    return redirect(url_for('listar_separacoes'))


@app.get("/separacoes")
@login_required
@module_required("separacoes")
def listar_separacoes() -> str:
    where, params = separation_visibility_clause()
    separacoes = query_all(
        f"""
        SELECT s.*, st.nome AS store_nome,
               r.nome AS responsavel_nome,
               c.nome AS conferente_nome
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        {where}
        ORDER BY CASE s.status
            WHEN 'ABERTA' THEN 1
            WHEN 'EM_SEPARACAO' THEN 2
            WHEN 'AGUARDANDO_CONFERENCIA' THEN 3
            WHEN 'FINALIZADA' THEN 4
            ELSE 5 END,
            s.id DESC
        """,
        params,
    )
    lotes = lotes_visiveis_para_usuario(g.user)
    return render_template("separacoes.html", title="Separações", separacoes=separacoes, lotes=lotes, usar_conferente=get_setting("usar_conferente", "1") == "1")


def can_access_separation(separation: sqlite3.Row) -> bool:
    if user_is_admin(g.user):
        return True
    return g.user["id"] in {separation["responsavel_id"], separation["conferente_id"]}


def load_separation(separation_id: int) -> sqlite3.Row | None:
    return query_one(
        """
        SELECT s.*, st.nome AS store_nome,
               r.nome AS responsavel_nome,
               c.nome AS conferente_nome,
               creator.nome AS criado_por_nome
        FROM separations s
        JOIN stores st ON st.id = s.store_id
        LEFT JOIN users r ON r.id = s.responsavel_id
        LEFT JOIN users c ON c.id = s.conferente_id
        LEFT JOIN users creator ON creator.id = s.criado_por
        WHERE s.id = ?
        """,
        (separation_id,),
    )




def separation_summary(separation_id: int) -> dict[str, float]:
    row = query_one(
        "SELECT COALESCE(SUM(quantidade_pedida),0) AS qtd_pedida, COALESCE(SUM(quantidade_separada),0) AS qtd_separada FROM separation_items WHERE separation_id = ?",
        (separation_id,),
    )
    return {"qtd_pedida": row["qtd_pedida"], "qtd_separada": row["qtd_separada"]}


@app.get("/separacoes/<int:separation_id>")
@login_required
@module_required("separacoes")
def detalhe_separacao(separation_id: int) -> str | Response:
    separation = load_separation(separation_id)
    if separation is None or not can_access_separation(separation):
        flash("Separação não encontrada ou sem permissão.", "error")
        return redirect(url_for("listar_separacoes"))

    items = query_all(
        """
        SELECT si.*, COALESCE(stk.quantidade_atual, 0) AS estoque_atual
        FROM separation_items si
        LEFT JOIN stock_items stk ON stk.codigo = si.codigo
        WHERE si.separation_id = ?
        ORDER BY si.descricao COLLATE NOCASE ASC, si.id ASC
        """,
        (separation_id,),
    )
    usar_conferente = get_setting("usar_conferente", "1") == "1"
    pode_editar_itens = user_is_admin(g.user) and separation["status"] != "FINALIZADA"
    pode_separar = (user_is_admin(g.user) or (normalize_role(g.user["role"]) == "separador" and g.user["id"] in {separation["responsavel_id"], separation["criado_por"]})) and separation["status"] in {"ABERTA", "EM_SEPARACAO"}
    pode_enviar_conferencia = usar_conferente and (user_is_admin(g.user) or (normalize_role(g.user["role"]) == "separador" and g.user["id"] in {separation["responsavel_id"], separation["criado_por"]})) and separation["status"] in {"ABERTA", "EM_SEPARACAO"}
    pode_finalizar = False
    if usar_conferente:
        pode_finalizar = (user_is_admin(g.user) or normalize_role(g.user["role"]) == "conferente") and (user_is_admin(g.user) or g.user["id"] == separation["conferente_id"]) and separation["status"] == "AGUARDANDO_CONFERENCIA"
        texto_fluxo = "O admin pode lançar itens. O separador marca a quantidade separada. O conferente finaliza."
        texto_botao_finalizar = "Finalizar separação"
    else:
        pode_finalizar = (user_is_admin(g.user) or normalize_role(g.user["role"]) == "separador") and (user_is_admin(g.user) or g.user["id"] == separation["responsavel_id"]) and separation["status"] in {"ABERTA", "EM_SEPARACAO"}
        texto_fluxo = "O admin pode lançar itens. O separador marca a quantidade separada. Como o conferente está desativado, o responsável ou admin finalizam direto."
        texto_botao_finalizar = "Finalizar direto"
    return render_template(
        "detalhe_separacao.html",
        title="Detalhe da separação",
        separation=separation,
        items=items,
        resumo=separation_summary(separation_id),
        pode_editar_itens=pode_editar_itens,
        pode_separar=pode_separar,
        pode_enviar_conferencia=pode_enviar_conferencia,
        pode_finalizar=pode_finalizar,
        texto_fluxo=texto_fluxo,
        texto_botao_finalizar=texto_botao_finalizar,
    )


@app.post("/separacoes/<int:separation_id>/itens")
@login_required
@module_required("separacoes")
@roles_required("admin")
def adicionar_item_separacao(separation_id: int) -> Response:
    separation = load_separation(separation_id)
    if separation is None:
        flash("Separação não encontrada.", "error")
        return redirect(url_for("listar_separacoes"))
    if separation["status"] == "FINALIZADA":
        flash("Não é possível alterar itens de uma separação finalizada.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation_id))

    codigo = request.form.get("codigo", "").strip()
    descricao = request.form.get("descricao", "").strip()
    if not codigo or not descricao:
        flash("Informe código e descrição do item.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation_id))
    try:
        fator_embalagem = parse_fator_embalagem(request.form.get("fator_embalagem", "1"))
        quantidade_base = parse_float(request.form.get("quantidade_pedida", ""), "Quantidade pedida")
        if quantidade_base <= 0:
            raise ValueError("Quantidade pedida deve ser maior que zero.")
        quantidade_pedida = quantidade_base * fator_embalagem
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation_id))

    stock = query_one("SELECT custo_unitario FROM stock_items WHERE codigo = ?", (codigo,))
    custo_ref = stock["custo_unitario"] if stock else 0
    with closing(get_conn()) as conn:
        conn.execute(
            """
            INSERT INTO separation_items (separation_id, codigo, descricao, fator_embalagem, quantidade_pedida, quantidade_separada, status, custo_unitario_ref, criado_em, atualizado_em)
            VALUES (?, ?, ?, ?, ?, 0, 'PENDENTE', ?, ?, ?)
            """,
            (separation_id, codigo, descricao, fator_embalagem, quantidade_pedida, custo_ref, agora_iso(), agora_iso()),
        )
        conn.commit()
    flash("Item do pedido adicionado.", "success")
    return redirect(url_for("detalhe_separacao", separation_id=separation_id))


@app.post("/separacoes/item/<int:item_id>/atualizar")
@login_required
@module_required("separacoes")
def atualizar_item_separacao(item_id: int) -> Response:
    separation_id = request.form.get("separation_id", "").strip()
    if not separation_id.isdigit():
        flash("Separação inválida.", "error")
        return redirect(url_for("listar_separacoes"))
    separation = load_separation(int(separation_id))
    if separation is None or not can_access_separation(separation):
        flash("Sem acesso a essa separação.", "error")
        return redirect(url_for("listar_separacoes"))
    if separation["status"] not in {"ABERTA", "EM_SEPARACAO"}:
        flash("Essa separação não aceita mais alterações de quantidade.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation["id"]))
    if (not user_is_admin(g.user) and normalize_role(g.user["role"]) != "separador") or (not user_is_admin(g.user) and normalize_role(g.user["role"]) == "separador" and g.user["id"] != separation["responsavel_id"]):
        flash("Somente o responsável ou admin podem informar a quantidade separada.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation["id"]))

    try:
        qtd_sep = parse_float(request.form.get("quantidade_separada", ""), "Quantidade separada")
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation["id"]))

    with closing(get_conn()) as conn:
        item = conn.execute("SELECT quantidade_pedida FROM separation_items WHERE id = ? AND separation_id = ?", (item_id, separation["id"])).fetchone()
        if item is None:
            flash("Item não encontrado.", "error")
            return redirect(url_for("detalhe_separacao", separation_id=separation["id"]))
        status = "PENDENTE"
        if qtd_sep <= 0:
            status = "PENDENTE"
        elif qtd_sep < float(item["quantidade_pedida"]):
            status = "PARCIAL"
        else:
            status = "SEPARADO"
        conn.execute(
            "UPDATE separation_items SET quantidade_separada = ?, status = ?, atualizado_em = ? WHERE id = ?",
            (qtd_sep, status, agora_iso(), item_id),
        )
        conn.execute("UPDATE separations SET status = 'EM_SEPARACAO' WHERE id = ? AND status = 'ABERTA'", (separation["id"],))
        conn.commit()
    flash("Quantidade separada atualizada.", "success")
    return redirect(url_for("detalhe_separacao", separation_id=separation["id"]))


@app.post("/separacoes/item/<int:item_id>/excluir")
@login_required
@module_required("separacoes")
@roles_required("admin")
def excluir_item_separacao(item_id: int) -> Response:
    separation_id = request.form.get("separation_id", "").strip()
    if not separation_id.isdigit():
        flash("Separação inválida.", "error")
        return redirect(url_for("listar_separacoes"))
    with closing(get_conn()) as conn:
        item = conn.execute("SELECT carryover_source_item_id FROM separation_items WHERE id = ? AND separation_id = ?", (item_id, int(separation_id))).fetchone()
        if item and item["carryover_source_item_id"]:
            conn.execute("UPDATE separation_items SET carryover_copied = 0, atualizado_em = ? WHERE id = ?", (agora_iso(), item["carryover_source_item_id"]))
        conn.execute("DELETE FROM separation_items WHERE id = ? AND separation_id = ?", (item_id, int(separation_id)))
        conn.commit()
    flash("Item removido da separação.", "success")
    return redirect(url_for("detalhe_separacao", separation_id=int(separation_id)))


@app.post("/separacoes/<int:separation_id>/excluir")
@login_required
@module_required("separacoes")
@roles_required("admin")
def excluir_separacao(separation_id: int) -> Response:
    separation = load_separation(separation_id)
    if separation is None:
        flash("Loja da separação não encontrada.", "error")
        return redirect(url_for("listar_separacoes"))
    if separation["status"] == "FINALIZADA":
        flash("Não é possível remover uma loja já finalizada. Use o histórico para manter os dados.", "error")
        return redirect(url_for("listar_separacoes"))

    operacao_chave = lote_operacao_chave_row(separation)
    with closing(get_conn()) as conn:
        desfazer_pendencias_transferidas(conn, separation_id)
        conn.execute("DELETE FROM separations WHERE id = ?", (separation_id,))
        conn.commit()

    restantes = carregar_lote(operacao_chave)
    if restantes:
        flash(f"Loja {separation['store_nome']} removida da separação.", "success")
        return redirect(url_for("grade_lote", lote_codigo=operacao_chave))

    flash("Loja removida. Como era a última do lote, o lote também saiu da lista.", "success")
    return redirect(url_for("listar_separacoes"))


@app.post("/lotes/<lote_codigo>/excluir")
@login_required
@module_required("lotes")
@roles_required("admin")
def excluir_lote(lote_codigo: str) -> Response:
    separacoes = carregar_lote(lote_codigo)
    if not separacoes:
        flash("Lote não encontrado.", "error")
        return redirect(url_for("listar_lotes"))

    finalizadas = [sep["store_nome"] for sep in separacoes if sep["status"] == "FINALIZADA"]
    if finalizadas:
        flash(
            "Não dá para excluir o lote inteiro porque existe loja finalizada nele: " + ", ".join(finalizadas) + ". Remova apenas as lojas abertas.",
            "error",
        )
        return redirect(url_for("grade_lote", lote_codigo=lote_codigo))

    with closing(get_conn()) as conn:
        for sep in separacoes:
            desfazer_pendencias_transferidas(conn, sep["id"])
        conn.execute(
            f"DELETE FROM separations WHERE id IN ({','.join('?' for _ in separacoes)})",
            tuple(sep["id"] for sep in separacoes),
        )
        conn.commit()

    flash(f"Lote removido com {len(separacoes)} loja(s).", "success")
    return redirect(url_for("listar_lotes"))


@app.post("/separacoes/<int:separation_id>/enviar-conferencia")
@login_required
@module_required("separacoes")
def enviar_conferencia(separation_id: int) -> Response:
    separation = load_separation(separation_id)
    if separation is None or not can_access_separation(separation):
        flash("Sem acesso a essa separação.", "error")
        return redirect(url_for("listar_separacoes"))
    if (not user_is_admin(g.user) and normalize_role(g.user["role"]) != "separador") or (not user_is_admin(g.user) and normalize_role(g.user["role"]) == "separador" and g.user["id"] != separation["responsavel_id"]):
        flash("Somente o responsável ou admin podem enviar para conferência.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation_id))

    if get_setting("usar_conferente", "1") != "1":
        flash("A função do conferente está desativada. Finalize direto pela própria separação.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation_id))

    total_itens = query_one("SELECT COUNT(*) AS c FROM separation_items WHERE separation_id = ?", (separation_id,))["c"]
    if total_itens == 0:
        flash("Adicione itens antes de enviar para conferência.", "error")
        return redirect(url_for("detalhe_separacao", separation_id=separation_id))

    with closing(get_conn()) as conn:
        conn.execute(
            "UPDATE separations SET status = 'AGUARDANDO_CONFERENCIA', enviado_conferencia_em = ? WHERE id = ?",
            (agora_iso(), separation_id),
        )
        conn.commit()
    flash("Separação enviada para conferência.", "success")
    return redirect(url_for("detalhe_separacao", separation_id=separation_id))


def validar_estoque_para_finalizacao(conn: sqlite3.Connection, separation_id: int) -> list[str]:
    problemas: list[str] = []
    itens = conn.execute(
        "SELECT codigo, descricao, quantidade_separada FROM separation_items WHERE separation_id = ?",
        (separation_id,),
    ).fetchall()
    for item in itens:
        stock = conn.execute("SELECT quantidade_atual FROM stock_items WHERE codigo = ?", (item["codigo"],)).fetchone()
        saldo = float(stock["quantidade_atual"]) if stock else 0.0
        if saldo < float(item["quantidade_separada"]):
            problemas.append(f"{item['descricao']} (saldo {fmt_num(saldo)} / separado {fmt_num(item['quantidade_separada'])})")
    return problemas


@app.post("/separacoes/<int:separation_id>/finalizar")
@login_required
@module_required("separacoes")
def finalizar_separacao(separation_id: int) -> Response:
    separation = load_separation(separation_id)
    if separation is None or not can_access_separation(separation):
        flash("Sem acesso a essa separação.", "error")
        return redirect(url_for("listar_separacoes"))

    usar_conferente = get_setting("usar_conferente", "1") == "1"
    if usar_conferente:
        if separation["status"] != "AGUARDANDO_CONFERENCIA":
            flash("A separação precisa estar aguardando conferência.", "error")
            return redirect(url_for("detalhe_separacao", separation_id=separation_id))
        if (not user_is_admin(g.user) and normalize_role(g.user["role"]) != "conferente") or (not user_is_admin(g.user) and normalize_role(g.user["role"]) == "conferente" and g.user["id"] != separation["conferente_id"]):
            flash("Somente o conferente designado ou admin podem finalizar.", "error")
            return redirect(url_for("detalhe_separacao", separation_id=separation_id))
    else:
        if separation["status"] not in {"ABERTA", "EM_SEPARACAO"}:
            flash("Com o conferente desligado, só é possível finalizar separações ainda em andamento.", "error")
            return redirect(url_for("detalhe_separacao", separation_id=separation_id))
        if (not user_is_admin(g.user) and normalize_role(g.user["role"]) != "separador") or (not user_is_admin(g.user) and normalize_role(g.user["role"]) == "separador" and g.user["id"] != separation["responsavel_id"]):
            flash("Somente o responsável ou admin podem finalizar quando o conferente estiver desligado.", "error")
            return redirect(url_for("detalhe_separacao", separation_id=separation_id))

    with closing(get_conn()) as conn:
        itens = conn.execute("SELECT * FROM separation_items WHERE separation_id = ?", (separation_id,)).fetchall()
        if not itens:
            flash("Essa separação não possui itens.", "error")
            return redirect(url_for("detalhe_separacao", separation_id=separation_id))

        usar_controle_global = get_setting("vincular_estoque", "1") == "1"
        precisa_abater = usar_controle_global and bool(separation["usar_estoque"])

        if precisa_abater:
            problemas = validar_estoque_para_finalizacao(conn, separation_id)
            if problemas:
                flash("Saldo insuficiente para finalizar com controle de estoque: " + "; ".join(problemas), "error")
                return redirect(url_for("detalhe_separacao", separation_id=separation_id))

            for item in itens:
                stock = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item["codigo"],)).fetchone()
                if stock is None:
                    conn.execute(
                        "INSERT INTO stock_items (codigo, descricao, quantidade_atual, custo_unitario, atualizado_em) VALUES (?, ?, 0, ?, ?)",
                        (item["codigo"], item["descricao"], item["custo_unitario_ref"], agora_iso()),
                    )
                    stock = conn.execute("SELECT * FROM stock_items WHERE codigo = ?", (item["codigo"],)).fetchone()
                novo_saldo = float(stock["quantidade_atual"]) - float(item["quantidade_separada"])
                conn.execute(
                    "UPDATE stock_items SET quantidade_atual = ?, atualizado_em = ? WHERE id = ?",
                    (novo_saldo, agora_iso(), stock["id"]),
                )
                conn.execute(
                    "INSERT INTO stock_movements (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em) VALUES (?, 'SAIDA_SEPARACAO', ?, ?, 'SEPARACAO', ?, ?, ?)",
                    (stock["id"], -float(item["quantidade_separada"]), f"Saída da separação {separation['lote_nome']} - {separation['store_nome']}", separation_id, g.user["id"], agora_iso()),
                )

        pendencias_restantes = sum(
            max(float(item["quantidade_pedida"] or 0) - float(item["quantidade_separada"] or 0), 0)
            for item in itens
        )
        conn.execute(
            "UPDATE separations SET status = 'FINALIZADA', finalizado_em = ? WHERE id = ?",
            (agora_iso(), separation_id),
        )
        conn.commit()

    if pendencias_restantes > 0:
        flash(
            "Separação finalizada parcialmente. O restante pendente poderá ser puxado automaticamente no próximo dia ao criar uma nova separação para essa loja.",
            "success",
        )
    else:
        flash("Separação finalizada e registrada no histórico.", "success")
    return redirect(url_for("detalhe_separacao", separation_id=separation_id))




@app.post("/relatorios/<int:separation_id>/apagar")
@login_required
@module_required("relatorios")
@roles_required("admin")
def apagar_historico_separacao(separation_id: int) -> Response:
    with closing(get_conn()) as conn:
        try:
            separation = conn.execute("SELECT status FROM separations WHERE id = ?", (separation_id,)).fetchone()
            if separation is None:
                raise ValueError("Separação não encontrada.")
            if separation["status"] == "CANCELADA":
                excluir_separacao_cancelada_no_conn(conn, separation_id)
                mensagem = "Separação cancelada excluída em definitivo."
            else:
                apagar_historico_separacao_no_conn(conn, separation_id, g.user["id"])
                mensagem = "Registro removido do histórico. Se usava estoque, o saldo foi estornado."
            conn.commit()
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("relatorios"))
    flash(mensagem, "success")
    return redirect(url_for("relatorios"))


@app.post("/relatorios/lotes/<lote_codigo>/apagar")
@login_required
@module_required("relatorios")
@roles_required("admin")
def apagar_historico_lote(lote_codigo: str) -> Response:
    separacoes = [s for s in carregar_lote_completo(lote_codigo) if s["status"] in {"FINALIZADA", "CANCELADA"}]
    if not separacoes:
        flash("Nenhum registro histórico encontrado nesse lote.", "error")
        return redirect(url_for("relatorios"))
    canceladas = 0
    excluidas = 0
    with closing(get_conn()) as conn:
        try:
            for sep in separacoes:
                if sep["status"] == "CANCELADA":
                    excluir_separacao_cancelada_no_conn(conn, sep["id"])
                    excluidas += 1
                else:
                    apagar_historico_separacao_no_conn(conn, sep["id"], g.user["id"])
                    canceladas += 1
            conn.commit()
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("relatorios"))
    partes = []
    if canceladas:
        partes.append(f"{canceladas} loja(s) cancelada(s) com estorno")
    if excluidas:
        partes.append(f"{excluidas} loja(s) cancelada(s) excluída(s) em definitivo")
    flash("Lote processado: " + "; ".join(partes) + ".", "success")
    return redirect(url_for("relatorios"))


@app.get("/relatorios/lotes/<lote_codigo>")
@login_required
def detalhe_historico_lote(lote_codigo: str) -> str | Response:
    if not (user_has_access(g.user, "relatorios") and (user_is_admin(g.user) or normalize_role(g.user["role"]) == "balanco")):
        return forbidden_redirect("Sem permissão para acessar o histórico.")
    historico = [s for s in carregar_lote_completo(lote_codigo) if s["status"] in {"FINALIZADA", "CANCELADA"}]
    if not historico:
        flash("Lote histórico não encontrado.", "error")
        return redirect(url_for("relatorios"))
    itens_por_separacao: dict[int, list[sqlite3.Row]] = {}
    for sep in historico:
        itens_por_separacao[sep["id"]] = query_all(
            "SELECT codigo, descricao, fator_embalagem, quantidade_pedida, quantidade_separada, custo_unitario_ref FROM separation_items WHERE separation_id = ? ORDER BY descricao COLLATE NOCASE ASC",
            (sep["id"],),
        )
    return render_template(
        "historico_lote.html",
        title="Histórico do lote",
        historico=historico,
        itens_por_separacao=itens_por_separacao,
    )


def _where_linha_relatorio(alias: str, linha_base: str = "", sublinha: str = "") -> tuple[str, list[Any]]:
    conds: list[str] = []
    params: list[Any] = []
    campos = f"LOWER(COALESCE({alias}.linha_caminho_erp, '') || ' ' || COALESCE({alias}.linha_erp, '') || ' ' || COALESCE({alias}.descricao, ''))"
    for termo in (linha_base, sublinha):
        termo = str(termo or "").strip()
        if not termo:
            continue
        palavras = [p for p in re.split(r"[\s\-/]+", termo.casefold()) if len(p) >= 2]
        for palavra in palavras:
            conds.append(f"{campos} LIKE ?")
            params.append(f"%{palavra}%")
    if not conds:
        return "", []
    return " AND " + " AND ".join(conds), params


def _dados_relatorio_gerencial(periodo_dias: int = 30, linha_base: str = "Padaria - Industria CD", sublinha: str = "") -> dict[str, Any]:
    periodo = max(1, min(int(periodo_dias or 30), 365))
    linha_base = str(linha_base or "Padaria - Industria CD").strip() or "Padaria - Industria CD"
    sublinha = str(sublinha or "").strip()
    where_st, params_st = _where_linha_relatorio("st", linha_base, sublinha)
    where_base, params_base = _where_linha_relatorio("st", linha_base, "")
    with closing(get_conn()) as conn:
        resumo = conn.execute(f"""
            SELECT COUNT(DISTINCT s.id) AS separacoes, COUNT(DISTINCT s.lote_codigo) AS lotes,
                   COUNT(DISTINCT s.store_id) AS lojas, COALESCE(SUM(si.quantidade_separada), 0) AS quantidade,
                   COALESCE(SUM(si.quantidade_separada * si.custo_unitario_ref), 0) AS custo
            FROM separations s
            LEFT JOIN separation_items si ON si.separation_id = s.id
            LEFT JOIN stock_items st ON st.codigo = si.codigo
            WHERE date(COALESCE(s.finalizado_em, s.criado_em)) >= date('now', ?) {where_st}
        """, [f"-{periodo} days", *params_st]).fetchone()
        estoque = conn.execute(f"""
            SELECT COUNT(*) AS produtos,
                   SUM(CASE WHEN quantidade_atual <= 0 THEN 1 ELSE 0 END) AS zerados,
                   SUM(CASE WHEN quantidade_atual > 0 AND quantidade_atual <= 10 THEN 1 ELSE 0 END) AS baixo,
                   COALESCE(SUM(quantidade_atual * custo_unitario), 0) AS valor_estoque
            FROM stock_items st WHERE ativo = 1 {where_st}
        """, params_st).fetchone()
        top_itens = conn.execute(f"""
            SELECT si.codigo, si.descricao, COALESCE(st.linha_erp, '') AS linha_erp,
                   SUM(si.quantidade_separada) AS quantidade, SUM(si.quantidade_separada * si.custo_unitario_ref) AS custo
            FROM separation_items si
            JOIN separations s ON s.id = si.separation_id
            LEFT JOIN stock_items st ON st.codigo = si.codigo
            WHERE date(COALESCE(s.finalizado_em, s.criado_em)) >= date('now', ?) {where_st}
            GROUP BY si.codigo, si.descricao, st.linha_erp
            ORDER BY quantidade DESC LIMIT 30
        """, [f"-{periodo} days", *params_st]).fetchall()
        criticos = conn.execute(f"""
            SELECT codigo, descricao, linha_erp, quantidade_atual, custo_unitario, atualizado_em
            FROM stock_items st
            WHERE ativo = 1 AND quantidade_atual <= 10 {where_st}
            ORDER BY quantidade_atual ASC, descricao COLLATE NOCASE ASC LIMIT 60
        """, params_st).fetchall()
        lojas = conn.execute(f"""
            SELECT store.nome AS loja, SUM(si.quantidade_separada) AS quantidade, SUM(si.quantidade_separada * si.custo_unitario_ref) AS custo
            FROM separations s
            JOIN stores store ON store.id = s.store_id
            LEFT JOIN separation_items si ON si.separation_id = s.id
            LEFT JOIN stock_items st ON st.codigo = si.codigo
            WHERE date(COALESCE(s.finalizado_em, s.criado_em)) >= date('now', ?) {where_st}
            GROUP BY store.id, store.nome ORDER BY quantidade DESC LIMIT 20
        """, [f"-{periodo} days", *params_st]).fetchall()
        sublinhas = conn.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(linha_erp), ''), 'Sem linha') AS linha, COUNT(*) AS produtos,
                   COALESCE(SUM(quantidade_atual), 0) AS saldo,
                   COALESCE(SUM(quantidade_atual * custo_unitario), 0) AS valor
            FROM stock_items st WHERE ativo = 1 {where_base}
            GROUP BY COALESCE(NULLIF(TRIM(linha_erp), ''), 'Sem linha')
            ORDER BY linha COLLATE NOCASE ASC
        """, params_base).fetchall()
        linhas_resumo = conn.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(linha_erp), ''), 'Sem linha') AS linha, COUNT(*) AS produtos,
                   SUM(CASE WHEN quantidade_atual <= 0 THEN 1 ELSE 0 END) AS zerados,
                   SUM(CASE WHEN quantidade_atual > 0 AND quantidade_atual <= 10 THEN 1 ELSE 0 END) AS baixo,
                   COALESCE(SUM(quantidade_atual), 0) AS saldo,
                   COALESCE(SUM(quantidade_atual * custo_unitario), 0) AS valor
            FROM stock_items st WHERE ativo = 1 {where_st}
            GROUP BY COALESCE(NULLIF(TRIM(linha_erp), ''), 'Sem linha')
            ORDER BY valor DESC, linha COLLATE NOCASE ASC LIMIT 50
        """, params_st).fetchall()
    observacoes = []
    filtro_nome = linha_base + ((" / " + sublinha) if sublinha else "")
    if int(estoque["zerados"] or 0) > 0:
        observacoes.append(f"Na linha {filtro_nome}, existem {int(estoque['zerados'] or 0)} item(ns) zerado(s) que merecem conferência.")
    if int(estoque["baixo"] or 0) > 0:
        observacoes.append(f"Na linha {filtro_nome}, existem {int(estoque['baixo'] or 0)} item(ns) com saldo baixo até 10 unidades.")
    if not observacoes:
        observacoes.append(f"Nenhum alerta crítico foi encontrado para {filtro_nome}.")
    return {"periodo": periodo, "gerado_em": agora_br(), "linha_base": linha_base, "sublinha": sublinha, "resumo": resumo, "estoque": estoque, "top_itens": top_itens, "criticos": criticos, "lojas": lojas, "sublinhas": sublinhas, "linhas_resumo": linhas_resumo, "observacoes": observacoes}


def _dados_resumo_padaria_industria_cd() -> dict[str, Any]:
    """Resumo simples da Padaria - Industria CD para a aba Relatórios.
    Não usa histórico de lote; olha direto para o estoque atual.
    """
    linha_base = "Padaria - Industria CD"
    where_st, params_st = _where_linha_relatorio("st", linha_base, "")
    with closing(get_conn()) as conn:
        estoque = conn.execute(f"""
            SELECT COUNT(*) AS produtos,
                   SUM(CASE WHEN quantidade_atual <= 0 THEN 1 ELSE 0 END) AS zerados,
                   SUM(CASE WHEN quantidade_atual > 0 AND quantidade_atual <= 10 THEN 1 ELSE 0 END) AS baixo,
                   COALESCE(SUM(quantidade_atual), 0) AS quantidade_total,
                   COALESCE(SUM(quantidade_atual * custo_unitario), 0) AS valor_estoque
            FROM stock_items st
            WHERE ativo = 1 {where_st}
        """, params_st).fetchone()
        linhas = conn.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(linha_erp), ''), 'Sem linha') AS linha,
                   COUNT(*) AS produtos,
                   SUM(CASE WHEN quantidade_atual <= 0 THEN 1 ELSE 0 END) AS zerados,
                   SUM(CASE WHEN quantidade_atual > 0 AND quantidade_atual <= 10 THEN 1 ELSE 0 END) AS baixo,
                   COALESCE(SUM(quantidade_atual), 0) AS saldo,
                   COALESCE(SUM(quantidade_atual * custo_unitario), 0) AS valor
            FROM stock_items st
            WHERE ativo = 1 {where_st}
            GROUP BY COALESCE(NULLIF(TRIM(linha_erp), ''), 'Sem linha')
            ORDER BY linha COLLATE NOCASE ASC
        """, params_st).fetchall()
    return {"linha_base": linha_base, "estoque": estoque, "linhas": linhas, "gerado_em": agora_br()}


@app.get("/relatorios/gerencial")
@login_required
@module_required("relatorios")
def relatorio_gerencial() -> str:
    periodo = request.args.get("periodo", "30")
    linha_base = request.args.get("linha_base", "Padaria - Industria CD")
    sublinha = request.args.get("sublinha", "")
    dados = _dados_relatorio_gerencial(int(periodo) if str(periodo).isdigit() else 30, linha_base, sublinha)
    registrar_auditoria("visualizar_relatorio_gerencial", "relatorio", str(dados["periodo"]), {"linha_base": linha_base, "sublinha": sublinha})
    return render_template("relatorio_gerencial.html", title="Relatório gerencial", dados=dados)


@app.get("/relatorios/gerencial/pdf")
@login_required
@module_required("relatorios")
def relatorio_gerencial_pdf() -> Response:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    periodo = request.args.get("periodo", "30")
    linha_base = request.args.get("linha_base", "Padaria - Industria CD")
    sublinha = request.args.get("sublinha", "")
    dados = _dados_relatorio_gerencial(int(periodo) if str(periodo).isdigit() else 30, linha_base, sublinha)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story: list[Any] = [Paragraph("Relatório gerencial - Sistema Alvorada", styles["Title"]), Paragraph(f"Período: últimos {dados['periodo']} dias • Gerado em {dados['gerado_em']}", styles["Normal"]), Spacer(1, .3*cm)]
    resumo = dados["resumo"]; estoque = dados["estoque"]
    cards = [["Indicador", "Valor"], ["Separações", str(resumo["separacoes"] or 0)], ["Lotes", str(resumo["lotes"] or 0)], ["Qtd. separada", fmt_num(resumo["quantidade"] or 0)], ["Custo separado", fmt_money(resumo["custo"] or 0)], ["Produtos ativos", str(estoque["produtos"] or 0)], ["Itens zerados", str(estoque["zerados"] or 0)], ["Estoque baixo", str(estoque["baixo"] or 0)], ["Valor estimado estoque", fmt_money(estoque["valor_estoque"] or 0)]]
    t = Table(cards, repeatRows=1)
    t.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2F5D2E")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.grey), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")]))
    story.append(t); story.append(Spacer(1, .35*cm))
    story.append(Paragraph("Observações automáticas", styles["Heading2"]))
    for obs in dados["observacoes"]:
        story.append(Paragraph("• " + obs, styles["Normal"]))
    if dados["criticos"]:
        story.append(Spacer(1, .25*cm)); story.append(Paragraph("Itens críticos", styles["Heading2"]))
        rows = [["Código", "Descrição", "Saldo"]] + [[r["codigo"], r["descricao"][:70], fmt_num(r["quantidade_atual"])] for r in dados["criticos"][:20]]
        t2 = Table(rows, repeatRows=1)
        t2.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2F5D2E")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.grey), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")]))
        story.append(t2)
    doc.build(story)
    buffer.seek(0)
    registrar_auditoria("exportar_relatorio_gerencial_pdf", "relatorio", str(dados["periodo"]), {})
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=f"relatorio_gerencial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")


@app.get("/relatorios")
@login_required
@module_required("relatorios")
def relatorios() -> str:
    if not (user_has_access(g.user, "relatorios") and (user_is_admin(g.user) or normalize_role(g.user["role"]) == "balanco")):
        return forbidden_redirect("Somente admin ou balanço podem acessar os relatórios.")
    dados_padaria = _dados_resumo_padaria_industria_cd()
    registrar_auditoria("visualizar_relatorios_padaria_industria_cd", "relatorio", "padaria_industria_cd", {})
    return render_template("relatorios.html", title="Relatórios", dados_padaria=dados_padaria)


@app.get("/estoque/historico/exportar.xlsx")
@login_required
@module_required("estoque")
def exportar_historico_estoque_excel() -> Response:
    filters = sanitize_stock_history_filters(request.args)
    movimentos = fetch_stock_movements(filters, limit=None)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        flash("Para exportar em Excel, instale a dependência openpyxl.", "error")
        return redirect(url_for("estoque", **request.args.to_dict(flat=True)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico estoque"

    filtros_aplicados = stock_history_filter_labels(filters)
    ws.append(["Histórico de movimentações do estoque"])
    ws.append(["Gerado em", agora_br()])
    ws.append(["Filtros", " | ".join(filtros_aplicados) if filtros_aplicados else "Sem filtros específicos"])
    ws.append([])

    headers = [
        "Data/Hora",
        "Tipo",
        "Código",
        "Código de barras",
        "Descrição",
        "Quantidade",
        "Usuário",
        "Observação",
        "Referência",
    ]
    ws.append(headers)
    for cell in ws[5]:
        cell.font = Font(bold=True)

    for mov in movimentos:
        usuario = mov["usuario_nome"] or mov["usuario_login"] or "Sistema"
        referencia = f"{mov['referencia_tipo'] or '-'} {mov['referencia_id'] or ''}".strip()
        ws.append([
            mov["criado_em"],
            stock_movement_label(mov["tipo"]),
            mov["codigo"],
            mov["codigo_barras"] or "",
            mov["descricao"],
            float(mov["quantidade"] or 0),
            usuario,
            mov["observacao"] or "",
            referencia,
        ])

    column_widths = {
        "A": 19,
        "B": 22,
        "C": 14,
        "D": 18,
        "E": 42,
        "F": 14,
        "G": 24,
        "H": 36,
        "I": 18,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=stock_history_export_filename("xlsx"),
    )


@app.get("/estoque/historico/exportar.pdf")
@login_required
@module_required("estoque")
def exportar_historico_estoque_pdf() -> Response:
    filters = sanitize_stock_history_filters(request.args)
    movimentos = fetch_stock_movements(filters, limit=None)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle
    except ImportError:
        flash("Para exportar em PDF, instale a dependência reportlab.", "error")
        return redirect(url_for("estoque", **request.args.to_dict(flat=True)))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Histórico de movimentações do estoque", styles["Title"]),
        Paragraph(f"Gerado em: {agora_br()}", styles["Normal"]),
    ]

    filtros_aplicados = stock_history_filter_labels(filters)
    story.append(Paragraph("Filtros: " + (" | ".join(filtros_aplicados) if filtros_aplicados else "Sem filtros específicos"), styles["Normal"]))
    story.append(Spacer(1, 10))

    data = [["Data/Hora", "Tipo", "Código", "Descrição", "Qtd.", "Usuário", "Observação"]]
    for mov in movimentos:
        usuario = mov["usuario_nome"] or mov["usuario_login"] or "Sistema"
        descricao = f"{mov['descricao']} ({mov['codigo']})"
        data.append([
            mov["criado_em"],
            stock_movement_label(mov["tipo"]),
            mov["codigo"],
            descricao,
            fmt_num(mov["quantidade"]),
            usuario,
            mov["observacao"] or "-",
        ])

    table = LongTable(data, repeatRows=1, colWidths=[80, 95, 55, 240, 50, 110, 130])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#eef2f7")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=stock_history_export_filename("pdf"),
    )


@app.get("/api/produto")
@login_required
@module_required("estoque")
def api_produto() -> Response:
    termo = request.args.get("codigo", "").strip()
    if not termo:
        return jsonify({"ok": False, "descricao": "", "quantidade_atual": 0})

    item = query_one(
        """
        SELECT *
        FROM stock_items
        WHERE ativo = 1
          AND (codigo = ? OR codigo_barras = ?)
        LIMIT 1
        """,
        (termo, termo),
    )

    if item is None:
        return jsonify({"ok": False, "descricao": "", "quantidade_atual": 0})

    return jsonify(
        {
            "ok": True,
            "codigo": item["codigo"],
            "codigo_barras": item["codigo_barras"],
            "descricao": item["descricao"],
            "fator_embalagem": item["fator_embalagem"],
            "quantidade_atual": item["quantidade_atual"],
            "custo_unitario": item["custo_unitario"],
        }
    )



# -----------------------------------------------------------------------------
# Balanço / Contagem de estoque
# -----------------------------------------------------------------------------

def _balance_row(balance_id: int) -> sqlite3.Row | None:
    return query_one(
        """
        SELECT b.*, u.nome AS criado_por_nome, uc.nome AS confirmado_por_nome
        FROM balance_counts b
        LEFT JOIN users u ON u.id = b.criado_por
        LEFT JOIN users uc ON uc.id = b.confirmado_por
        WHERE b.id = ?
        """,
        (balance_id,),
    )


def _balance_items(balance_id: int) -> list[sqlite3.Row]:
    return query_all(
        """
        SELECT *
        FROM balance_count_items
        WHERE balance_count_id = ?
        ORDER BY linha_erp IS NULL, linha_erp, descricao
        """,
        (balance_id,),
    )


def _balance_summary(items: list[sqlite3.Row]) -> dict[str, Any]:
    total = len(items)
    divergentes = sum(1 for i in items if abs(float(i["delta"] or 0)) > 0.000001)
    sobras = sum(1 for i in items if float(i["delta"] or 0) > 0)
    faltas = sum(1 for i in items if float(i["delta"] or 0) < 0)
    valor_delta = sum(float(i["delta"] or 0) * float(i["custo_unitario"] or 0) for i in items)
    return {"total": total, "divergentes": divergentes, "sobras": sobras, "faltas": faltas, "valor_delta": valor_delta}


@app.route("/balanco", methods=["GET", "POST"])
@login_required
@module_required("balanco")
def balancos() -> str | Response:
    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip() or f"Balanço {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        observacao = request.form.get("observacao", "").strip()
        with closing(get_conn()) as conn:
            cur = conn.execute(
                "INSERT INTO balance_counts (titulo, observacao, status, criado_por, criado_em) VALUES (?, ?, 'ABERTO', ?, ?)",
                (titulo, observacao, g.user["id"], agora_iso()),
            )
            conn.commit()
            balance_id = cur.lastrowid
        registrar_auditoria("criar_balanco", "balance_counts", str(balance_id), {"titulo": titulo})
        flash("Contagem de estoque criada. Agora registre as mercadorias contadas.", "success")
        return redirect(url_for("detalhe_balanco", balance_id=balance_id))

    balances = query_all(
        """
        SELECT b.*, u.nome AS criado_por_nome,
               (SELECT COUNT(*) FROM balance_count_items i WHERE i.balance_count_id = b.id) AS total_itens,
               (SELECT COUNT(*) FROM balance_count_items i WHERE i.balance_count_id = b.id AND ABS(i.delta) > 0.000001) AS divergencias
        FROM balance_counts b
        LEFT JOIN users u ON u.id = b.criado_por
        ORDER BY b.id DESC
        LIMIT 80
        """
    )
    return render_template("balancos.html", title="Balanço de estoque", balances=balances)


@app.route("/balanco/<int:balance_id>", methods=["GET", "POST"])
@login_required
@module_required("balanco")
def detalhe_balanco(balance_id: int) -> str | Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        flash("Balanço não encontrado.", "error")
        return redirect(url_for("balancos"))

    if request.method == "POST":
        if balanco["status"] != "ABERTO":
            flash("Este balanço já foi confirmado/fechado e não aceita novas contagens.", "error")
            return redirect(url_for("detalhe_balanco", balance_id=balance_id))
        codigo = request.form.get("codigo", "").strip()
        quantidade_raw = request.form.get("quantidade_contada", "").strip()
        modo = request.form.get("modo", "substituir")
        if not codigo:
            flash("Informe o código ou código de barras do produto.", "error")
            return redirect(url_for("detalhe_balanco", balance_id=balance_id))
        try:
            quantidade = parse_float(quantidade_raw, "Quantidade contada")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("detalhe_balanco", balance_id=balance_id))

        with closing(get_conn()) as conn:
            item = conn.execute(
                """
                SELECT * FROM stock_items
                WHERE ativo = 1 AND (codigo = ? OR codigo_barras = ?)
                LIMIT 1
                """,
                (codigo, codigo),
            ).fetchone()
            if item is None:
                flash("Produto não encontrado no estoque.", "error")
                return redirect(url_for("detalhe_balanco", balance_id=balance_id))
            existente = conn.execute(
                "SELECT * FROM balance_count_items WHERE balance_count_id = ? AND stock_item_id = ?",
                (balance_id, item["id"]),
            ).fetchone()
            agora = agora_iso()
            sistema = float(item["quantidade_atual"] or 0)
            if existente:
                nova_contada = float(existente["quantidade_contada"] or 0) + quantidade if modo == "somar" else quantidade
                conn.execute(
                    """
                    UPDATE balance_count_items
                    SET quantidade_contada = ?, delta = ?, atualizado_em = ?
                    WHERE id = ?
                    """,
                    (nova_contada, nova_contada - float(existente["quantidade_sistema"] or 0), agora, existente["id"]),
                )
                flash("Item atualizado na contagem.", "success")
            else:
                conn.execute(
                    """
                    INSERT INTO balance_count_items
                    (balance_count_id, stock_item_id, codigo, descricao, linha_erp, quantidade_sistema, quantidade_contada, delta, custo_unitario, criado_em, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (balance_id, item["id"], item["codigo"], item["descricao"], item["linha_erp"], sistema, quantidade, quantidade - sistema, item["custo_unitario"] or 0, agora, agora),
                )
                flash("Item registrado na contagem.", "success")
            conn.commit()
        registrar_auditoria("registrar_item_balanco", "balance_counts", str(balance_id), {"codigo": codigo, "quantidade": quantidade})
        return redirect(url_for("detalhe_balanco", balance_id=balance_id))

    items = _balance_items(balance_id)
    resumo = _balance_summary(items)
    return render_template("balanco_detalhe.html", title=f"Balanço #{balance_id}", balanco=balanco, items=items, resumo=resumo)



@app.get("/balanco/<int:balance_id>/checar-codigo")
@login_required
@module_required("balanco")
def checar_codigo_balanco(balance_id: int) -> Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        return jsonify({"ok": False, "error": "Balanço não encontrado."}), 404
    codigo = request.args.get("codigo", "").strip()
    if not codigo:
        return jsonify({"ok": False, "error": "Código vazio."}), 400
    with closing(get_conn()) as conn:
        item = conn.execute("""
            SELECT id, codigo, codigo_barras, descricao
            FROM stock_items
            WHERE ativo = 1 AND (codigo = ? OR codigo_barras = ?)
            LIMIT 1
            """, (codigo, codigo)).fetchone()
        if item is None:
            return jsonify({"ok": True, "found": False, "exists": False})
        existente = conn.execute("""
            SELECT quantidade_contada
            FROM balance_count_items
            WHERE balance_count_id = ? AND stock_item_id = ?
            """, (balance_id, item["id"])).fetchone()
    return jsonify({
        "ok": True,
        "found": True,
        "exists": existente is not None,
        "codigo": item["codigo"],
        "descricao": item["descricao"],
        "quantidade_contada": float(existente["quantidade_contada"] or 0) if existente else 0,
    })
@app.post("/balanco/<int:balance_id>/item/<int:item_id>/remover")
@login_required
@module_required("balanco")
def remover_item_balanco(balance_id: int, item_id: int) -> Response:
    balanco = _balance_row(balance_id)
    if balanco is None or balanco["status"] != "ABERTO":
        flash("Não é possível remover item deste balanço.", "error")
        return redirect(url_for("balancos"))
    if not (user_is_admin(g.user) or int(balanco["criado_por"] or 0) == int(g.user["id"])):
        return forbidden_redirect("Somente o admin ou o usuário que criou este balanço pode remover a contagem.")
    with closing(get_conn()) as conn:
        conn.execute("DELETE FROM balance_count_items WHERE id = ? AND balance_count_id = ?", (item_id, balance_id))
        conn.commit()
    registrar_auditoria("remover_item_balanco", "balance_counts", str(balance_id), {"item_id": item_id})
    flash("Item removido da contagem.", "success")
    return redirect(url_for("detalhe_balanco", balance_id=balance_id))


@app.post("/balanco/<int:balance_id>/confirmar")
@login_required
@module_required("balanco")
def confirmar_balanco(balance_id: int) -> Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        flash("Balanço não encontrado.", "error")
        return redirect(url_for("balancos"))
    if balanco["status"] != "ABERTO":
        flash("Este balanço já foi confirmado.", "error")
        return redirect(url_for("detalhe_balanco", balance_id=balance_id))
    if not can_edit_stock_registration(g.user):
        return forbidden_redirect("Somente admin pode confirmar balanço e atualizar o estoque.")
    items = _balance_items(balance_id)
    if not items:
        flash("Inclua ao menos um item antes de confirmar.", "error")
        return redirect(url_for("detalhe_balanco", balance_id=balance_id))
    agora = agora_iso()
    with closing(get_conn()) as conn:
        for item in items:
            contada = float(item["quantidade_contada"] or 0)
            atual = conn.execute("SELECT quantidade_atual FROM stock_items WHERE id = ?", (item["stock_item_id"],)).fetchone()
            atual_qtd = float(atual["quantidade_atual"] or 0) if atual else float(item["quantidade_sistema"] or 0)
            delta_real = contada - atual_qtd
            conn.execute("UPDATE stock_items SET quantidade_atual = ?, atualizado_em = ? WHERE id = ?", (contada, agora, item["stock_item_id"]))
            conn.execute(
                """
                INSERT INTO stock_movements
                (stock_item_id, tipo, quantidade, observacao, referencia_tipo, referencia_id, criado_por, criado_em)
                VALUES (?, 'BALANCO_ESTOQUE', ?, ?, 'BALANCO', ?, ?, ?)
                """,
                (item["stock_item_id"], delta_real, f"Atualização por balanço #{balance_id}: sistema {fmt_num(atual_qtd)} → contado {fmt_num(contada)}", balance_id, g.user["id"], agora),
            )
        conn.execute("UPDATE balance_counts SET status = 'CONFIRMADO', confirmado_por = ?, confirmado_em = ? WHERE id = ?", (g.user["id"], agora, balance_id))
        conn.commit()
    registrar_auditoria("confirmar_balanco", "balance_counts", str(balance_id), {"total_itens": len(items)})
    flash("Balanço confirmado e estoque atualizado com base na contagem.", "success")
    return redirect(url_for("detalhe_balanco", balance_id=balance_id))


@app.get("/balanco/<int:balance_id>/remover")
@login_required
@module_required("balanco")
def confirmar_remover_balanco(balance_id: int) -> str | Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        flash("Contagem não encontrada.", "error")
        return redirect(url_for("balancos"))
    if not (user_is_admin(g.user) or int(balanco["criado_por"] or 0) == int(g.user["id"])):
        return forbidden_redirect("Somente o admin ou o usuário que criou este balanço pode remover a contagem.")
    with closing(get_conn()) as conn:
        resumo = conn.execute(
            """
            SELECT COUNT(*) AS total_itens,
                   COALESCE(SUM(CASE WHEN ABS(delta) > 0.000001 THEN 1 ELSE 0 END), 0) AS divergencias
            FROM balance_count_items
            WHERE balance_count_id = ?
            """,
            (balance_id,),
        ).fetchone()
    return render_template("balanco_remover.html", balanco=balanco, resumo=resumo)


@app.post("/balanco/<int:balance_id>/remover")
@login_required
@module_required("balanco")
def remover_balanco(balance_id: int) -> Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        flash("Contagem não encontrada.", "error")
        return redirect(url_for("balancos"))

    if not (user_is_admin(g.user) or int(balanco["criado_por"] or 0) == int(g.user["id"])):
        return forbidden_redirect("Somente o admin ou o usuário que criou este balanço pode remover a contagem.")

    senha_admin = request.form.get("senha_admin", "")
    if not check_password_hash(g.user["password_hash"], senha_admin):
        flash("Senha incorreta. A contagem não foi removida.", "error")
        return redirect(url_for("balancos"))

    titulo = str(balanco["titulo"] or "")
    status = str(balanco["status"] or "")
    with closing(get_conn()) as conn:
        resumo = conn.execute(
            """
            SELECT COUNT(*) AS total_itens,
                   COALESCE(SUM(CASE WHEN ABS(delta) > 0.000001 THEN 1 ELSE 0 END), 0) AS divergencias
            FROM balance_count_items
            WHERE balance_count_id = ?
            """,
            (balance_id,),
        ).fetchone()
        conn.execute("DELETE FROM balance_count_items WHERE balance_count_id = ?", (balance_id,))
        conn.execute("DELETE FROM balance_counts WHERE id = ?", (balance_id,))
        conn.commit()

    registrar_auditoria(
        "remover_balanco_estoque",
        "balance_counts",
        str(balance_id),
        {
            "titulo": titulo,
            "status": status,
            "total_itens": int(resumo["total_itens"] or 0) if resumo else 0,
            "divergencias": int(resumo["divergencias"] or 0) if resumo else 0,
            "observacao": "Remove apenas o registro da contagem. Se o balanço já estava confirmado, o estoque não é desfeito automaticamente.",
        },
    )
    flash("Contagem removida com sucesso.", "success")
    return redirect(url_for("balancos"))


@app.get("/balanco/<int:balance_id>/exportar.xlsx")
@login_required
@module_required("balanco")
def exportar_balanco_excel(balance_id: int) -> Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        flash("Balanço não encontrado.", "error")
        return redirect(url_for("balancos"))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash("Para exportar em Excel, instale openpyxl.", "error")
        return redirect(url_for("detalhe_balanco", balance_id=balance_id))
    items = _balance_items(balance_id)
    resumo = _balance_summary(items)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balanço"
    ws.append(["Balanço de estoque", f"#{balance_id}"])
    ws.append(["Título", balanco["titulo"]])
    ws.append(["Status", balanco["status"]])
    ws.append(["Criado em", balanco["criado_em"]])
    ws.append(["Total itens", resumo["total"]])
    ws.append(["Divergências", resumo["divergentes"]])
    ws.append([])
    headers = ["Linha", "Código", "Descrição", "Qtd sistema", "Qtd contada", "Diferença", "Custo unit.", "Valor diferença"]
    ws.append(headers)
    for cell in ws[8]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F7E33")
    for i in items:
        delta = float(i["delta"] or 0)
        custo = float(i["custo_unitario"] or 0)
        ws.append([i["linha_erp"] or "-", i["codigo"], i["descricao"], i["quantidade_sistema"], i["quantidade_contada"], delta, custo, delta*custo])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 10), 45)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    registrar_auditoria("exportar_balanco_excel", "balance_counts", str(balance_id), {})
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"balanco_{balance_id}.xlsx")


@app.get("/balanco/<int:balance_id>/exportar.pdf")
@login_required
@module_required("balanco")
def exportar_balanco_pdf(balance_id: int) -> Response:
    balanco = _balance_row(balance_id)
    if balanco is None:
        flash("Balanço não encontrado.", "error")
        return redirect(url_for("balancos"))
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    items = _balance_items(balance_id)
    resumo = _balance_summary(items)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=.8*cm, leftMargin=.8*cm, topMargin=.8*cm, bottomMargin=.8*cm)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"Balanço de estoque #{balance_id}", styles["Title"]), Paragraph(f"{balanco['titulo']} • Status: {balanco['status']} • Criado em: {balanco['criado_em']}", styles["Normal"]), Spacer(1,.2*cm)]
    cards = [["Indicador", "Valor"], ["Itens contados", str(resumo["total"])], ["Divergências", str(resumo["divergentes"])], ["Sobras", str(resumo["sobras"])], ["Faltas", str(resumo["faltas"])], ["Valor diferença", fmt_money(resumo["valor_delta"])]]
    t = Table(cards, repeatRows=1)
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#3f7e33")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.25,colors.grey),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")]))
    story.append(t); story.append(Spacer(1,.25*cm))
    rows = [["Linha", "Código", "Descrição", "Sistema", "Contado", "Dif."]]
    rows.extend([[(i["linha_erp"] or "-")[:24], i["codigo"], (i["descricao"] or "")[:75], fmt_num(i["quantidade_sistema"]), fmt_num(i["quantidade_contada"]), fmt_num(i["delta"])] for i in items[:160]])
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#3f7e33")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.2,colors.grey),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7)]))
    story.append(table)
    doc.build(story); buf.seek(0)
    registrar_auditoria("exportar_balanco_pdf", "balance_counts", str(balance_id), {})
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=f"balanco_{balance_id}.pdf")




# =========================
# Comunicação: chat, tarefas e pedidos agendados
# =========================

def _upload_chat_dir() -> str:
    folder = os.path.join(BASE_DIR, "static", "uploads", "chat_temp")
    os.makedirs(folder, exist_ok=True)
    return folder


def _is_group_member(conn: sqlite3.Connection, group_id: int, user_id: int) -> bool:
    if user_is_admin(g.user):
        return True
    row = conn.execute("SELECT 1 FROM chat_group_members WHERE group_id = ? AND user_id = ?", (group_id, user_id)).fetchone()
    return row is not None


@app.get("/comunicacao")
@login_required
@module_required("comunicacao")
def comunicacao() -> str:
    group_id = request.args.get("group_id", type=int)
    with closing(get_conn()) as conn:
        usuarios = conn.execute("SELECT id, nome, role FROM users WHERE ativo = 1 ORDER BY nome").fetchall()
        lojas = conn.execute("SELECT id, nome FROM stores WHERE ativo = 1 ORDER BY nome").fetchall()
        if user_is_admin(g.user):
            grupos = conn.execute("""
                SELECT g.*, COUNT(m.user_id) AS total_membros
                FROM chat_groups g
                LEFT JOIN chat_group_members m ON m.group_id = g.id
                WHERE g.ativo = 1
                GROUP BY g.id
                ORDER BY g.criado_em DESC
            """).fetchall()
        else:
            grupos = conn.execute("""
                SELECT g.*, COUNT(m2.user_id) AS total_membros
                FROM chat_groups g
                JOIN chat_group_members m ON m.group_id = g.id AND m.user_id = ?
                LEFT JOIN chat_group_members m2 ON m2.group_id = g.id
                WHERE g.ativo = 1
                GROUP BY g.id
                ORDER BY g.criado_em DESC
            """, (g.user["id"],)).fetchall()
        grupo_atual = None
        mensagens = []
        if group_id:
            if _is_group_member(conn, group_id, g.user["id"]):
                grupo_atual = conn.execute("SELECT * FROM chat_groups WHERE id = ? AND ativo = 1", (group_id,)).fetchone()
                mensagens = conn.execute("""
                    SELECT msg.*, u.nome AS usuario_nome
                    FROM chat_messages msg
                    LEFT JOIN users u ON u.id = msg.user_id
                    WHERE msg.group_id = ?
                    ORDER BY msg.criado_em ASC, msg.id ASC
                """, (group_id,)).fetchall()
            else:
                flash("Você não participa desse grupo.", "error")
        tarefas = conn.execute("""
            SELECT t.*, u.nome AS responsavel_nome
            FROM team_tasks t
            LEFT JOIN users u ON u.id = t.responsavel_id
            ORDER BY CASE t.status WHEN 'ABERTA' THEN 0 ELSE 1 END, t.prazo IS NULL, t.prazo, t.id DESC
            LIMIT 80
        """).fetchall()
        pedidos_agendados = conn.execute("""
            SELECT o.*, s.nome AS loja_nome
            FROM scheduled_orders o
            LEFT JOIN stores s ON s.id = o.loja_id
            ORDER BY CASE o.status WHEN 'AGENDADO' THEN 0 ELSE 1 END, o.agendado_para IS NULL, o.agendado_para, o.id DESC
            LIMIT 80
        """).fetchall()
    return render_template(
        "comunicacao.html",
        title="Chat/Tarefas",
        usuarios=usuarios,
        lojas=lojas,
        grupos=grupos,
        grupo_atual=grupo_atual,
        mensagens=mensagens,
        tarefas=tarefas,
        pedidos_agendados=pedidos_agendados,
        pode_criar_pedidos=user_has_access(g.user, "pedidos"),
    )



@app.post("/comunicacao/direto/abrir")
@login_required
@module_required("comunicacao")
def abrir_chat_direto() -> Response:
    usuario_id = request.form.get("usuario_id", type=int)
    if not usuario_id or usuario_id == g.user["id"]:
        flash("Selecione um usuário válido.", "error")
        return redirect(url_for("comunicacao"))
    a, b = sorted([int(g.user["id"]), int(usuario_id)])
    direto_key = f"{a}:{b}"
    with closing(get_conn()) as conn:
        destino = conn.execute("SELECT id, nome FROM users WHERE id = ? AND ativo = 1", (usuario_id,)).fetchone()
        if destino is None:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("comunicacao"))
        grupo = conn.execute("SELECT id FROM chat_groups WHERE direto_key = ? AND ativo = 1", (direto_key,)).fetchone()
        if grupo is None:
            nome = f"{g.user['nome']} ↔ {destino['nome']}"
            cur = conn.execute("INSERT INTO chat_groups (nome, criado_por, criado_em, tipo_chat, direto_key) VALUES (?, ?, ?, 'direto', ?)", (nome, g.user["id"], agora_iso(), direto_key))
            group_id = cur.lastrowid
            conn.executemany("INSERT OR IGNORE INTO chat_group_members (group_id, user_id, criado_em) VALUES (?, ?, ?)", [(group_id, g.user["id"], agora_iso()), (group_id, usuario_id, agora_iso())])
            conn.commit()
        else:
            group_id = grupo["id"]
    return redirect(url_for("comunicacao", group_id=group_id))


@app.post("/comunicacao/grupos/criar")
@login_required
@module_required("comunicacao")
def criar_grupo_chat() -> Response:
    if not user_is_admin(g.user):
        return forbidden_redirect("Somente o admin pode criar grupos.")
    nome = request.form.get("nome", "").strip()
    participantes = [int(x) for x in request.form.getlist("participantes") if str(x).isdigit()]
    if not nome:
        flash("Informe o nome do grupo.", "error")
        return redirect(url_for("comunicacao"))
    if g.user["id"] not in participantes:
        participantes.append(g.user["id"])
    with closing(get_conn()) as conn:
        cur = conn.execute("INSERT INTO chat_groups (nome, criado_por, criado_em) VALUES (?, ?, ?)", (nome, g.user["id"], agora_iso()))
        gid = cur.lastrowid
        conn.executemany("INSERT OR IGNORE INTO chat_group_members (group_id, user_id, criado_em) VALUES (?, ?, ?)", [(gid, uid, agora_iso()) for uid in participantes])
        conn.commit()
    registrar_auditoria("criar_grupo_chat", "chat_groups", str(gid), {"nome": nome, "participantes": participantes})
    flash("Grupo criado.", "success")
    return redirect(url_for("comunicacao", group_id=gid))




@app.post("/comunicacao/grupos/<int:group_id>/excluir")
@login_required
@module_required("comunicacao")
def excluir_grupo_chat(group_id: int) -> Response:
    if not user_is_admin(g.user):
        return forbidden_redirect("Somente o admin pode excluir chats.")
    with closing(get_conn()) as conn:
        grupo = conn.execute("SELECT * FROM chat_groups WHERE id = ? AND ativo = 1", (group_id,)).fetchone()
        if grupo is None:
            flash("Chat não encontrado ou já excluído.", "error")
            return redirect(url_for("comunicacao"))
        arquivos = conn.execute("SELECT id, arquivo_path FROM chat_messages WHERE group_id = ? AND arquivo_path IS NOT NULL AND arquivo_status = 'pendente'", (group_id,)).fetchall()
        for arq in arquivos:
            if arq["arquivo_path"] and os.path.exists(arq["arquivo_path"]):
                try:
                    os.remove(arq["arquivo_path"])
                except OSError:
                    pass
        conn.execute("UPDATE chat_groups SET ativo = 0 WHERE id = ?", (group_id,))
        conn.execute("UPDATE chat_messages SET arquivo_status = CASE WHEN arquivo_path IS NOT NULL AND arquivo_status = 'pendente' THEN 'apagado_chat_excluido' ELSE arquivo_status END WHERE group_id = ?", (group_id,))
        conn.commit()
    registrar_auditoria("excluir_grupo_chat", "chat_groups", str(group_id), {"nome": grupo["nome"], "arquivos_apagados": len(arquivos)})
    flash("Chat excluído pelo admin. Arquivos temporários pendentes foram apagados.", "success")
    return redirect(url_for("comunicacao"))


@app.post("/comunicacao/grupos/<int:group_id>/mensagem")
@login_required
@module_required("comunicacao")
def enviar_mensagem_chat(group_id: int) -> Response:
    mensagem = request.form.get("mensagem", "").strip()
    arquivo = request.files.get("arquivo")
    arquivo_nome = arquivo_path = None
    status = "sem_arquivo"
    if arquivo and arquivo.filename:
        safe = secure_filename(arquivo.filename)
        arquivo_nome = safe or "arquivo"
        unique = f"{uuid.uuid4().hex}_{arquivo_nome}"
        full = os.path.join(_upload_chat_dir(), unique)
        arquivo.save(full)
        arquivo_path = full
        status = "pendente"
    if not mensagem and not arquivo_nome:
        flash("Digite uma mensagem ou envie um arquivo.", "error")
        return redirect(url_for("comunicacao", group_id=group_id))
    with closing(get_conn()) as conn:
        if not _is_group_member(conn, group_id, g.user["id"]):
            flash("Você não participa desse grupo.", "error")
            return redirect(url_for("comunicacao"))
        cur = conn.execute("""
            INSERT INTO chat_messages (group_id, user_id, mensagem, arquivo_nome, arquivo_path, arquivo_status, expira_em, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now', '+7 days'), ?)
        """, (group_id, g.user["id"], mensagem, arquivo_nome, arquivo_path, status, agora_iso()))
        mid = cur.lastrowid
        conn.commit()
    registrar_auditoria("enviar_mensagem_chat", "chat_messages", str(mid), {"group_id": group_id, "tem_arquivo": bool(arquivo_nome)})
    return redirect(url_for("comunicacao", group_id=group_id))


@app.get("/comunicacao/arquivo/<int:message_id>")
@login_required
@module_required("comunicacao")
def baixar_arquivo_chat(message_id: int) -> Response:
    with closing(get_conn()) as conn:
        msg = conn.execute("SELECT * FROM chat_messages WHERE id = ?", (message_id,)).fetchone()
        if msg is None or not _is_group_member(conn, int(msg["group_id"]), g.user["id"]):
            flash("Arquivo não encontrado.", "error")
            return redirect(url_for("comunicacao"))
    if not msg["arquivo_path"] or msg["arquivo_status"] != "pendente" or not os.path.exists(msg["arquivo_path"]):
        flash("Arquivo expirado ou já confirmado.", "error")
        return redirect(url_for("comunicacao", group_id=msg["group_id"]))
    return send_file(msg["arquivo_path"], as_attachment=True, download_name=msg["arquivo_nome"])


@app.post("/comunicacao/arquivo/<int:message_id>/confirmar")
@login_required
@module_required("comunicacao")
def confirmar_arquivo_chat(message_id: int) -> Response:
    with closing(get_conn()) as conn:
        msg = conn.execute("SELECT * FROM chat_messages WHERE id = ?", (message_id,)).fetchone()
        if msg is None or not _is_group_member(conn, int(msg["group_id"]), g.user["id"]):
            flash("Arquivo não encontrado.", "error")
            return redirect(url_for("comunicacao"))
        if msg["arquivo_path"] and os.path.exists(msg["arquivo_path"]):
            try:
                os.remove(msg["arquivo_path"])
            except OSError:
                pass
        conn.execute("UPDATE chat_messages SET arquivo_status = 'recebido_apagado', recebido_por = ?, recebido_em = ? WHERE id = ?", (g.user["id"], agora_iso(), message_id))
        conn.commit()
    registrar_auditoria("confirmar_recebimento_arquivo_chat", "chat_messages", str(message_id), {"group_id": msg["group_id"]})
    flash("Recebimento confirmado. Arquivo temporário apagado.", "success")
    return redirect(url_for("comunicacao", group_id=msg["group_id"]))


@app.post("/comunicacao/tarefas/criar")
@login_required
@module_required("comunicacao")
def criar_tarefa() -> Response:
    titulo = request.form.get("titulo", "").strip()
    descricao = request.form.get("descricao", "").strip()
    responsavel_id = request.form.get("responsavel_id", type=int)
    prazo = request.form.get("prazo", "").strip()
    if not titulo:
        flash("Informe o título da tarefa.", "error")
        return redirect(url_for("comunicacao"))
    redirect_group_id = request.form.get("redirect_group_id", type=int)
    with closing(get_conn()) as conn:
        cur = conn.execute("INSERT INTO team_tasks (titulo, descricao, responsavel_id, prazo, criado_por, criado_em) VALUES (?, ?, ?, ?, ?, ?)", (titulo, descricao, responsavel_id, prazo, g.user["id"], agora_iso()))
        tid = cur.lastrowid
        if redirect_group_id and _is_group_member(conn, redirect_group_id, g.user["id"]):
            aviso = f"📌 Tarefa criada: {titulo}"
            if descricao:
                aviso += f"\n{descricao}"
            conn.execute("INSERT INTO chat_messages (group_id, user_id, mensagem, arquivo_status, criado_em) VALUES (?, ?, ?, 'sem_arquivo', ?)", (redirect_group_id, g.user["id"], aviso, agora_iso()))
        conn.commit()
    registrar_auditoria("criar_tarefa", "team_tasks", str(tid), {"titulo": titulo, "responsavel_id": responsavel_id})
    flash("Tarefa criada.", "success")
    return redirect(url_for("comunicacao", group_id=redirect_group_id) if redirect_group_id else url_for("comunicacao"))


@app.post("/comunicacao/tarefas/<int:task_id>/concluir")
@login_required
@module_required("comunicacao")
def concluir_tarefa(task_id: int) -> Response:
    with closing(get_conn()) as conn:
        conn.execute("UPDATE team_tasks SET status = 'CONCLUIDA', finalizado_em = ? WHERE id = ?", (agora_iso(), task_id))
        conn.commit()
    registrar_auditoria("concluir_tarefa", "team_tasks", str(task_id), {})
    flash("Tarefa concluída.", "success")
    return redirect(url_for("comunicacao"))


@app.post("/comunicacao/pedidos-agendados/criar")
@login_required
@module_required("comunicacao")
def criar_pedido_agendado() -> Response:
    titulo = request.form.get("titulo", "").strip()
    itens_texto = request.form.get("itens_texto", "").strip()
    loja_id = request.form.get("loja_id", type=int)
    agendado_para = request.form.get("agendado_para", "").strip()
    if not titulo or not itens_texto:
        flash("Informe título e itens do pedido.", "error")
        return redirect(url_for("comunicacao"))
    redirect_group_id = request.form.get("redirect_group_id", type=int)
    with closing(get_conn()) as conn:
        cur = conn.execute("INSERT INTO scheduled_orders (loja_id, titulo, itens_texto, agendado_para, criado_por, criado_em) VALUES (?, ?, ?, ?, ?, ?)", (loja_id, titulo, itens_texto, agendado_para, g.user["id"], agora_iso()))
        oid = cur.lastrowid
        if redirect_group_id and _is_group_member(conn, redirect_group_id, g.user["id"]):
            aviso = f"🧾 Pedido agendado: {titulo}"
            if itens_texto:
                aviso += f"\n{itens_texto}"
            conn.execute("INSERT INTO chat_messages (group_id, user_id, mensagem, arquivo_status, criado_em) VALUES (?, ?, ?, 'sem_arquivo', ?)", (redirect_group_id, g.user["id"], aviso, agora_iso()))
        conn.commit()
    registrar_auditoria("criar_pedido_agendado", "scheduled_orders", str(oid), {"titulo": titulo, "loja_id": loja_id})
    flash("Pedido agendado criado.", "success")
    return redirect(url_for("comunicacao", group_id=redirect_group_id) if redirect_group_id else url_for("comunicacao"))


@app.post("/comunicacao/pedidos-agendados/<int:order_id>/enviar")
@login_required
@module_required("pedidos")
def enviar_pedido_agendado(order_id: int) -> Response:
    with closing(get_conn()) as conn:
        pedido = conn.execute("SELECT * FROM scheduled_orders WHERE id = ?", (order_id,)).fetchone()
        if pedido is None:
            flash("Pedido agendado não encontrado.", "error")
            return redirect(url_for("comunicacao"))
        conn.execute("UPDATE scheduled_orders SET status = 'ENVIADO_AO_SEPARADOR', enviado_por = ?, enviado_em = ? WHERE id = ?", (g.user["id"], agora_iso(), order_id))
        conn.commit()
    registrar_auditoria("enviar_pedido_agendado_ao_separador", "scheduled_orders", str(order_id), {"titulo": pedido["titulo"], "loja_id": pedido["loja_id"]})
    flash("Pedido marcado como enviado ao separador. Use Criar pedidos para montar a separação oficial quando necessário.", "success")
    return redirect(url_for("comunicacao"))

@app.get("/favicon.ico")
def favicon() -> Response:
    return send_from_directory(os.path.join(BASE_DIR, "static", "img"), "favicon.ico", mimetype="image/vnd.microsoft.icon")

@app.get("/health")
def health() -> Response:
    try:
        with closing(get_conn()) as conn:
            conn.execute("SELECT 1").fetchone()
        return jsonify({"status": "ok", "db": DB_PATH})
    except Exception as exc:
        return jsonify({"status": "error", "error": str(exc), "db": DB_PATH}), 500




def _mcp_encontrar_usuario_por_texto(conn: sqlite3.Connection, texto: str) -> sqlite3.Row | None:
    termo = str(texto or "").strip().casefold()
    if not termo:
        return None
    usuarios = conn.execute("""
        SELECT u.id, u.nome, u.username, u.store_id, st.nome AS loja_nome
        FROM users u
        LEFT JOIN stores st ON st.id = u.store_id
        WHERE u.ativo = 1
        ORDER BY u.nome
    """).fetchall()
    for u in usuarios:
        nome = str(u["nome"] or "").casefold()
        username = str(u["username"] or "").casefold()
        if termo in {nome, username} or termo in nome or termo in username:
            return u
    return None


def _mcp_encontrar_loja_por_texto(conn: sqlite3.Connection, texto: str) -> sqlite3.Row | None:
    termo = str(texto or "").strip().casefold()
    if not termo:
        return None
    lojas = conn.execute("SELECT id, nome FROM stores WHERE ativo = 1 ORDER BY nome").fetchall()
    for loja in lojas:
        nome = str(loja["nome"] or "").casefold()
        if termo == nome or termo in nome:
            return loja
    if termo.isdigit():
        return conn.execute("SELECT id, nome FROM stores WHERE id = ?", (int(termo),)).fetchone()
    return None


def _mcp_obter_grupo_direto(conn: sqlite3.Connection, destino_user_id: int) -> int:
    atual = int(g.user["id"])
    destino = conn.execute("SELECT nome, username FROM users WHERE id = ?", (destino_user_id,)).fetchone()
    nome_destino = destino["nome"] if destino else str(destino_user_id)
    nome_atual = g.user["nome"] or g.user["username"]
    nome = f"Ocorrências - {nome_atual} / {nome_destino}"
    grupo = conn.execute("SELECT id FROM chat_groups WHERE nome = ? AND ativo = 1 LIMIT 1", (nome,)).fetchone()
    if grupo:
        gid = int(grupo["id"])
    else:
        cur = conn.execute("INSERT INTO chat_groups (nome, criado_por, criado_em) VALUES (?, ?, ?)", (nome, atual, agora_iso()))
        gid = int(cur.lastrowid)
    for uid in {atual, int(destino_user_id)}:
        conn.execute("INSERT OR IGNORE INTO chat_group_members (group_id, user_id, criado_em) VALUES (?, ?, ?)", (gid, uid, agora_iso()))
    return gid


def _formatar_ocorrencia_mcp(tipo: str, loja_nome: str, nota_numero: str, itens: list[dict[str, Any]], relato: str) -> str:
    labels = {
        "mercadoria_sem_nota": "Mercadoria chegou sem nota",
        "nota_sem_mercadoria": "Nota chegou, mas mercadoria veio faltando/passando",
        "outro": "Outro problema informado pela loja",
    }
    linhas = ["Ocorrência registrada via Assistente MCP/IA", f"Tipo: {labels.get(tipo, tipo)}"]
    if loja_nome:
        linhas.append(f"Loja: {loja_nome}")
    if nota_numero:
        linhas.append(f"Nota: {nota_numero}")
    if relato:
        linhas.append(f"Relato: {relato}")
    if itens:
        linhas.append("Itens:")
        for idx, item in enumerate(itens, 1):
            cod = item.get("codigo") or item.get("codigo_barras") or "-"
            desc = item.get("descricao") or "-"
            qtd = item.get("quantidade") or item.get("diferenca") or "-"
            sinal = item.get("sinal") or ""
            linhas.append(f"{idx}. Código/Barras: {cod} | Descrição: {desc} | Qtd: {sinal}{qtd}")
    return "\n".join(linhas)


@app.post("/api/mcp/anexar-temp")
@login_required
def api_mcp_anexar_temp() -> Response:
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"error": "Nenhum arquivo recebido."}), 400
    original = secure_filename(arquivo.filename) or "arquivo"
    ext = os.path.splitext(original)[1].lower()
    allowed = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".pdf", ".txt", ".xlsx", ".xls", ".csv", ".doc", ".docx"}
    if ext not in allowed:
        return jsonify({"error": "Tipo de arquivo não permitido para anexo temporário."}), 400
    unique = f"mcp_{g.user['id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    full = os.path.join(_upload_chat_dir(), unique)
    arquivo.save(full)
    url = url_for('static', filename='uploads/chat_temp/' + unique)
    registrar_auditoria("mcp_anexar_arquivo_temp", "chat_temp", unique, {"nome_original": original})
    return jsonify({"ok": True, "nome": original, "url": url, "path": unique})


@app.get("/api/mcp/usuarios")
@login_required
def api_mcp_usuarios() -> Response:
    with closing(get_conn()) as conn:
        rows = conn.execute("""
            SELECT u.id, u.nome, u.username, u.role, u.store_id, st.nome AS loja_nome
            FROM users u
            LEFT JOIN stores st ON st.id = u.store_id
            WHERE u.ativo = 1
            ORDER BY u.nome
        """).fetchall()
    return jsonify({"usuarios": [dict(r) for r in rows]})


@app.post("/api/mcp/registrar-ocorrencia")
@login_required
@module_required("comunicacao")
def api_mcp_registrar_ocorrencia() -> Response:
    data = request.get_json(silent=True) or {}
    tipo = str(data.get("tipo") or "").strip()
    destino_txt = str(data.get("destino") or "").strip()
    loja_txt = str(data.get("loja") or "").strip()
    nota_numero = str(data.get("nota_numero") or "").strip()
    relato = str(data.get("relato") or "").strip()
    itens = data.get("itens") if isinstance(data.get("itens"), list) else []
    anexos = data.get("anexos") if isinstance(data.get("anexos"), list) else []
    if anexos:
        linhas_anexo = []
        for anexo in anexos[:8]:
            if isinstance(anexo, dict):
                nome = str(anexo.get("nome") or "arquivo").strip()
                url = str(anexo.get("url") or "").strip()
                linhas_anexo.append(f"- {nome}: {url}" if url else f"- {nome}")
        if linhas_anexo:
            relato = (relato + "\n\nAnexos temporários enviados pelo MCP:\n" + "\n".join(linhas_anexo)).strip()

    if tipo not in {"mercadoria_sem_nota", "nota_sem_mercadoria", "outro"}:
        return jsonify({"error": "Tipo de problema inválido."}), 400
    if not destino_txt:
        return jsonify({"error": "Informe o usuário de destino."}), 400
    if tipo == "nota_sem_mercadoria" and not nota_numero:
        return jsonify({"error": "Informe o número da nota."}), 400
    if tipo in {"mercadoria_sem_nota", "nota_sem_mercadoria"} and not itens:
        return jsonify({"error": "Informe pelo menos um item."}), 400

    with closing(get_conn()) as conn:
        destino = _mcp_encontrar_usuario_por_texto(conn, destino_txt)
        if destino is None:
            return jsonify({"error": f"Não encontrei o usuário de destino: {destino_txt}."}), 404
        loja = _mcp_encontrar_loja_por_texto(conn, loja_txt) if loja_txt else None
        if loja is None and g.user["store_id"]:
            loja = conn.execute("SELECT id, nome FROM stores WHERE id = ?", (int(g.user["store_id"]),)).fetchone()
        gid = _mcp_obter_grupo_direto(conn, int(destino["id"]))
        mensagem = _formatar_ocorrencia_mcp(tipo, loja["nome"] if loja else loja_txt, nota_numero, itens, relato)
        cur = conn.execute("""
            INSERT INTO store_issue_reports (tipo, destino_user_id, loja_id, nota_numero, itens_json, relato, chat_group_id, criado_por, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (tipo, int(destino["id"]), int(loja["id"]) if loja else None, nota_numero, json.dumps(itens, ensure_ascii=False), relato, gid, int(g.user["id"]), agora_iso()))
        issue_id = int(cur.lastrowid)
        conn.execute("""
            INSERT INTO chat_messages (group_id, user_id, mensagem, arquivo_status, criado_em)
            VALUES (?, ?, ?, 'sem_arquivo', ?)
        """, (gid, int(g.user["id"]), mensagem, agora_iso()))
        conn.commit()

    registrar_auditoria("mcp_registrar_ocorrencia_loja", "store_issue_reports", str(issue_id), {
        "tipo": tipo, "destino": destino_txt, "loja": (loja["nome"] if loja else loja_txt), "nota_numero": nota_numero, "itens": itens
    })
    return jsonify({
        "ok": True,
        "mensagem": "Ocorrência enviada no chat para " + (destino["nome"] or destino["username"]) + ".",
        "redirect_url": url_for("comunicacao", group_id=gid),
        "issue_id": issue_id,
    })
if __name__ == "__main__":
    ensure_default_data()
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
